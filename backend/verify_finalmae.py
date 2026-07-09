import sys, pandas as pd
sys.path.insert(0, '/app')
from services.algotest_job import execute_algotest_job
from services.optimizer import excel_builder as eb
import openpyxl, io

def run(legs, tag):
    req={"index":"NIFTY","from_date":"2024-01-01","to_date":"2024-06-30",
     "strategy_type":"positional","underlying":"cash","expiry_type":"WEEKLY",
     "entry_dte":1,"exit_dte":0,"slippage_pct":0,"charges_enabled":False,
     "square_off_mode":"partial","rollover_toggle":True,"no_cache":True,"legs":legs}
    res = execute_algotest_job(req)
    trades = res.get("trades") or []
    df = pd.DataFrame(trades)
    xb = eb.build_combo_xlsx(df, res.get("summary") or {}, combo_label=tag,
                             from_date=req["from_date"], to_date=req["to_date"])
    wb = openpyxl.load_workbook(io.BytesIO(xb))
    ws = wb["Trade Sheet"]
    hdr = [c.value for c in ws[1]]
    idx = {h:i for i,h in enumerate(hdr) if h}
    print(f"=== {tag} ===")
    mism = 0; checked = 0
    for r in range(2, min(ws.max_row+1, 40)):
        nm1 = ws.cell(row=r, column=idx.get('Net MAE 1',0)+1).value
        nm2 = ws.cell(row=r, column=idx.get('Net MAE 2',0)+1).value
        fm  = ws.cell(row=r, column=idx.get('Final MAE',0)+1).value
        pct = ws.cell(row=r, column=idx.get('% P&L',0)+1).value
        if nm1 in (None,'') or nm2 in (None,'') or fm in (None,''): continue
        checked += 1
        # excel stores % P&L as fraction (0.00% fmt) — value is the raw pct? check both
        # In sheet, % P&L cell.value is stored as the number; MAE cols are TRUE_PCT too.
        exp_multi = round(min(nm1, nm2, pct), 4)
        exp_single = round(min(nm1, nm2), 4)
        if r <= 7:
            print(f"  row {r}: nm1={nm1} nm2={nm2} pct={pct} -> Final={fm} | min(nm1,nm2,pct)={exp_multi} min(nm1,nm2)={exp_single}")
    print(f"  checked {checked} trades")

# 2-leg CE+PE sell (multi-leg -> should floor by pct)
run([
 {"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
 {"segment":"OPTIONS","option_type":"PE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
], "2LEG_CE_PE_SELL")

# 1-leg CE sell (single -> no floor)
run([
 {"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
], "1LEG_CE_SELL")
