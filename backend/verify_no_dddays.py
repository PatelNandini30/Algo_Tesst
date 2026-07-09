import sys, pandas as pd
sys.path.insert(0,'/app')
from services.algotest_job import execute_algotest_job
from services.optimizer import excel_builder as eb
import openpyxl, io
req={"index":"NIFTY","from_date":"2023-01-01","to_date":"2024-12-31",
 "strategy_type":"positional","underlying":"cash","expiry_type":"WEEKLY",
 "entry_dte":1,"exit_dte":0,"slippage_pct":0,"charges_enabled":False,
 "square_off_mode":"partial","rollover_toggle":True,"no_cache":True,
 "legs":[{"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
         {"segment":"OPTIONS","option_type":"PE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}}]}
res=execute_algotest_job(req)
df=pd.DataFrame(res.get("trades") or [])
xb=eb.build_combo_xlsx(df,res.get("summary") or {},combo_label="T",from_date=req["from_date"],to_date=req["to_date"])
wb=openpyxl.load_workbook(io.BytesIO(xb)); ws=wb["Summary"]
for r in range(1,ws.max_row+1):
    a=ws.cell(row=r,column=1).value
    if a=="Year":
        hdr=[ws.cell(row=r,column=c).value for c in range(1,18) if ws.cell(row=r,column=c).value not in (None,)]
        print("HEADER:", hdr)
        # next data row
        dat=[ws.cell(row=r+1,column=c).value for c in range(1,len(hdr)+1)]
        print("ROW   :", dat)
        break
