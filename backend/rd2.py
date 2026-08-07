import warnings; warnings.filterwarnings("ignore")
import openpyxl
wb=openpyxl.load_workbook("/app/_optfile.xlsx", data_only=True)
print("========== RULES ==========")
ws=wb["Rules"]
for row in ws.iter_rows(values_only=True):
    cells=[str(c) for c in row if c not in (None,"")]
    if cells: print(" | ".join(cells))
print("\n========== SUMMARY ==========")
ws=wb["Summary"]
for row in ws.iter_rows(values_only=True):
    cells=[str(c) for c in row if c not in (None,"")]
    if cells: print(" | ".join(cells))
