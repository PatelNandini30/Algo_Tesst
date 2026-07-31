"""MIN-of-Final-MAE / MIN-of-Actual-Live-DD grids under each WOW & MOM block.

Four grids per sheet: the weekly pair under the WOW block (Year x W1..Wn) and
the monthly pair under the MOM block (Row Labels x Jan..Dec + Grand Total).
Values are RAW percent points aggregated with MIN — never summed or averaged.

Must work for all four overlay shapes: midcap, MIDCPNIFTY, both, neither. The
field is resolved by PRESENCE of the Combined column, not by the has_midcap
flag, so a MIDCPNIFTY-only run isn't silently left with blank grids.
"""
import os
import sys
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook  # noqa: E402

from services.optimizer.wow_mom import (  # noqa: E402
    _min_grid_fields,
    _wm_from_cleaned,
    write_wow_mom_combined,
)


def _row(expiry, exit_date, pnl, mae, ldd, *, combined=False):
    pre = "Combined " if combined else ""
    r = {
        "Expiry": expiry, "Exit Date": exit_date,
        f"{pre}Net P&L %" if combined else "% P&L": pnl,
        f"{pre}%DD" if combined else "%DD": -1.0,
        f"{pre}Final MAE": mae,
        f"{pre}Actual Live DD": ldd,
    }
    if combined:
        r["Combined Net P&L %"] = pnl
        r["Combined %DD"] = -1.0
    return r


# Two trades in Jan-2024, one in Feb-2024, two in Mar-2025 — enough to prove
# MIN (not sum/avg) per cell, per-year Grand Total, and the cross-year row.
CLEANED = [
    _row("15-01-2024", "15-01-2024", 1.0, -2.0, -3.0),
    _row("22-01-2024", "22-01-2024", 2.0, -5.0, -1.0),   # same month, lower MAE
    _row("12-02-2024", "12-02-2024", 1.5, -1.0, -8.0),
    _row("11-03-2025", "11-03-2025", -1.0, -4.0, -2.0),
    _row("18-03-2025", "18-03-2025", 3.0, -0.5, -9.0),
]

CLEANED_MIDCAP = [
    _row("15-01-2024", "15-01-2024", 1.0, -2.0, -3.0, combined=True),
    _row("12-02-2024", "12-02-2024", 1.5, -1.0, -8.0, combined=True),
]


def _sheet(cleaned, has_midcap=False):
    # NOTE: no bare `assert` here — the backend image runs PYTHONOPTIMIZE=1,
    # which strips assert statements AND the call inside them.
    wb = Workbook()
    wb.remove(wb.active)
    if write_wow_mom_combined(wb, cleaned, has_midcap, "Strategy") is not True:
        raise AssertionError("write_wow_mom_combined returned falsy")
    return wb["WOW & MOM Summary"]


def _find_all(ws, title):
    """Every (row, col) where this grid title appears, top to bottom.

    The weekly and monthly copies share the same title, so callers must say
    which one they mean — weekly is written first (under the WOW block).
    """
    hits = []
    for row in ws.iter_rows():
        for c in row:
            if c.value == title:
                hits.append((c.row, c.column))
    return sorted(hits)


def _grid(ws, title, axis="mom"):
    """{row_label: {col_label: value}} for the weekly or monthly copy."""
    hits = _find_all(ws, title)
    if not hits:
        raise AssertionError(f"grid {title!r} not found")
    idx = 0 if axis == "wow" else len(hits) - 1
    r0, c0 = hits[idx]
    # Weekly: title, Month band, Year header. Monthly: title, column header.
    hdr_row = r0 + 2 if axis == "wow" else r0 + 1
    labels, out = {}, {}
    for c in range(c0 + 1, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v in (None, ""):
            break
        labels[c] = v
    r = hdr_row + 1
    while True:
        rl = ws.cell(r, c0).value
        if rl in (None, ""):
            break
        out[str(rl)] = {labels[c]: ws.cell(r, c).value for c in labels}
        if str(rl) == "Grand Total":
            break
        r += 1
    return out


class TestMonthlyMinGrids(unittest.TestCase):

    def setUp(self):
        self.ws = _sheet(CLEANED)

    def test_min_not_sum_per_month(self):
        g = _grid(self.ws, "Min of Final MAE")
        # Jan-2024 has MAE -2.0 and -5.0 → MIN is -5.0, not -7.0 and not -3.5.
        self.assertEqual(g["2024"]["Jan"], -5.0)
        self.assertEqual(g["2024"]["Feb"], -1.0)
        self.assertEqual(g["2025"]["Mar"], -4.0)

    def test_row_grand_total_is_min_across_months(self):
        g = _grid(self.ws, "Min of Final MAE")
        self.assertEqual(g["2024"]["Grand Total"], -5.0)
        self.assertEqual(g["2025"]["Grand Total"], -4.0)

    def test_grand_total_row_is_min_across_years(self):
        g = _grid(self.ws, "Min of Final MAE")
        self.assertEqual(g["Grand Total"]["Jan"], -5.0)
        self.assertEqual(g["Grand Total"]["Mar"], -4.0)
        self.assertEqual(g["Grand Total"]["Grand Total"], -5.0)

    def test_live_dd_grid_is_separate_and_unscaled(self):
        g = _grid(self.ws, "Min of Actual Live DD")
        self.assertEqual(g["2024"]["Jan"], -3.0)      # min(-3.0, -1.0)
        self.assertEqual(g["2024"]["Feb"], -8.0)
        self.assertEqual(g["2025"]["Mar"], -9.0)
        self.assertEqual(g["Grand Total"]["Grand Total"], -9.0)

    def test_months_with_no_trades_stay_blank(self):
        g = _grid(self.ws, "Min of Final MAE")
        self.assertEqual(g["2024"]["Jun"], "")
        self.assertEqual(g["2025"]["Jan"], "")


class TestWeeklyMinGrids(unittest.TestCase):

    def test_weekly_grid_uses_week_columns(self):
        ws = _sheet(CLEANED)
        # The weekly copy is the FIRST one on the sheet (under the WOW block);
        # its header row is week numbers, not month names.
        r0, c0 = _find_all(ws, "Min of Final MAE")[0]
        hdr = [ws.cell(r0 + 2, c).value for c in range(c0 + 1, c0 + 4)]
        self.assertEqual(hdr, ["W1", "W2", "W3"])

    def test_weekly_values_bucket_by_iso_week(self):
        ws = _sheet(CLEANED)
        g = _grid(ws, "Min of Final MAE", axis="wow")
        # 15-01-2024 -> ISO W3, 22-01-2024 -> W4, 12-02-2024 -> W7
        self.assertEqual(g["2024"]["W3"], -2.0)
        self.assertEqual(g["2024"]["W4"], -5.0)
        self.assertEqual(g["2024"]["W7"], -1.0)
        self.assertEqual(g["2024"]["Grand Total"], -5.0)

    def test_all_four_grids_present(self):
        ws = _sheet(CLEANED)
        titles = [c.value for row in ws.iter_rows() for c in row
                  if c.value in ("Min of Final MAE", "Min of Actual Live DD")]
        self.assertEqual(titles.count("Min of Final MAE"), 2, "weekly + monthly")
        self.assertEqual(titles.count("Min of Actual Live DD"), 2)

    def test_grids_do_not_overwrite_the_blocks(self):
        """WOW block, its grids, the MOM block, then its grids — in that order."""
        ws = _sheet(CLEANED)
        rows = [c.row for row in ws.iter_rows() for c in row
                if c.value in ("Min of Final MAE", "Min of Actual Live DD")]
        mom_hdr = None
        for row in ws.iter_rows():
            for c in row:
                if c.value == "Row Labels":
                    continue
                if c.value == "Jan" and ws.cell(c.row, c.column - 1).value == "Year":
                    mom_hdr = c.row
                    break
            if mom_hdr:
                break
        self.assertIsNotNone(mom_hdr, "MOM block header not found")
        weekly = sorted(rows)[:2]
        monthly = sorted(rows)[2:]
        self.assertTrue(all(r < mom_hdr for r in weekly),
                        "weekly grids must sit above the MOM block")
        self.assertTrue(all(r > mom_hdr for r in monthly),
                        "monthly grids must sit below the MOM block")


class TestOverlayShapes(unittest.TestCase):
    """midcap / MIDCPNIFTY / both / neither — all four must produce real grids."""

    def test_plain_run_uses_plain_columns(self):
        mae, live, mt, lt = _min_grid_fields(CLEANED)
        self.assertEqual((mae, live), ("Final MAE", "Actual Live DD"))
        self.assertEqual((mt, lt), ("Min of Final MAE", "Min of Actual Live DD"))

    def test_combined_run_uses_combined_columns(self):
        mae, live, mt, lt = _min_grid_fields(CLEANED_MIDCAP)
        self.assertEqual((mae, live),
                         ("Combined Final MAE", "Combined Actual Live DD"))
        self.assertEqual(mt, "Min of Combined Final MAE")
        self.assertEqual(lt, "Min of Combined Actual Live DD")

    def test_combined_grid_has_values(self):
        ws = _sheet(CLEANED_MIDCAP, has_midcap=True)
        g = _grid(ws, "Min of Combined Final MAE")
        self.assertEqual(g["2024"]["Jan"], -2.0)
        self.assertEqual(g["2024"]["Feb"], -1.0)
        self.assertEqual(g["Grand Total"]["Grand Total"], -2.0)

    def test_missing_columns_degrade_to_blank_not_crash(self):
        rows = [{"Expiry": "15-01-2024", "Exit Date": "15-01-2024",
                 "% P&L": 1.0, "%DD": -1.0}]
        wm = _wm_from_cleaned(rows, False)
        self.assertEqual(wm["mom_mae"], {})
        self.assertEqual(wm["mom_ldd"], {})
        ws = _sheet(rows)          # must still render
        g = _grid(ws, "Min of Final MAE")
        self.assertEqual(g["2024"]["Jan"], "")


class TestMergedSummaryHasNoGrids(unittest.TestCase):
    """Summary sheets stay plain cross-combo grids; pivots live on their own.

    The MIN grids under all 24+ blocks made the summary unreadable, so they moved
    to dedicated sheets where each combination is captioned by name.
    """

    ADJS = ("No Adj", "Rise 1%")

    def _merged(self):
        from openpyxl import Workbook as WB
        from services.optimizer.wow_mom import write_merged_wow_mom
        combos = [{
            "title": f"PE ATM | {adj}", "cleaned": CLEANED, "has_midcap": False,
            "adj_key": adj, "adj_label": adj, "row_key": "PE ATM||",
        } for adj in self.ADJS]
        wb = WB()
        wb.remove(wb.active)
        if write_merged_wow_mom(wb, combos) is not True:
            raise AssertionError("write_merged_wow_mom returned falsy")
        return wb

    def test_no_min_grids_on_either_summary_sheet(self):
        wb = self._merged()
        for name in ("WOW Summary", "MOM Summary"):
            ws = wb[name]
            found = [c.value for row in ws.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("Min of")]
            self.assertEqual(found, [], f"{name} must carry no MIN grids")

    def test_pivot_sheets_exist(self):
        wb = self._merged()
        self.assertIn("WOW Min Pivots", wb.sheetnames)
        self.assertIn("MOM Min Pivots", wb.sheetnames)

    def test_every_combination_is_captioned_by_name(self):
        wb = self._merged()
        for name in ("WOW Min Pivots", "MOM Min Pivots"):
            ws = wb[name]
            captions = [c.value for row in ws.iter_rows() for c in row
                        if isinstance(c.value, str) and c.value.startswith("PE ATM | ")]
            self.assertEqual(len(captions), len(self.ADJS), f"{name} caption count")
            self.assertEqual(sorted(captions),
                             sorted(f"PE ATM | {a}" for a in self.ADJS))

    def test_each_combination_gets_both_pivots(self):
        wb = self._merged()
        for name in ("WOW Min Pivots", "MOM Min Pivots"):
            ws = wb[name]
            grids = [c.value for row in ws.iter_rows() for c in row
                     if isinstance(c.value, str) and c.value.startswith("Min of")]
            self.assertEqual(len(grids), 2 * len(self.ADJS), f"{name} grid count")

    def _merged_n(self, n):
        from openpyxl import Workbook as WB
        from services.optimizer.wow_mom import write_merged_wow_mom
        combos = [{
            "title": f"PE ATM | V{i}", "cleaned": CLEANED, "has_midcap": False,
            "adj_key": f"V{i}", "adj_label": f"V{i}", "row_key": "PE ATM||",
        } for i in range(n)]
        wb = WB()
        wb.remove(wb.active)
        if write_merged_wow_mom(wb, combos) is not True:
            raise AssertionError("write_merged_wow_mom returned falsy")
        return wb

    def test_units_tile_three_across_then_wrap(self):
        """7 combos -> rows of 3, 3, 1; columns reused, no two units overlap."""
        wb = self._merged_n(7)
        for name in ("WOW Min Pivots", "MOM Min Pivots"):
            ws = wb[name]
            caps = sorted((c.row, c.column) for row in ws.iter_rows() for c in row
                          if isinstance(c.value, str) and c.value.startswith("PE ATM | V"))
            self.assertEqual(len(caps), 7, name)
            by_row = {}
            for r, c in caps:
                by_row.setdefault(r, []).append(c)
            widths = [len(v) for _, v in sorted(by_row.items())]
            self.assertEqual(widths, [3, 3, 1], f"{name} should tile 3 per line")
            # every band starts at the same three column positions
            bands = [sorted(v) for _, v in sorted(by_row.items())]
            self.assertEqual(bands[0][:len(bands[1])], bands[1])
            self.assertEqual(len(set(caps)), len(caps), "units overlap")

    def test_bands_do_not_collide_vertically(self):
        """The second band must start below the first band's last grid row."""
        wb = self._merged_n(4)
        for name in ("WOW Min Pivots", "MOM Min Pivots"):
            ws = wb[name]
            caps = sorted((c.row, c.column) for row in ws.iter_rows() for c in row
                          if isinstance(c.value, str) and c.value.startswith("PE ATM | V"))
            first_band = min(r for r, _ in caps)
            second_band = max(r for r, _ in caps)
            self.assertGreater(second_band, first_band, name)
            # last "Grand Total" of band 1 sits above band 2's caption
            gts = [c.row for row in ws.iter_rows() for c in row
                   if c.value == "Grand Total" and c.row < second_band]
            self.assertTrue(gts and max(gts) < second_band,
                            f"{name}: band 2 overlaps band 1")

    def test_pivot_sheets_use_the_right_axis(self):
        wb = self._merged()
        wow, mom = wb["WOW Min Pivots"], wb["MOM Min Pivots"]
        r0, c0 = _find_all(wow, "Min of Final MAE")[0]
        self.assertEqual([wow.cell(r0 + 2, c0 + i).value for i in (1, 2)], ["W1", "W2"])
        r0, c0 = _find_all(mom, "Min of Final MAE")[0]
        self.assertEqual([mom.cell(r0 + 1, c0 + i).value for i in (1, 2)], ["Jan", "Feb"])

    def test_blocks_still_present(self):
        wb = self._merged()
        for name in ("WOW Summary", "MOM Summary"):
            ws = wb[name]
            blocks = [c.value for row in ws.iter_rows() for c in row
                      if isinstance(c.value, str) and " | " in c.value]
            self.assertEqual(len(blocks), 2, f"{name} lost its combo blocks")

    def test_per_combo_sheet_still_has_them(self):
        """Removing them from the merged sheet must not touch the tradesheet."""
        ws = _sheet(CLEANED)
        found = [c.value for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith("Min of")]
        self.assertEqual(len(found), 4)


class TestExistingBlocksUnchanged(unittest.TestCase):

    def test_wow_and_mom_numbers_are_untouched(self):
        """Adding grids must not perturb the blocks above them."""
        wm = _wm_from_cleaned(CLEANED, False)
        self.assertEqual(wm["n_trades"], 5)
        self.assertEqual(wm["mom_years"], [2024, 2025])
        # MOM Jan-2024 total is still the SUM of both trades' returns.
        self.assertAlmostEqual(wm["mom"][2024]["months"]["Jan"], 0.03, places=9)


if __name__ == "__main__":
    unittest.main()
