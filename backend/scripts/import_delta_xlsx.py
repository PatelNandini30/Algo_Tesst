#!/usr/bin/env python3
"""
import_delta_xlsx.py — load eod_delta_full_history.xlsx into option_data.delta/iv_pct.

Sheet → DB symbol mapping:
  NIFTY      → option_data.symbol = 'NIFTY'
  MIDCPNIFTY → option_data.symbol = 'MIDCPNIFTY'

Columns used: Date, Symbol, Expiry, Strike, Type, IV%, Delta
Key match:    option_data.(date, symbol, expiry_date, strike_price, option_type)

Runs an UPDATE (never INSERT) — only patches existing option rows.

Usage (inside backend container):
    python scripts/import_delta_xlsx.py /tmp/delta.xlsx
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, datetime
from pathlib import Path

_BACKEND_DIR = Path(__file__).resolve().parent.parent
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

import openpyxl
from migrate_data import _conn

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("import_delta_xlsx")

SHEET_SYMBOL = {
    "NIFTY":      "NIFTY",
    "MIDCPNIFTY": "MIDCPNIFTY",
}


def _to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet %r not found, skipping", sheet_name)
        return []
    ws = wb[sheet_name]
    # Header: Date, Symbol, Expiry, Strike, Type, DaysToExpiry, SpotClose, Premium, IV%, Delta
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        if not row or row[0] is None:
            continue
        dt      = _to_date(row[0])
        expiry  = _to_date(row[2])
        strike  = _to_float(row[3])
        opt_type = str(row[4]).strip().upper() if row[4] else None
        iv_pct  = _to_float(row[8])
        delta   = _to_float(row[9])
        if dt is None or expiry is None or strike is None or not opt_type:
            continue
        rows.append({
            "date": dt, "expiry": expiry, "strike": strike,
            "type": opt_type, "iv_pct": iv_pct, "delta": delta,
        })
    return rows


def _update_batch(cur, symbol: str, rows: list[dict]) -> int:
    """UPDATE option_data rows by key; returns rows updated."""
    if not rows:
        return 0
    # Build temp table, bulk COPY, then UPDATE via JOIN.
    cur.execute("""
        CREATE TEMP TABLE tmp_delta (
            date        DATE,
            symbol      VARCHAR,
            expiry_date DATE,
            strike_price NUMERIC,
            option_type VARCHAR,
            delta       NUMERIC,
            iv_pct      NUMERIC
        ) ON COMMIT DROP
    """)

    import io
    import csv as _csv
    buf = io.StringIO()
    w = _csv.writer(buf)
    for r in rows:
        w.writerow([
            r["date"], symbol, r["expiry"], r["strike"],
            r["type"],
            "" if r["delta"] is None else r["delta"],
            "" if r["iv_pct"] is None else r["iv_pct"],
        ])
    buf.seek(0)
    cur.copy_expert(
        "COPY tmp_delta(date,symbol,expiry_date,strike_price,option_type,delta,iv_pct) "
        "FROM STDIN WITH (FORMAT csv, NULL '')",
        buf,
    )

    cur.execute("""
        UPDATE option_data o
        SET delta   = t.delta,
            iv_pct  = t.iv_pct
        FROM tmp_delta t
        WHERE o.date         = t.date
          AND o.symbol       = t.symbol
          AND o.expiry_date  = t.expiry_date
          AND o.strike_price = t.strike_price
          AND o.option_type  = t.option_type
    """)
    return cur.rowcount


def run(xlsx_path: Path) -> None:
    logger.info("Opening %s", xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    conn = _conn()
    total = 0
    for sheet_name, db_symbol in SHEET_SYMBOL.items():
        rows = _read_sheet(wb, sheet_name)
        if not rows:
            logger.info("  %s → 0 rows", sheet_name)
            continue
        # Batch by 100k to keep memory reasonable
        BATCH = 100_000
        updated = 0
        for i in range(0, len(rows), BATCH):
            chunk = rows[i:i + BATCH]
            with conn.cursor() as cur:
                updated += _update_batch(cur, db_symbol, chunk)
            conn.commit()
        logger.info("  %s → %s: %d rows updated", sheet_name, db_symbol, updated)
        total += updated
    wb.close()
    logger.info("Total rows updated: %d", total)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Path to eod_delta_full_history.xlsx")
    args = ap.parse_args()
    run(Path(args.xlsx))


if __name__ == "__main__":
    main()
