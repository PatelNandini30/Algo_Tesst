#!/usr/bin/env python3
"""
import_ohlc_xlsx.py — load DATA OF OHLC xlsx into spot_data + index_ohlc tables.

Only daily sheets (_D suffix) are imported.

Symbol mapping (xlsx Ticker → DB symbol):
  "NIFTY 50"           → spot_data.symbol = 'NIFTY'
  "NIFTYBANK"          → spot_data.symbol = 'BANKNIFTY'
  "NIFTYMIDCAPSELECT"  → spot_data.symbol = 'MIDCPNIFTY'
  "NIFTYMIDCAP100"     → index_ohlc.symbol = 'NIFTYMIDCAP100'

Note: FINNIFTY is NOT in this xlsx. Update it separately from NSE archives.

Usage (inside backend container):
    python scripts/import_ohlc_xlsx.py /tmp/ohlc.xlsx
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, datetime
from pathlib import Path

_BACKEND_DIR = Path(__file__).resolve().parent.parent
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

import openpyxl
from migrate_data import _conn, _copy_stage, _upsert

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("import_ohlc_xlsx")

# Daily sheet name → (DB table, DB symbol)
DAILY_SHEETS = {
    "Nifty_D":       ("spot_data",   "NIFTY"),
    "BNF_D":         ("spot_data",   "BANKNIFTY"),
    "Mid_Select_D":  ("spot_data",   "MIDCPNIFTY"),
    "Mid_D":         ("index_ohlc",  "NIFTYMIDCAP100"),
}

SPOT_COLS    = ["date", "symbol", "open", "high", "low", "close"]
SPOT_KEY     = ["date", "symbol"]
SPOT_TYPES   = {"date": "DATE", "symbol": "VARCHAR", "open": "NUMERIC",
                "high": "NUMERIC", "low": "NUMERIC", "close": "NUMERIC"}

OHLC_COLS    = ["symbol", "trade_date", "open_price", "high_price", "low_price", "close_price"]
OHLC_KEY     = ["symbol", "trade_date"]
OHLC_TYPES   = {"symbol": "VARCHAR", "trade_date": "DATE", "open_price": "NUMERIC",
                "high_price": "NUMERIC", "low_price": "NUMERIC", "close_price": "NUMERIC"}


def _to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_sheet(wb, sheet_name: str, db_symbol: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet %r not found in workbook, skipping", sheet_name)
        return []

    ws = wb[sheet_name]
    rows = []
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        # Skip header row (first row)
        if not header_skipped:
            header_skipped = True
            continue
        if not row or row[0] is None:
            continue
        # Columns: Ticker, Date/Time, Open, High, Low, Close, %Chg, LogReturn
        dt = _to_date(row[1])
        if dt is None:
            continue
        o = _to_float(row[2])
        h = _to_float(row[3])
        lo = _to_float(row[4])
        c = _to_float(row[5])
        if c is None:
            continue
        rows.append({
            "dt": dt, "open": o, "high": h, "low": lo, "close": c,
        })
    return rows


def _upsert_spot(conn, rows: list[dict], symbol: str) -> tuple[int, int]:
    import pandas as pd
    df = pd.DataFrame([{
        "date": r["dt"], "symbol": symbol,
        "open": r["open"], "high": r["high"], "low": r["low"], "close": r["close"],
    } for r in rows])
    if df.empty:
        return 0, 0
    df["date"] = pd.to_datetime(df["date"]).dt.date
    stage = f"stg_ohlc_spot_{symbol.lower()}"
    with conn.cursor() as cur:
        _copy_stage(cur, df, stage)
        upd, ins = _upsert(cur, stage, "spot_data", SPOT_KEY, SPOT_COLS, SPOT_TYPES)
    conn.commit()
    return upd, ins


def _upsert_index_ohlc(conn, rows: list[dict], symbol: str) -> tuple[int, int]:
    import pandas as pd
    df = pd.DataFrame([{
        "symbol": symbol, "trade_date": r["dt"],
        "open_price": r["open"], "high_price": r["high"],
        "low_price": r["low"], "close_price": r["close"],
    } for r in rows])
    if df.empty:
        return 0, 0
    df["trade_date"] = pd.to_datetime(df["trade_date"]).dt.date
    stage = f"stg_ohlc_idx_{symbol.lower()}"
    with conn.cursor() as cur:
        _copy_stage(cur, df, stage)
        upd, ins = _upsert(cur, stage, "index_ohlc", OHLC_KEY, OHLC_COLS, OHLC_TYPES)
    conn.commit()
    return upd, ins


def run(xlsx_path: Path) -> None:
    logger.info("Opening %s", xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    conn = _conn()

    total_ins = total_upd = 0
    for sheet_name, (table, db_symbol) in DAILY_SHEETS.items():
        rows = _read_sheet(wb, sheet_name, db_symbol)
        if not rows:
            logger.info("  %s → 0 rows, skipped", sheet_name)
            continue

        if table == "spot_data":
            upd, ins = _upsert_spot(conn, rows, db_symbol)
        else:
            upd, ins = _upsert_index_ohlc(conn, rows, db_symbol)

        logger.info("  %s → %s.%s: %d inserted, %d updated",
                    sheet_name, table, db_symbol, ins, upd)
        total_ins += ins
        total_upd += upd

    wb.close()
    logger.info("Total: %d inserted, %d updated", total_ins, total_upd)


def main() -> None:
    ap = argparse.ArgumentParser(description="Import OHLC xlsx into spot_data + index_ohlc")
    ap.add_argument("xlsx", help="Path to DATA OF OHLC xlsx file")
    args = ap.parse_args()
    run(Path(args.xlsx))


if __name__ == "__main__":
    main()
