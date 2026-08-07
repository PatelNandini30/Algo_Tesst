import warnings; warnings.filterwarnings("ignore")
import openpyxl
wb=openpyxl.load_workbook("/app/_f5.xlsx", data_only=True)
print("SHEETS:",wb.sheetnames)
print("\n===== RULES =====")
for row in wb["Rules"].iter_rows(values_only=True):
    c=[str(x) for x in row if x not in (None,"")]
    if c: print(" | ".join(c))
