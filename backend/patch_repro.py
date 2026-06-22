import math
import pandas as pd
from datetime import datetime
from base import bulk_load_options, bulk_clear_options
from services.algotest_job import _try_rust_engine, _build_fast_lookup_from_bulk
from services.optimizer.excel_builder import compute_midcap_for_rows
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

segs = [("2019-03-29","2019-05-10"),("2019-09-30","2020-02-28"),("2022-07-29","2022-12-23"),
        ("2023-04-28","2024-03-28"),("2024-04-30","2024-12-31"),("2025-03-28","2026-02-25")]
ef, et = "2019-01-01", "2026-06-01"
payload = {
  "index":"NIFTY","from_date":ef,"to_date":et,"date_from":ef,"date_to":et,
  "expiry_type":"weekly","strategy_type":"positional",
  "entry_dte":0,"exit_dte":0,"rollover_toggle":True,
  "slippage_pct":0,"charges_enabled":False,"square_off_mode":"partial",
  "filter":"custom","filter_config":"custom",
  "filter_segments":[{"start":s,"end":e} for s,e in segs],
  "legs":[{"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY",
           "strike_interval":100,"strike_selection":{"type":"strike_type","strike_type":"ATM"}}],
  "midcap_legs":[{"symbol":"NIFTYMIDCAP100","midcap_mode":"hypothetical","cost_pct_per_month":0.5,"position":"SELL","lots":1}],
}
bulk_load_options("NIFTY",ef,et); _build_fast_lookup_from_bulk("NIFTY",ef,et)
df,summ,_ = _try_rust_engine(payload,"NIFTY",ef,et)
bulk_clear_options()
if df is None or df.empty:
    print("NO TRADES"); raise SystemExit
rows = df.where(df.notna(), None).to_dict("records")
mbt, msum, hasm = compute_midcap_for_rows(rows, payload["midcap_legs"], None)
print("trades(legs):", len(rows), "midcap available:", hasm, "midcap trades:", len(mbt))

def numv(v):
    try:
        if v is None or v=="" : return None
        return float(v)
    except: return None
def pdate(v):
    if not v: return None
    s=str(v)[:10]
    try: return datetime.strptime(s, "%Y-%m-%d")
    except:
        try: return datetime.strptime(s, "%d-%m-%Y")
        except: return None

# group by trade, chronological
grp={}
for r in rows:
    k=str(r.get("Trade") or 1); grp.setdefault(k,[]).append(r)
def keyfn(k):
    legs=grp[k]; m=next((l for l in legs if not l.get("ReEntryIndex") and not l.get("ReEntryTrigger") and not l.get("ReEntryMode")), legs[0])
    d=pdate(m.get("Entry Date")) or datetime.max
    return (d, int(k))
okeys=sorted(grp.keys(), key=keyfn)

tdata=[]
for k in okeys:
    legs=grp[k]
    main=next((l for l in legs if not l.get("ReEntryIndex") and not l.get("ReEntryTrigger") and not l.get("ReEntryMode")), legs[0])
    ce=next((l for l in legs if str(l.get("Type","")).upper() in ("CE","CALL")), main)
    spot=numv(main.get("Entry Spot")) or 0
    mc=mbt.get(k) or {}
    cep=numv(ce.get("CE P&L"))
    niftyMae=numv(ce.get("MAE")) or 0; niftyMfe=numv(ce.get("MFE")) or 0
    midMae=numv(mc.get("Midcap MAE")) or 0; midMfe=numv(mc.get("Midcap MFE")) or 0
    cpct=numv(mc.get("Combined Net P&L %"))
    nm1=midMfe+niftyMae; nm2=midMae+niftyMfe
    cfm = min(nm1,nm2,cpct) if cpct is not None else None
    tdata.append(dict(
        entry=str(main.get("Entry Date"))[:10], exit=str(main.get("Exit Date"))[:10],
        entryMs=(pdate(main.get("Entry Date")).timestamp()*1000 if pdate(main.get("Entry Date")) else None),
        exitMs=(pdate(main.get("Exit Date")).timestamp()*1000 if pdate(main.get("Exit Date")) else None),
        midcapPct=numv(mc.get("Midcap Leg P&L %")), midcapMae=numv(mc.get("Midcap MAE")), midcapClose=numv(mc.get("Midcap Entry Spot")),
        callPct=((cep/spot*100) if (cep is not None and spot>0) else None), callMae=numv(ce.get("MAE")),
        combinedPct=cpct, combinedMae=cfm))

# gap detection -> patches
GAP=30*86400000; patches=[]; last=None
for td in tdata:
    gap = (td["entryMs"]-last) if (last is not None and td["entryMs"] is not None) else 0
    if not patches or gap>GAP: patches.append([])
    patches[-1].append(td)
    if td["exitMs"] is not None: last=td["exitMs"]
print("patches:", len(patches), [(p[0]["entry"], p[-1]["exit"], len(p)) for p in patches])

def chain(trades, dkey, mkey):
    prev=100.0; peak=100.0; out=[]; pnl=0.0; lmin=math.inf
    for td in trades:
        d=td[dkey] if isinstance(td[dkey],float) else (td[dkey] or 0)
        d=d if d is not None else 0
        cumm=prev*(1+d/100); peak=max(peak,cumm); dd=cumm-peak; pdd=(dd/peak) if peak else 0
        m=td[mkey]; m=m if m is not None else 0
        lnav=prev*(1+m/100); live=((lnav/peak-1)*100) if peak else 0
        out.append(dict(td=td,drive=d,cumm=cumm,peak=peak,dd=dd,pdd=pdd,mae=m,lnav=lnav,live=live))
        pnl+=d; lmin=min(lmin,live); prev=cumm
    f=trades[0]; l=trades[-1]
    days=((l["exitMs"]-f["entryMs"])/86400000) if (f["entryMs"] and l["exitMs"]) else None
    cagr=((math.pow(out[-1]["cumm"]/100,365/days)-1)*100) if (days and days>0 and out[-1]["cumm"]>0) else None
    return dict(rows=out, entry=f["entry"], exit=l["exit"], cagr=cagr, pnl=pnl, lmin=(None if lmin==math.inf else lmin))

PHASES=[("Midcap Future","midcapPct","midcapMae","Future P&L %"),
        ("Nifty CE","callPct","callMae","Net P&L %"),
        ("Nifty CE + Midcap Future","combinedPct","combinedMae","Net P&L %")]
for title,dk,mk,plab in PHASES:
    print(f"\n=== {title}  side table (Entry|Exit|CAGR|{plab}|LiveDD) ===")
    for p in patches:
        ch=chain(p,dk,mk)
        print(f"  {ch['entry']}  {ch['exit']}  CAGR={ch['cagr']:.2f}%  {plab}={ch['pnl']:.2f}  LiveDD={ch['lmin']:.2f}")

# write xlsx
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Patch wise"
hf=Font(bold=True,color="FFFFFFFF"); hfill=PatternFill("solid",fgColor="FF34495E")
tfill=PatternFill("solid",fgColor="FF1F3864")
col=1
for title,dk,mk,plab in PHASES:
    chains=[chain(p,dk,mk) for p in patches]
    midcap = (dk=="midcapPct")
    dhdr = (["Entry Date","Exit Date","Midcap Hypo P&L %","cumm","Peak","Close","Hypo MAE","Lowest NAV","Live DD"] if midcap
            else ["Net P&L %","Cumulative","Peak","DD","%DD","MAE","Lowest NAV","Actual Live DD"])
    dW=len(dhdr); ds=col; ss=col+dW+1
    c=ws.cell(1,ds,title); c.font=Font(bold=True,color="FFFFFFFF"); c.fill=tfill
    ws.cell(2,ds,"Phase wise Distribution").font=Font(bold=True)
    for i,h in enumerate(dhdr):
        cc=ws.cell(4,ds+i,h); cc.font=hf; cc.fill=hfill
    rr=5
    for ch in chains:
        for rw in ch["rows"]:
            cc=ds
            if midcap:
                ws.cell(rr,cc,rw["td"]["entry"]); cc+=1; ws.cell(rr,cc,rw["td"]["exit"]); cc+=1
                ws.cell(rr,cc,round(rw["drive"],4)); cc+=1; ws.cell(rr,cc,round(rw["cumm"],4)); cc+=1
                ws.cell(rr,cc,round(rw["peak"],4)); cc+=1; ws.cell(rr,cc,rw["td"]["midcapClose"]); cc+=1
                ws.cell(rr,cc,round(rw["mae"],4)); cc+=1; ws.cell(rr,cc,round(rw["lnav"],4)); cc+=1; ws.cell(rr,cc,round(rw["live"],4)); cc+=1
            else:
                ws.cell(rr,cc,round(rw["drive"],4)); cc+=1; ws.cell(rr,cc,round(rw["cumm"],4)); cc+=1
                ws.cell(rr,cc,round(rw["peak"],4)); cc+=1; ws.cell(rr,cc,round(rw["dd"],4)); cc+=1
                ws.cell(rr,cc,round(rw["pdd"],4)); cc+=1; ws.cell(rr,cc,round(rw["mae"],4)); cc+=1
                ws.cell(rr,cc,round(rw["lnav"],4)); cc+=1; ws.cell(rr,cc,round(rw["live"],4)); cc+=1
            rr+=1
    shdr=["Entry","Exit","CAGR",plab,"Live DD"]
    for i,h in enumerate(shdr):
        cc=ws.cell(4,ss+i,h); cc.font=hf; cc.fill=PatternFill("solid",fgColor="FF2C5F8A")
    for i,ch in enumerate(chains):
        sr=5+i
        ws.cell(sr,ss,ch["entry"]); ws.cell(sr,ss+1,ch["exit"])
        ws.cell(sr,ss+2,(round(ch["cagr"],2) if ch["cagr"] is not None else ""))
        ws.cell(sr,ss+3,round(ch["pnl"],2)); ws.cell(sr,ss+4,(round(ch["lmin"],2) if ch["lmin"] is not None else ""))
    col=ss+len(shdr)+1
wb.save("/app/patch_wise_test.xlsx")
print("\nWROTE /app/patch_wise_test.xlsx")
