import warnings; warnings.filterwarnings("ignore")
import openpyxl
wb=openpyxl.load_workbook("/app/_optfile.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
for sn in wb.sheetnames:
    ws=wb[sn]
    print("\n===== %s (%dx%d) ====="%(sn,ws.max_row,ws.max_column))
