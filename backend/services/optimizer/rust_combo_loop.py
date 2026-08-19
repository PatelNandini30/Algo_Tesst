"""
Rust combo-loop scaffold (Phase 1 groundwork) — flag + whitelist + shadow differ.

This module is PURELY ADDITIVE. Nothing in the live optimizer path imports it yet;
Phase 1 wires it in. With ``OPTIMIZE_RUST_LOOP=0`` (the default) the Rust batch path
is never reached and the optimizer behaves exactly as today.

HARD RULE (user): the optimizer must be **Rust-bounded only — NO Python fallback at
any point**. See [[rust-only-no-python-fallback]]. There is no pure-Python engine on
any live path (``run_algotest_backtest`` is already deleted). Today's per-combo path
(``run_rust_engine_pipeline`` — Rust-priced, Python-orchestrated) is retained **solely
as an offline / shadow PARITY REFERENCE**, never as a runtime fallback. When the Rust
batch is authoritative, a combo it does not own **hard-fails (RuntimeError)** — it is
never silently routed anywhere else.

Three pieces (see RUST_COMBO_LOOP_DESIGN.md §5, §6.3):

1. ``rust_loop_mode()`` — the ``OPTIMIZE_RUST_LOOP`` flag: ``{"0","shadow","1"}``.
     0       today's engine only (Rust batch unreached) — the default while porting.
     shadow  today's engine produces the real output AND is the parity reference; the
             Rust batch runs alongside for the shapes it owns and a differ logs any
             mismatch. Risk-free evidence on real jobs. (This is the ONLY place today's
             engine is used at runtime — as the reference, not a fallback.)
     1       Rust batch is the ONLY runtime calculation path. A combo the batch does not
             own hard-fails via ``require_rust_supported`` — NO fallback. Turned on only
             after every in-scope feature is ported to Rust and shadow-clean.

2. ``rust_batch_unsupported(merged_payload)`` — a FAIL-CLOSED coverage gate (design R5).
   It inspects the EFFECTIVE post-merge config (the output of
   ``param_expander.apply_combo_for_optim``, which force-enables spot-adj / midcap /
   pct_of_atm / straddle_width) and returns a reason string for any shape the pure-Rust
   batch does not yet OWN. Returns ``None`` only for shapes the batch fully owns.
   ``require_rust_supported`` raises RuntimeError on any non-None reason (the authoritative
   hard-fail). As features are ported to Rust, the corresponding reject is lifted — and
   only after that feature's parity corpus is clean — until the gate returns None for
   everything and there is nothing left to hard-fail on.

3. The shadow differ — ``diff_summary`` / ``diff_trades`` / ``diff_redis_row`` /
   ``diff_xlsx_cells`` — compares the reference (today's engine) artifacts against the
   Rust batch artifacts and returns human-readable diffs (logged, never raised).

Nothing here changes a single number: the flag gates whether the Rust path runs at all,
the coverage gate only ever *narrows* what Rust touches (and hard-fails otherwise —
never falls back), and the differ is read-only.
"""
from __future__ import annotations

import logging
import math
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


# ── 1. Flag ──────────────────────────────────────────────────────────────────

def rust_loop_mode() -> str:
    """Return the OPTIMIZE_RUST_LOOP mode: one of "0" | "shadow" | "1".

    Unknown / unset values fall back to "0" (today's fork-pool path) so a typo can
    never silently enable the Rust path.
    """
    v = str(os.environ.get("OPTIMIZE_RUST_LOOP", "0")).strip().lower()
    return v if v in ("0", "shadow", "1") else "0"


def rust_loop_enabled() -> bool:
    """True when the Rust batch should run at all (shadow OR authoritative)."""
    return rust_loop_mode() in ("shadow", "1")


def rust_loop_authoritative() -> bool:
    """True only in mode "1" — the Rust batch is the ONLY runtime path and a combo it
    does not own hard-fails (no fallback)."""
    return rust_loop_mode() == "1"


# ── 2. Rust-batch coverage gate — fail-closed, hard-fail (no Python fallback) ─
#
# The current pure-Rust batch owns ONLY: resolve_trade_specs_core +
# simulate_trades_batch_core + (Python-side, for now) analytics/MAE-MFE. It does NOT
# yet own the SL/Target/Trail scan (Phase 0b) nor any Python-orchestrated feature
# (spot-adj, midcap, re-entry, next-weekly, lazy, filters, futures). So the gate below
# rejects ALL of those. A rejected combo does NOT fall back to Python — in authoritative
# mode it hard-fails (require_rust_supported → RuntimeError). As each capability lands
# the corresponding reject is lifted — and only after the parity corpus is clean for it
# (design R5, §7) — until the gate returns None for everything.

# Strike-selection modes compute_strike_for_leg resolves in Rust (engine_rust.py:1189+).
_RUST_STRIKE_TYPES = frozenset({
    "", "strike_type",
    "rel_leg", "pct_of_atm",
    "closest_premium", "premium_gte", "premium_lte", "premium_range",
    "time_value", "time_value_gte", "time_value_lte",
    "straddle_width", "atm_straddle_prem_pct",
})
# expiry values that mean "trade a contract one expiry further out" — Python-orchestrated.
_NEXT_EXPIRY_TYPES = frozenset({"NEXT_WEEKLY", "WEEKLY_T1", "NEXT_MONTHLY", "MONTHLY_T1"})


def _truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return v.strip().lower() not in ("", "0", "false", "no", "none")
    if isinstance(v, (list, tuple, dict, set)):
        return len(v) > 0
    return bool(v)


def _leg_has_exit_scan(leg: Dict[str, Any]) -> bool:
    """True if the leg carries an SL/Target/Trail or buffer-strike exit — none of
    which the pure-Rust batch owns yet (Phase 0b extracts the SL scan)."""
    for k in ("stopLoss", "targetProfit", "trailSL"):
        v = leg.get(k)
        if isinstance(v, dict) and any(_truthy(v.get(sub)) for sub in
                                       ("enabled", "value", "trigger", "move", "type")):
            return True
        elif not isinstance(v, dict) and _truthy(v):
            return True
    if _truthy(leg.get("buffer_strike_enabled")):
        return True
    return False


def _leg_has_spot_adjustment(leg: Dict[str, Any]) -> bool:
    """True if the leg carries its OWN spot_adjustment (the per-leg breach/re-strike
    feature). The pure-Rust batch does NOT implement the spot-adjustment mark-timeline
    / earliest-wins cascade / re-anchor logic (that lives in the Python engine_rust
    orchestration), so claiming such a combo would price it WITHOUT any adjustment —
    silently wrong. The strategy-level knob is already excluded above via
    `spot_adjustment_enabled`; this closes the per-leg case (leg['spot_adjustment'])."""
    for k in ("spot_adjustment", "spotAdjustment"):
        v = leg.get(k)
        if isinstance(v, dict) and _truthy(v.get("enabled")):
            return True
    return False


def _leg_is_futures(leg: Dict[str, Any]) -> bool:
    seg = str(leg.get("segment") or "").strip().lower()
    ot = str(leg.get("option_type") or "").strip().upper()
    return seg in ("futures", "future", "fut") or ot in ("FUT", "FUTIDX")


def _leg_is_lazy(leg: Dict[str, Any]) -> bool:
    return (_truthy(leg.get("lazy_leg_config")) or _truthy(leg.get("lazyLegConfig"))
            or _truthy(leg.get("is_lazy")))


def _leg_is_next_expiry(leg: Dict[str, Any]) -> bool:
    return str(leg.get("expiry") or "").strip().upper() in _NEXT_EXPIRY_TYPES


def _leg_has_reentry(leg: Dict[str, Any]) -> bool:
    return _truthy(leg.get("reEntryOnSL")) or _truthy(leg.get("reEntryOnTarget"))


def rust_batch_unsupported(merged_payload: Dict[str, Any]) -> Optional[str]:
    """Return None if the pure-Rust batch fully OWNS this combo, else a short reason
    string it does not (yet).

    This is a COVERAGE gate, NOT a routing-to-Python gate: a non-None reason means the
    Rust batch can't produce this combo — in authoritative mode that HARD-FAILS
    (see require_rust_supported); there is no Python fallback.

    FAIL-CLOSED: `merged_payload` MUST be the effective post-merge config
    (apply_combo_for_optim output) so force-enabled flags are visible. Anything not
    positively recognized returns a reason. Never returns None on doubt.
    """
    try:
        p = merged_payload or {}
        if not isinstance(p, dict):
            return "payload-not-dict"

        # ── Payload-level orchestration features (all Python-orchestrated today) ──
        # Multi-index (per-leg index) runs price each leg from its OWN index series
        # via run_multi_index_feature / run_sync_weekly_cadence — the single-index
        # Rust batch must never claim them (it would misprice cross-index legs).
        # LIFTED 2026-08-14 (multi_index). The comment above is a PRICING concern —
        # mode 1 does not price: run_multi_index_feature / run_sync_weekly_cadence
        # already produced trades_df, and rust_authoritative_summary only summarises
        # it. compute_summary_metrics is ALREADY the live path for every multi-index
        # sweep today (excel_builder._cmetrics is Rust-only, ungated on multi_index).
        # Measured on 24 real combos of a multi_index + sync_weekly_roll + FUTURES +
        # YEARLY + 14-filter-segment sweep (job 2e133fd4): 2,688 fields compared,
        # CLEAN=24, WITH-DIFFS=0, overall AND patchwise.
        # if _truthy(p.get("multi_index_mode")):
        #     return "multi_index"
        # LIFTED 2026-08-17 (payload-level spot_adjustment). The adjustment is applied
        # by the Python-orchestrated engine BEFORE the summary runs, so trades_df
        # already carries the re-strikes/re-entries. Corpus job a2dedd4c
        # (spot_adjustment_enabled=True, 1% rise) was genuinely active — 56 SPOT_ADJ
        # exits across 504 rows — 7 combos / 784 fields / 0 diffs.
        # if _truthy(p.get("spot_adjustment_enabled")):
        #     return "spot_adjustment"
        # LIFTED 2026-08-14 (midcap). rust_authoritative_summary computes the overlay
        # itself via compute_midcap_for_rows and passes midcap_by_trade/midcap_summary
        # into compute_summary_metrics — the same inputs mode 0 gives _cmetrics.
        # Corpus job 8c6273a4 (hypothetical buy, cost 0.5%/mo): has_midcap=True,
        # midcap_leg_pnl_sum=20,392.97, combined_pnl_sum=21,314.80 — i.e. the overlay
        # was genuinely active, not inert. 6 combos / 738 fields / 0 diffs.
        # msa = p.get("midcap_spot_adjustment")
        # if _truthy(p.get("midcap_legs")) or (isinstance(msa, dict) and _truthy(msa.get("enabled"))):
        #     return "midcap"
        # LIFTED 2026-08-13 (filter), same basis as the Phase 0b SL/Target lift:
        # in mode 1 the trades still come from the per-combo Rust engine — only the
        # SUMMARY is Rust-authoritative — and compute_summary_metrics TAKES
        # filter_segments as a parameter (it is already the live production path via
        # excel_builder._cmetrics for every filtered sweep). Parity measured on 96
        # real combos of a 7-filter-segment + YEARLY + per-leg-spot-adj sweep:
        # CLEAN=96, WITH-DIFFS=0 across overall (71 keys) and patchwise (41 keys),
        # plus tools/mode1_parity on the 3-payload corpus. Re-add this reject if
        # a future filter feature stops being expressible as filter_segments.
        # if _truthy(p.get("filter_segments")) or _truthy(p.get("filter_config")):
        #     return "filter"
        # Per-leg individual filter files are applied by the Python spec
        # post-pass (services/leg_filter.py); the Rust batch has no notion of
        # them and would price the masked leg over its full window.
        # LIFTED 2026-08-14 (leg_filter). The per-leg mask is applied by the Python
        # spec post-pass BEFORE the summary runs, so trades_df already reflects it.
        # Corpus job 26736821 (leg0 masked to 3 of the payload's 7 segments) proved the
        # mask was live: rows per leg = {1: 73, 2: 263, 3: 263}. 6 combos / 672 fields
        # / 0 diffs.
        # for _leg in (p.get("legs") or []):
        #     if isinstance(_leg, dict) and _truthy(_leg.get("filter_segments")):
        #         return "leg_filter"
        # STILL REJECTED. Attempted 2026-08-14 and NOT lifted: two corpus sweeps
        # (overall_sl_value=25 then =2) produced ZERO SL exits across 4,536 rows —
        # every exit was SCHEDULED_EXIT / SPOT_ADJ / FILTER_END. A "0 diffs" result on
        # a stop that never fires proves nothing, so this stays gated until a corpus
        # that actually triggers it exists.
        if _truthy(p.get("overall_sl_value")) or _truthy(p.get("overall_target_value")):
            return "overall_sl_target"  # Phase 0b (SL scan) not yet extracted
        # YEARLY pins the contract to a December expiry while the cadence list
        # drives entry/exit, and needs the Python-resolved `yearly_cycles` to do
        # it. This batch loop builds its own expiry inputs, so it would silently
        # trade the CADENCE contract. Exclude explicitly rather than let an
        # unrecognised expiry kind fall through as "supported".
        # LIFTED 2026-08-13 (yearly_expiry): the concern above is a PRICING one —
        # this batch "would silently trade the CADENCE contract" — but mode 1 does
        # not price; _run_single_backtest already produced the trades. The same 96
        # real combos carried expiry_type=YEARLY and matched exactly.
        # if str(p.get("expiry_type") or "").upper() == "YEARLY":
        #     return "yearly_expiry"
        # SAME-INDEX MIXED EXPIRY — identical hazard to YEARLY above. A MONTHLY
        # leg under a WEEKLY cadence is pinned to its own monthly contract by
        # _build_fixed_entry_specs; this batch loop builds its own expiry inputs
        # and would silently trade the CADENCE (weekly) contract instead.
        # Fail closed so mixed baskets fall back to the per-combo path that
        # actually resolves the pin.
        # LIFTED 2026-08-17 (mixed_expiry). Same reasoning as YEARLY above: the pin is
        # resolved by _build_fixed_entry_specs during PRICING, and mode 1 only
        # summarises the finished trades. Corpus job 6f15ab87 (WEEKLY cadence +
        # MONTHLY option legs) genuinely mixed — 17 distinct expiries across 483 rows
        # — 7 combos / 784 fields / 0 diffs. This shape hard-failed job 027a9a6a.
        # if str(p.get("expiry_type") or "").upper() in ("WEEKLY", "NEXT_WEEKLY", "WEEKLY_T1"):
        #     for _l in p.get("legs") or []:
        #         if (... MONTHLY option leg ...):
        #             return "mixed_expiry"

        legs = p.get("legs") or []
        if not isinstance(legs, list) or len(legs) == 0:
            return "no-legs"

        for i, leg in enumerate(legs):
            if not isinstance(leg, dict):
                return f"leg{i}-not-dict"
            # LIFTED 2026-08-13 (per-leg spot_adjustment): all three legs of the
            # measured corpus had spot_adjustment enabled and swept; summaries matched.
            # The PAYLOAD-level spot_adjustment_enabled reject above is UNCHANGED.
            # if _leg_has_spot_adjustment(leg):
            #     return f"leg{i}-spot_adjustment"
            # LIFTED 2026-08-14 (futures leg): same basis — the futures leg is priced
            # by the Python-orchestrated path before the summary runs. The corpus above
            # carried a NIFTY FUTURES monthly leg; fut_pnl_total/fut_pnl_pct are
            # supplemented explicitly in rust_authoritative_summary and matched exactly.
            # if _leg_is_futures(leg):
            #     return f"leg{i}-futures"
            if _leg_is_lazy(leg):
                return f"leg{i}-lazy"
            if _leg_is_next_expiry(leg):
                return f"leg{i}-next_expiry"
            if _leg_has_reentry(leg):
                return f"leg{i}-reentry"
            # Phase 0b LIFTED (2026-07-20): SL / Target / Trail / buffer-strike.
            # In mode-1 the trades still come from the per-combo Rust engine
            # (_run_single_backtest) exactly as in mode 0 — only the SUMMARY is
            # Rust-authoritative here — and the Rust summary was proven byte-identical
            # to the Python summary across a 19-case SL/Target/Trail corpus (ATM/ITM/
            # OTM/premium strikes, 1–4 legs, buffer-strike, tight-SL, 2y + COVID-2020
            # ranges): tools/phase0b_sl_parity.py. So this shape no longer hard-fails.
            # (The remaining rejects below are still Python-orchestrated.)
            # if _leg_has_exit_scan(leg):
            #     return f"leg{i}-sl_target_trail"
            sel = leg.get("strike_selection") or {}
            sel_type = str((sel.get("type") if isinstance(sel, dict) else "") or "").strip().lower()
            recognized = (sel_type in _RUST_STRIKE_TYPES
                          or sel_type.startswith(("atm", "itm", "otm")))
            if not recognized:
                return f"leg{i}-strike:{sel_type or 'unknown'}"

        return None  # every gate passed → the pure-Rust batch owns this shape
    except Exception as exc:  # fail-closed on ANY unexpected shape
        return f"coverage-gate-error:{exc}"


def combo_supported(merged_payload: Dict[str, Any]) -> bool:
    """Convenience boolean — True iff the pure-Rust batch owns this combo."""
    return rust_batch_unsupported(merged_payload) is None


def require_rust_supported(merged_payload: Dict[str, Any]) -> None:
    """Authoritative-mode hard-fail: raise RuntimeError if the Rust batch does not own
    this combo. This is the enforcement of the "Rust-bounded only, no Python fallback"
    rule — an unsupported combo errors out; it is NEVER routed to today's engine at
    runtime (that engine exists only as the offline/shadow parity reference).
    """
    reason = rust_batch_unsupported(merged_payload)
    if reason is not None:
        raise RuntimeError(
            f"OPTIMIZE_RUST_LOOP=1 (Rust-bounded, no fallback): combo not owned by the "
            f"Rust batch yet ({reason}). Port this feature to Rust before enabling "
            f"authoritative mode for it."
        )


# ── 3. Shadow differ ─────────────────────────────────────────────────────────
#
# Each field is rounded before storage, so exact equality is the target; a tiny
# tolerance guards against float-repr noise only (design §5). Any real mismatch is a
# bug, not noise. All differs are READ-ONLY and return lists of human strings.

_DEFAULT_TOL = 1e-6


def _num(v: Any) -> Optional[float]:
    try:
        if v is None or isinstance(v, bool):
            return None
        f = float(v)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def _values_match(a: Any, b: Any, tol: float) -> bool:
    fa, fb = _num(a), _num(b)
    if fa is not None and fb is not None:
        return abs(fa - fb) <= tol + tol * max(abs(fa), abs(fb))
    return str(a) == str(b)


def diff_summary(py: Dict[str, Any], rust: Dict[str, Any], tol: float = _DEFAULT_TOL) -> List[str]:
    """Key-by-key diff of two summary dicts over the keys they share, plus any key
    present in one but not the other. Numbers compared with a tiny tolerance."""
    diffs: List[str] = []
    py = py or {}
    rust = rust or {}
    for k in sorted(set(py) | set(rust)):
        if k not in py:
            diffs.append(f"summary[{k}]: only in rust ({rust[k]!r})")
        elif k not in rust:
            diffs.append(f"summary[{k}]: only in python ({py[k]!r})")
        elif not _values_match(py[k], rust[k], tol):
            diffs.append(f"summary[{k}]: py={py[k]!r} rust={rust[k]!r}")
    return diffs


def diff_trades(py_df: Any, rust_df: Any, tol: float = _DEFAULT_TOL,
                max_report: int = 40) -> List[str]:
    """Cell-level diff of two trades DataFrames (the CSV/tradesheet source). Reports
    shape/column mismatches first, then up to `max_report` differing cells."""
    diffs: List[str] = []
    try:
        py_cols = list(getattr(py_df, "columns", []) or [])
        rust_cols = list(getattr(rust_df, "columns", []) or [])
        if py_cols != rust_cols:
            only_py = [c for c in py_cols if c not in rust_cols]
            only_rs = [c for c in rust_cols if c not in py_cols]
            if only_py:
                diffs.append(f"trades cols only in python: {only_py}")
            if only_rs:
                diffs.append(f"trades cols only in rust: {only_rs}")
        n_py, n_rs = len(py_df), len(rust_df)
        if n_py != n_rs:
            diffs.append(f"trades row count: py={n_py} rust={n_rs}")
        common_cols = [c for c in py_cols if c in rust_cols]
        for r in range(min(n_py, n_rs)):
            for c in common_cols:
                a = py_df.iloc[r][c]
                b = rust_df.iloc[r][c]
                if not _values_match(a, b, tol):
                    diffs.append(f"trades[row={r},{c}]: py={a!r} rust={b!r}")
                    if len(diffs) >= max_report:
                        diffs.append("… (truncated)")
                        return diffs
    except Exception as exc:
        diffs.append(f"trades-diff-error: {exc}")
    return diffs


def diff_redis_row(py_row: Dict[str, Any], rust_row: Dict[str, Any],
                   tol: float = _DEFAULT_TOL) -> List[str]:
    """Deep-diff the per-combo Redis result row (combo_columns, objective_value,
    trade_count, has_midcap, inline_finalized, nested summary, …)."""
    diffs: List[str] = []

    def _walk(a: Any, b: Any, path: str) -> None:
        if isinstance(a, dict) or isinstance(b, dict):
            a = a if isinstance(a, dict) else {}
            b = b if isinstance(b, dict) else {}
            for k in sorted(set(a) | set(b)):
                _walk(a.get(k), b.get(k), f"{path}.{k}")
        elif isinstance(a, (list, tuple)) or isinstance(b, (list, tuple)):
            a = list(a) if isinstance(a, (list, tuple)) else []
            b = list(b) if isinstance(b, (list, tuple)) else []
            if len(a) != len(b):
                diffs.append(f"row{path}: len py={len(a)} rust={len(b)}")
            for i in range(min(len(a), len(b))):
                _walk(a[i], b[i], f"{path}[{i}]")
        elif not _values_match(a, b, tol):
            diffs.append(f"row{path}: py={a!r} rust={b!r}")

    _walk(py_row or {}, rust_row or {}, "")
    return diffs


def diff_xlsx_cells(py_path: str, rust_path: str, max_report: int = 40) -> List[str]:
    """Cell-diff two .xlsx files (two writers differ in zip/XML metadata, so byte-diff
    is meaningless — compare sheet names/order, dims, and per-cell (value, number_format)).
    """
    diffs: List[str] = []
    try:
        from openpyxl import load_workbook
        wa = load_workbook(py_path)
        wb = load_workbook(rust_path)
        if wa.sheetnames != wb.sheetnames:
            diffs.append(f"xlsx sheets: py={wa.sheetnames} rust={wb.sheetnames}")
        for sheet in [s for s in wa.sheetnames if s in wb.sheetnames]:
            sa, sb = wa[sheet], wb[sheet]
            if (sa.max_row, sa.max_column) != (sb.max_row, sb.max_column):
                diffs.append(f"xlsx[{sheet}] dims: py={sa.max_row}x{sa.max_column} "
                             f"rust={sb.max_row}x{sb.max_column}")
            for row in range(1, min(sa.max_row, sb.max_row) + 1):
                for col in range(1, min(sa.max_column, sb.max_column) + 1):
                    ca, cb = sa.cell(row=row, column=col), sb.cell(row=row, column=col)
                    if not _values_match(ca.value, cb.value, _DEFAULT_TOL) \
                            or ca.number_format != cb.number_format:
                        diffs.append(
                            f"xlsx[{sheet}]!{ca.coordinate}: "
                            f"py=({ca.value!r},{ca.number_format}) "
                            f"rust=({cb.value!r},{cb.number_format})")
                        if len(diffs) >= max_report:
                            diffs.append("… (truncated)")
                            return diffs
    except Exception as exc:
        diffs.append(f"xlsx-diff-error: {exc}")
    return diffs


def log_shadow_diffs(job_id: str, combo_label: str, kind: str, diffs: List[str]) -> bool:
    """Emit shadow-diff results to the log. Returns True if clean (no diffs)."""
    tag = f"[RUST_SHADOW job={str(job_id)[:8]} combo={combo_label!r} {kind}]"
    if not diffs:
        logger.info("%s CLEAN", tag)
        return True
    logger.warning("%s %d diff(s):", tag, len(diffs))
    for d in diffs[:50]:
        logger.warning("%s   %s", tag, d)
    return False


# ── Shadow wiring — prove the ported Rust summary on real combos ─────────────

def run_shadow_summary_check(
    combo_id, combo_label, merged_payload, trades_df,
    base_summary, python_flat_summary, python_summary_pw,
    midcap_legs=None, midcap_spot_adjustment=None, midcap_symbol="NIFTYMIDCAP100",
    filter_segments=None,
) -> None:
    """Shadow mode: recompute the per-combo summary with the RUST engine
    (compute_optim_metrics + compute_summary_metrics, overall + patchwise) and diff it,
    key-by-key, against the authoritative Python summary. Read-only — logs diffs, never
    raises, never changes output. Only runs when OPTIMIZE_RUST_LOOP=shadow.

    This is the design's §5 shadow gate: run it across a real sweep and it exercises the
    ported summary on every strategy the corpus can't cover; a family graduates to
    Rust-authoritative only when this is clean.
    """
    if rust_loop_mode() != "shadow":
        return
    reason = rust_batch_unsupported(merged_payload)
    if reason is not None:
        logger.info("[RUST_SHADOW combo=%s] not owned by Rust yet (%s) — skipped", combo_id, reason)
        return
    try:
        import algotest_native  # type: ignore
        if trades_df is None or (hasattr(trades_df, "empty") and trades_df.empty):
            return
        records = trades_df.where(trades_df.notna(), None).to_dict("records")
        mbt, msumm = None, None
        if midcap_legs:
            from services.optimizer.excel_builder import compute_midcap_for_rows
            mbt, msumm, _has = compute_midcap_for_rows(
                records, midcap_legs, midcap_spot_adjustment, midcap_symbol or "NIFTYMIDCAP100")
            mbt = mbt or None

        opt_m = algotest_native.compute_optim_metrics(records, base_summary)
        merged_summary = {**base_summary, **opt_m}
        sm_over = algotest_native.compute_summary_metrics(records, merged_summary, False, filter_segments, mbt, msumm)
        rust_flat = {**merged_summary, **sm_over}
        log_shadow_diffs(combo_id, combo_label, "summary-overall",
                         diff_summary(python_flat_summary or {}, rust_flat))

        if python_summary_pw is not None:
            # python_summary_pw is the RAW _cmetrics(patchwise) output (not merged with
            # the base summary), so diff against the raw Rust _cmetrics output.
            sm_pw = algotest_native.compute_summary_metrics(records, merged_summary, True, filter_segments, mbt, msumm)
            log_shadow_diffs(combo_id, combo_label, "summary-patchwise",
                             diff_summary(python_summary_pw or {}, sm_pw))
    except Exception as exc:  # shadow is best-effort; never disturb the real run
        logger.warning("[RUST_SHADOW combo=%s] shadow check errored: %s", combo_id, exc)


def rust_authoritative_summary(
    trades_df, base_summary, merged_payload,
    midcap_legs=None, midcap_spot_adjustment=None, midcap_symbol="NIFTYMIDCAP100",
    filter_segments=None,
):
    """OPTIMIZE_RUST_LOOP=1: compute the per-combo summary ENTIRELY in Rust — the
    Rust-bounded, NO-Python-fallback path. The caller MUST first call
    require_rust_supported(merged_payload) (raises on any un-ported shape); this
    function does not fall back to Python for anything.

    Returns (flat_summary, summary_pw) exactly as the Python path would have built them:
      flat_summary = {**base_summary, **compute_optim_metrics, **compute_summary_metrics(overall)}
      summary_pw   = compute_summary_metrics(patchwise)   (raw, as _cmetrics returns it)
    """
    import algotest_native  # type: ignore
    if trades_df is None or (hasattr(trades_df, "empty") and trades_df.empty):
        return dict(base_summary or {}), None
    records = trades_df.where(trades_df.notna(), None).to_dict("records")
    mbt, msumm = None, None
    if midcap_legs:
        from services.optimizer.excel_builder import compute_midcap_for_rows
        mbt, msumm, _has = compute_midcap_for_rows(
            records, midcap_legs, midcap_spot_adjustment, midcap_symbol or "NIFTYMIDCAP100")
        mbt = mbt or None
    opt_m = algotest_native.compute_optim_metrics(records, base_summary or {})
    flat = {**(base_summary or {}), **opt_m}
    over = algotest_native.compute_summary_metrics(records, flat, False, filter_segments, mbt, msumm)
    flat = {**flat, **over}
    pw = algotest_native.compute_summary_metrics(records, flat, True, filter_segments, mbt, msumm)
    # The Rust engines emit CE/PE totals but NOT futures. compute_xlsx_summary_metrics
    # already supplements them (excel_builder.py, "if 'fut_pnl_total' not in _rust_sm"),
    # but this path calls algotest_native directly and so bypassed that patch — mode 1
    # therefore dropped fut_pnl_total / fut_pnl_pct from every combo's summary.
    # tools/optim_metrics_parity measured it: 34 keys, exactly these 2 diverging
    # (py=0.0 vs rust=<MISS>) on all 3 corpus payloads. Same formula as the original
    # (Sigma FUT P&L; pct = Sigma FUT P&L / Entry Spot * 100) so the two paths agree.
    # Purely additive — never overwrites a value Rust did produce.
    # Applied to BOTH dicts: the patchwise summary is a separate Rust call and
    # drops the same two keys, so patching only `flat` left download_mode=patchwise
    # (the default here) still missing them — caught by tools/mode1_parity.
    def _add_fut(d):
        if d is None or "fut_pnl_total" in d:
            return d
        _ft = _fp = 0.0
        for _r in records:
            _fu = _num(_r.get("FUT P&L"))
            if _fu is None:
                continue
            _ft += _fu
            _es = _num(_r.get("Entry Spot"))
            if _es not in (None, 0):
                _fp += _fu / _es
        d["fut_pnl_total"] = round(_ft, 2)
        d["fut_pnl_pct"] = round(_fp * 100, 4)
        return d

    return _add_fut(flat), _add_fut(pw)
