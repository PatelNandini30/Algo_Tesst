"""
WOW & MOM summary — Python port of the frontend `wowMom.js` + `wowMomSheet.js`.

Produces the same Week-on-Week / Month-on-Month tables the backtest and the
optimizer per-combo tradesheets show, so all three outputs match:
  • backtest export        (frontend ResultsPanel.jsx → wowMomSheet.js)
  • optim per-combo export (frontend buildTradeExcel.js → wowMomSheet.js,
                            and backend ZIP excel_builder.py → here)
  • optim merged summary   (backend → here, one stacked block per combo)

Computation mirrors wowMom.js exactly (verified to ~1e-15 and against the
research team's hand-corrected "Rectify" sheet). The WOW drawdown CONTINUES
across blank weeks (gaps are skipped, not breaks). Total / Max DD / R/MDD are
written as computed VALUES (the codebase convention) — identical displayed
numbers to the JS formula results.
"""
from __future__ import annotations

import json
import logging
import math
from functools import lru_cache
import re
from datetime import datetime
from statistics import mode
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# write_merged_wow_mom's column-pagination fallback (below) has called
# logger.warning(...) since it was written, but nothing in this module ever
# defined `logger` — so the first sweep wide enough to need a second sheet
# raised NameError and destroyed the very workbook the paging existed to
# rescue, instead of just logging that it split.
logger = logging.getLogger(__name__)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

WOW_RF = 0.06 / 52
MOM_RF = 0.06 / 12
WOW_NANN = 52
MOM_NANN = 12
SLOPE_CAP = 307


# ── date helpers ───────────────────────────────────────────────────────────
def _parse_date(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if v is None or v == "":
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d",
                "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _iso_year_week(d: datetime) -> Tuple[int, int]:
    iso = d.isocalendar()
    return int(iso[0]), int(iso[1])


def _first_thu_week(year: int, month: int) -> int:
    d = datetime(year, month, 1)
    dow = d.weekday()              # Mon=0
    shift = (3 - dow) % 7
    d = datetime.fromordinal(d.toordinal() + shift)
    return _iso_year_week(d)[1]


def get_month_maps(years: List[int], n_weeks: int) -> Tuple[Dict[int, int], Dict[int, int]]:
    sw: Dict[int, int] = {}
    for m in range(1, 13):
        weeks = [_first_thu_week(y, m) for y in years] if years else [1]
        try:
            sw[m] = mode(weeks)
        except Exception:
            sw[m] = weeks[0]
    ew: Dict[int, int] = {}
    for m in range(1, 12):
        ew[m] = sw[m + 1] - 1
    ew[12] = n_weeks
    return sw, ew


# ── ratio engine (port of compute_ratios) ──────────────────────────────────
def _std(arr: List[float], mean: Optional[float] = None) -> float:
    if not arr:
        return 0.0
    m = (sum(arr) / len(arr)) if mean is None else mean
    return math.sqrt(sum((x - m) ** 2 for x in arr) / len(arr))


def compute_ratios(returns: List[float], rf: float, n_ann: int) -> Optional[Dict[str, float]]:
    if len(returns) < 3:
        return None
    R = returns
    N = len(R)
    avg = sum(R) / N
    sd = _std(R, avg)
    pos = [v for v in R if v > 0]
    neg = [v for v in R if v < 0]
    wp = len(pos) / N
    wa = (sum(pos) / len(pos)) if pos else 0.0
    lp = len(neg) / N
    la = (sum(neg) / len(neg)) if neg else 0.0
    exp = ((wp / abs(la)) * wa - lp) if la else 0.0
    sd_neg = _std(neg) if neg else 0.0
    sh = ((avg - rf) / sd) * math.sqrt(n_ann) if sd else 0.0
    so = ((avg - rf) / sd_neg) * math.sqrt(n_ann) if sd_neg else 0.0
    sqn = (avg / sd) * math.sqrt(N) if sd else 0.0

    n_s = min(N, SLOPE_CAP)
    eq = [100.0]
    for v in R:
        eq.append(eq[-1] * (1 + v))
    Y = eq[1:n_s + 1]
    X = list(range(1, n_s + 1))
    mx = sum(X) / n_s
    my = sum(Y) / n_s
    sxy = sum((X[i] - mx) * (Y[i] - my) for i in range(n_s))
    sxx = sum((X[i] - mx) ** 2 for i in range(n_s))
    slope = (sxy / sxx) if sxx else 0.0
    intercept = my - slope * mx
    sse = sum((Y[i] - (slope * X[i] + intercept)) ** 2 for i in range(n_s))
    st_err = math.sqrt(max(sse / (n_s - 2), 1e-30))
    base = (slope * math.sqrt(sxx)) / st_err
    k1 = base / math.sqrt(N)
    k2 = base / N
    k3 = (base * math.sqrt(252)) / N

    c = 1.0
    for v in R:
        c *= (1 + v)
    cg = c ** (n_ann / N) - 1
    return dict(wp=wp, wa=wa, lp=lp, la=la, exp=exp, sh=sh, so=so,
                k1=k1, k2=k2, k3=k3, sqn=sqn, cg=cg, n=N)


def wow_year_drawdown(week_map: Dict[int, float], nw: int) -> Dict[str, Optional[float]]:
    """Running cumulative drawdown over present weeks; gaps are SKIPPED, not breaks."""
    vals = [(w, week_map[w]) for w in range(1, nw + 1) if week_map.get(w) is not None]
    if not vals:
        return {"maxdd": None, "start": None, "end": None}
    g_min = 0.0
    best: Tuple[Optional[int], Optional[int]] = (None, None)
    cum = 0.0
    started = False
    s_start = None
    s_min = 0.0
    s_mc = None
    for col, val in vals:
        if not started:
            if val < 0:
                started, s_start, cum, s_min, s_mc = True, col, val, val, col
        else:
            cum += val
            if cum >= 0:
                if s_min < g_min:
                    g_min, best = s_min, (s_start, s_mc)
                started, cum, s_start, s_min, s_mc = False, 0.0, None, 0.0, None
            elif cum < s_min:
                s_min, s_mc = cum, col
    if started and s_min < g_min:
        g_min, best = s_min, (s_start, s_mc)
    sc, ec = best
    if sc is None:
        return {"maxdd": None, "start": None, "end": None}
    return {"maxdd": round(g_min, 4), "start": sc, "end": ec}


# ── bucketing (port of buildWowMom) ────────────────────────────────────────
def _to_num(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        f = float(str(v).replace(",", "").replace("%", "").replace("₹", "").strip())
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None


def build_wow_mom(cleaned: List[Dict], ret_field: str, dd_field: str,
                  dd_is_percent: bool, live_field: Optional[str] = None,
                  yearly: bool = False, mae_field: Optional[str] = None) -> Dict[str, Any]:
    wow: Dict[int, Dict[int, float]] = {}
    mom_monthly: Dict[int, Dict[int, float]] = {}
    mom_dd: Dict[int, List[float]] = {}
    # MIN-of-Final-MAE / MIN-of-Actual-Live-DD grids rendered below each block.
    # Same bucketing as the block above them (WOW by Expiry week, MOM by Exit
    # month) so the columns line up, but aggregated with MIN and kept in RAW
    # percent points — these fields are already percent points, and the source
    # workbook shows them unscaled (e.g. -1.7526), unlike the Live DD % column
    # in the block header which divides by 100.
    wow_mae: Dict[int, Dict[int, float]] = {}
    wow_ldd: Dict[int, Dict[int, float]] = {}
    mom_mae: Dict[int, Dict[int, float]] = {}
    mom_ldd: Dict[int, Dict[int, float]] = {}

    def _keep_min(store: Dict[int, Dict[int, float]], y: int, k: int, v: Optional[float]) -> None:
        if v is None:
            return
        cur = store.setdefault(y, {}).get(k)
        if cur is None or v < cur:
            store[y][k] = v
    # Live DD % — MIN of "Actual Live DD" (percent points → /100) per year.
    # WOW is bucketed by Expiry year, MOM by Exit year (to match each block's
    # year rows). Written between Max DD and R/MDD; R/MDD still uses Max DD.
    wow_live: Dict[int, List[float]] = {}
    mom_live: Dict[int, List[float]] = {}
    n_trades = 0

    for t in cleaned:
        raw = t.get(ret_field)
        if raw == "" or raw is None:
            continue
        ret = _to_num(raw)
        if ret is None:
            continue
        n_trades += 1
        dec = ret / 100.0

        live_dec = None
        live_raw = None
        if live_field:
            live_num = _to_num(t.get(live_field))
            if live_num is not None:
                live_dec = live_num / 100.0
                live_raw = live_num
        mae_raw = _to_num(t.get(mae_field)) if mae_field else None

        # WOW week identity.
        #
        # Normally the Expiry: for a weekly/monthly strategy the trade IS its
        # contract, so Expiry is the trade's natural week — and it deliberately
        # keeps a contract's P&L in ONE week even when a T-n exit lands in the
        # previous calendar week.
        #
        # Under YEARLY that same principle breaks it: the contract IS the whole
        # year, so every trade shares one Expiry and the entire year collapses
        # into a single cell (2019-12-26 → ISO week 52; 2020-12-31 → week 53).
        # There the roll segment is the week, so the week identity comes from the
        # Exit Date — mirroring what MOM already does below.
        # SAME-INDEX MIXED EXPIRY: with a weekly cadence and a pinned monthly leg,
        # the two legs of ONE trade carry different Expiry values, so bucketing on
        # Expiry splits a single trade across weeks (and which week wins depends on
        # leg order). "Cadence Expiry" is the trade's shared weekly contract, so
        # both legs land together. It is emitted for every row and equals Expiry
        # unless a leg is genuinely pinned, so non-mixed WOW output is unchanged.
        _wk_src = (
            t.get("Exit Date") if yearly
            else (t.get("Cadence Expiry") or t.get("Expiry"))
        )
        e = _parse_date(_wk_src)
        if e is not None:
            y, w = _iso_year_week(e)
            wow.setdefault(y, {})
            wow[y][w] = wow[y].get(w, 0.0) + dec
            if live_dec is not None:
                wow_live.setdefault(y, []).append(live_dec)
            _keep_min(wow_mae, y, w, mae_raw)
            _keep_min(wow_ldd, y, w, live_raw)

        x = _parse_date(t.get("Exit Date"))
        if x is not None:
            y = x.year
            mi = x.month - 1
            mom_monthly.setdefault(y, {})
            mom_monthly[y][mi] = mom_monthly[y].get(mi, 0.0) + dec
            dd_raw = t.get(dd_field)
            dd_num = _to_num(dd_raw)
            if dd_num is not None:
                dd_dec = (dd_num / 100.0) if dd_is_percent else dd_num
                mom_dd.setdefault(y, []).append(dd_dec)
            if live_dec is not None:
                mom_live.setdefault(y, []).append(live_dec)
            _keep_min(mom_mae, y, mi, mae_raw)
            _keep_min(mom_ldd, y, mi, live_raw)

    for y in list(wow.keys()):
        for w in list(wow[y].keys()):
            if abs(wow[y][w]) <= 1e-9:
                del wow[y][w]
        if not wow[y]:
            del wow[y]

    mom: Dict[int, Dict] = {}
    for y in mom_monthly:
        months = {}
        for mi in range(12):
            v = mom_monthly[y].get(mi)
            if v is not None and abs(v) > 1e-9:
                months[MONTHS[mi]] = v
        if not months:
            continue
        total = sum(months.values())
        dd_arr = mom_dd.get(y, [])
        if dd_arr:
            maxdd = min(dd_arr)
        else:
            # running cumulative DD over the monthly series (fallback)
            cum = 0.0
            mdd = 0.0
            for mn in MONTHS:
                cum += months.get(mn, 0.0)
                if cum > 0:
                    cum = 0.0
                mdd = min(mdd, cum)
            maxdd = mdd
        live_arr = mom_live.get(y, [])
        livedd = min(live_arr) if live_arr else None
        mom[y] = {"months": months, "total": total, "maxdd": maxdd, "livedd": livedd}

    wow_live_min = {y: (min(v) if v else None) for y, v in wow_live.items()}
    wow_years = sorted(wow.keys())
    mom_years = sorted(mom.keys())
    return {"wow": wow, "mom": mom, "wow_years": wow_years,
            "mom_years": mom_years, "n_weeks": 53, "n_trades": n_trades,
            "wow_live": wow_live_min,
            "wow_mae": wow_mae, "wow_ldd": wow_ldd,
            "mom_mae": mom_mae, "mom_ldd": mom_ldd}


def flat_weekly(wow: Dict[int, Dict[int, float]]) -> List[float]:
    return [wow[y][w] for y in wow for w in wow[y]]


def flat_monthly(mom: Dict[int, Dict]) -> List[float]:
    return [mom[y]["months"][k] for y in mom for k in mom[y]["months"]]


# ── title (port of buildWowMomTitle) ───────────────────────────────────────
def _fmt_num(v: float) -> str:
    return str(int(v)) if float(v).is_integer() else str(round(float(v), 2))


def build_wow_mom_title(config: Optional[Dict]) -> str:
    if not config:
        return "Strategy"
    legs = config.get("legs") or []
    opt_leg = None
    for l in legs:
        if l.get("segment") not in ("midcap100", "futures") and l.get("option_type"):
            opt_leg = l
            break
    if opt_leg is None and legs:
        opt_leg = legs[0]
    cepe, strike = "", ""
    if opt_leg:
        o = str(opt_leg.get("option_type") or "").lower()
        cepe = "CE" if o == "call" else "PE" if o == "put" else (o.upper() if o else "")
        criteria = opt_leg.get("strike_criteria") or "strike_type"
        if criteria == "pct_of_atm":
            pct_val = _to_num(opt_leg.get("pct_value")) or 0
            if not pct_val:
                strike = "ATM"
            else:
                moneyness = str(opt_leg.get("pct_atm_moneyness") or "OTM").upper()
                strike = f"{_fmt_num(pct_val)}% {moneyness}"
        elif criteria == "atm_straddle_prem_pct":
            strike = "STRADDLE"
        else:
            strike = str(opt_leg.get("strike_type") or "ATM").upper()
    adj = "No Adj"
    if config.get("spotAdjustmentEnabled"):
        d = config.get("spotAdjustmentDirection") or "rise"
        val = config.get("spotAdjustmentValue")
        unit = "%" if (config.get("spotAdjustmentUnits") or "percent") == "percent" else "pts"
        val_str = _fmt_num(val) if val is not None else ""
        word = "Rise or Fall" if d == "both" else "Fall" if d == "fall" else "Rise"
        adj = f"{word}{(' ' + val_str + unit) if val_str else ''}"
    left = " ".join(x for x in (cepe, strike) if x) or "Strategy"
    return f"{left} | {adj}"


# ── styling ────────────────────────────────────────────────────────────────
_NAVY_BG, _NAVY_TX = "FF1F3864", "FFFFFFFF"
_SECTION_BG = "FF2C5F8A"
_HEADER_BG, _HEADER_TX = "FF34495E", "FFFFFFFF"
_SUB_BG, _SUB_TX = "FFD6E4F7", "FF1F3864"
_GREEN_BG, _GREEN_TX = "FFD4EFDF", "FF1E7E34"
_RED_BG, _RED_TX = "FFFDE8E8", "FFC0392B"
_LABEL_BG = "FFF2F6FA"
_BORDER_CLR = "FFB0C4D8"
_BLACK = "FF000000"

PCT_FMT, RAT_FMT, K_FMT, INT_FMT = "0.00%", "0.00", "0.0000", "0"
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

STAT_LBL = ["Win %", "Win avg", "Loss %", "Loss Avg", "Expectancy",
            "No. of Trades", "Sharpe", "Sortino", "K1", "K2", "K3", "SQN", "CAGR"]
STAT_FMT = [PCT_FMT, PCT_FMT, PCT_FMT, PCT_FMT, K_FMT, INT_FMT,
            RAT_FMT, RAT_FMT, K_FMT, K_FMT, K_FMT, RAT_FMT, PCT_FMT]


# Style objects are IMMUTABLE and shareable in openpyxl, but building them is
# expensive — and _Sheet.h/v built a fresh Font + PatternFill + Border for EVERY
# cell. On a merged grid that is millions of allocations: profiling the WOW/MOM
# pre-build showed 46.0s of 47.9s (96%) in cell writing versus 1.8s in save().
# Caching by value hands out the same object for the same style, which is what
# openpyxl serializes anyway, so the workbook is byte-for-byte unchanged.
@lru_cache(maxsize=None)
def _thin(color: str = _BORDER_CLR) -> Border:
    s = Side(style="thin", color=color)
    return Border(top=s, left=s, bottom=s, right=s)


@lru_cache(maxsize=None)
def _cached_font(bold: bool, size, color: str) -> Font:
    return Font(bold=bold, size=size, color=color, name="Calibri")


@lru_cache(maxsize=None)
def _cached_fill(bg: str) -> PatternFill:
    return PatternFill("solid", fgColor=bg)


def _sign_fill(v):
    if v is None:
        return None
    return _GREEN_BG if v >= 0 else _RED_BG


def _sign_tx(v):
    if v is None:
        return _BLACK
    return _GREEN_TX if v >= 0 else _RED_TX


class _Sheet:
    """Thin cell helper bound to a worksheet (mirrors hCell/vCell in JS)."""

    def __init__(self, ws):
        self.ws = ws

    def h(self, r, c, val, *, size=9, tx=_HEADER_TX, bg=_HEADER_BG, align=None):
        cell = self.ws.cell(r, c, val)
        cell.font = _cached_font(True, size, tx)
        cell.fill = _cached_fill(bg)
        cell.alignment = align or _CENTER
        cell.border = _thin()
        return cell

    def v(self, r, c, val, fmt=None, *, size=9, tx=None, bg=None):
        cell = self.ws.cell(r, c)
        cell.value = "" if val is None else val
        cell.font = _cached_font(False, size, tx or _BLACK)
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt
        cell.alignment = _CENTER
        cell.border = _thin()
        if bg:
            cell.fill = _cached_fill(bg)
        return cell


def _stat_vals(m, n):
    if m:
        return [m["wp"], m["wa"], m["lp"], m["la"], m["exp"], n,
                m["sh"], m["so"], m["k1"], m["k2"], m["k3"], m["sqn"], m["cg"]]
    return [None, None, None, None, None, n, None, None, None, None, None, None, None]


def _avg(a):
    return (sum(a) / len(a)) if a else None


def _write_stat_header(S, wm, title, base_row, title_merge_cols, stat_col0, stats, cc=0):
    S.h(base_row, cc + 1, title, size=10, tx=_NAVY_TX, bg=_NAVY_BG, align=_LEFT)
    S.ws.merge_cells(start_row=base_row, start_column=cc + 1, end_row=base_row, end_column=cc + title_merge_cols)
    S.h(base_row + 1, cc + 1, "", bg=_NAVY_BG)
    S.ws.merge_cells(start_row=base_row + 1, start_column=cc + 1, end_row=base_row + 1, end_column=cc + title_merge_cols)
    vals = _stat_vals(stats, wm["n_trades"])
    for i, lbl in enumerate(STAT_LBL):
        S.h(base_row, cc + stat_col0 + i, lbl, bg=_SECTION_BG)
    for i, val in enumerate(vals):
        S.v(base_row + 1, cc + stat_col0 + i, val, STAT_FMT[i], bg=_LABEL_BG)


def _write_wow_block(S, wm, title, base_row: int, base_col: int = 1) -> int:
    """Write one WOW block at (base_row, base_col). Returns the Total-row index.

    Columns (relative to base_col): Year | W1..Wn | Total | Max DD | Live DD % | R/MDD.
    """
    cc = base_col - 1
    nw = wm["n_weeks"]
    w_tot, w_mdd, w_live, w_rmdd = 2 + nw, 3 + nw, 4 + nw, 5 + nw
    wow_live = wm.get("wow_live", {})
    sw, ew = get_month_maps(wm["wow_years"], nw)
    month_ends = set(ew.values())
    blk = Side(style="medium", color=_BLACK)

    def month_edge(cell, w):
        if w in month_ends:
            cell.border = Border(top=cell.border.top, left=cell.border.left,
                                 bottom=cell.border.bottom, right=blk)

    _write_stat_header(S, wm, title, base_row, 5, 7,
                       compute_ratios(flat_weekly(wm["wow"]), WOW_RF, WOW_NANN), cc=cc)
    w_hdr, w_month, w_data0 = base_row + 2, base_row + 3, base_row + 4

    S.h(w_hdr, cc + 1, "Year")
    for w in range(1, nw + 1):
        c = S.h(w_hdr, cc + 1 + w, f"W{w}", size=7)
        month_edge(c, w)
    for col, h in ((w_tot, "Total"), (w_mdd, "Max DD"), (w_live, "Live DD %"), (w_rmdd, "R/MDD")):
        S.h(w_hdr, cc + col, h)
        S.h(w_month, cc + col, "")
        S.ws.merge_cells(start_row=w_hdr, start_column=cc + col, end_row=w_month, end_column=cc + col)

    S.h(w_month, cc + 1, "Month", size=8, tx=_SUB_TX, bg=_SUB_BG)
    for w in range(1, nw + 1):
        c = S.h(w_month, cc + 1 + w, "", bg=_SUB_BG)
        month_edge(c, w)
    for mi, mn in enumerate(MONTHS):
        start_w = sw.get(mi + 1, 1)
        if 1 <= start_w <= nw:
            cell = S.ws.cell(w_month, cc + 1 + start_w, mn)
            cell.font = _cached_font(True, 8, _SUB_TX)
            cell.alignment = _CENTER

    r = w_data0
    w_tots, w_mdds, w_rmdds, w_lives = [], [], [], []
    for yr in wm["wow_years"]:
        S.h(r, cc + 1, yr, tx=_SUB_TX, bg=_LABEL_BG)
        tot, cnt = 0.0, 0
        yd = wm["wow"].get(yr, {})
        for w in range(1, nw + 1):
            val = yd.get(w)
            if val is not None:
                cell = S.v(r, cc + 1 + w, round(val, 6), PCT_FMT, size=7, bg=_sign_fill(val), tx=_sign_tx(val))
                tot += val
                cnt += 1
            else:
                cell = S.v(r, cc + 1 + w, "", PCT_FMT, size=7)
            month_edge(cell, w)
        S.v(r, cc + w_tot, round(tot, 6) if cnt else "", PCT_FMT,
            bg=_LABEL_BG, tx=_sign_tx(tot if cnt else None))

        dd = wow_year_drawdown(yd, nw)
        S.v(r, cc + w_mdd, dd["maxdd"] if dd["maxdd"] is not None else "", PCT_FMT, bg=_RED_BG, tx=_RED_TX)
        live = wow_live.get(yr)
        S.v(r, cc + w_live, round(live, 6) if live is not None else "", PCT_FMT,
            bg=_RED_BG if (live is not None and live < 0) else _LABEL_BG, tx=_sign_tx(live))
        rmdd = (tot / abs(dd["maxdd"])) if (cnt and dd["maxdd"]) else None
        S.v(r, cc + w_rmdd, round(rmdd, 2) if rmdd is not None else "", RAT_FMT,
            bg=_sign_fill(rmdd), tx=_sign_tx(rmdd))
        if cnt:
            w_tots.append(tot)
        if dd["maxdd"] is not None:
            w_mdds.append(dd["maxdd"])
        if rmdd is not None:
            w_rmdds.append(rmdd)
        if live is not None:
            w_lives.append(live)
        if dd["start"] is not None:
            for w in range(dd["start"], dd["end"] + 1):
                cell = S.ws.cell(r, cc + 1 + w)
                cell.border = Border(
                    top=blk, bottom=blk,
                    left=blk if w == dd["start"] else None,
                    right=blk if w == dd["end"] else None,
                )
        r += 1

    data_r1 = r - 1
    S.h(r, cc + 1, "Total")
    if data_r1 >= w_data0:
        t_sum = sum(w_tots)
        S.v(r, cc + w_tot, round(t_sum, 6), PCT_FMT, bg=_sign_fill(t_sum), tx=_sign_tx(t_sum))
        am = _avg(w_mdds) or 0
        S.v(r, cc + w_mdd, round(am, 6), PCT_FMT, bg=_RED_BG, tx=_RED_TX)
        al = _avg(w_lives)
        S.v(r, cc + w_live, round(al, 6) if al is not None else "", PCT_FMT,
            bg=_RED_BG if (al is not None and al < 0) else _LABEL_BG, tx=_sign_tx(al))
        ar = _avg(w_rmdds)
        S.v(r, cc + w_rmdd, round(ar or 0, 2), RAT_FMT, bg=_sign_fill(ar), tx=_sign_tx(ar))
    return r


def _write_mom_block(S, wm, title, base_row: int, base_col: int = 1) -> int:
    """Columns (relative to base_col): Year | Jan..Dec | Total | Max DD | Live DD % | R/MDD."""
    cc = base_col - 1
    _write_stat_header(S, wm, title, base_row, 4, 6,
                       compute_ratios(flat_monthly(wm["mom"]), MOM_RF, MOM_NANN), cc=cc)
    m_hdr, m_data0 = base_row + 2, base_row + 3
    for i, h in enumerate(["Year"] + MONTHS + ["Total", "Max DD", "Live DD %", "R/MDD"]):
        S.h(m_hdr, cc + 1 + i, h, size=8)

    r = m_data0
    for yr in wm["mom_years"]:
        yd = wm["mom"].get(yr, {"months": {}, "total": None, "maxdd": None, "livedd": None})
        S.h(r, cc + 1, yr, tx=_SUB_TX, bg=_LABEL_BG)
        months = yd.get("months", {})
        for mi, mn in enumerate(MONTHS):
            val = months.get(mn)
            if val is not None:
                S.v(r, cc + 2 + mi, round(val, 6), PCT_FMT, bg=_sign_fill(val), tx=_sign_tx(val))
            else:
                S.v(r, cc + 2 + mi, "", PCT_FMT)
        tot, mdd, live = yd.get("total"), yd.get("maxdd"), yd.get("livedd")
        S.v(r, cc + 14, round(tot, 6) if tot is not None else "", PCT_FMT,
            bg=_sign_fill(tot), tx=_sign_tx(tot))
        S.v(r, cc + 15, round(mdd, 6) if mdd is not None else "", PCT_FMT, bg=_RED_BG, tx=_RED_TX)
        S.v(r, cc + 16, round(live, 6) if live is not None else "", PCT_FMT,
            bg=_RED_BG if (live is not None and live < 0) else _LABEL_BG, tx=_sign_tx(live))
        rmdd = (tot / abs(mdd)) if (tot is not None and mdd) else None
        S.v(r, cc + 17, round(rmdd, 2) if rmdd is not None else "", RAT_FMT,
            bg=_sign_fill(rmdd), tx=_sign_tx(rmdd))
        r += 1

    data_r1 = r - 1
    S.h(r, cc + 1, "Total")
    for mi in range(12):
        S.h(r, cc + 2 + mi, "")
    if data_r1 >= m_data0:
        tots = [wm["mom"].get(y, {})["total"] for y in wm["mom_years"] if wm["mom"].get(y, {}).get("total") is not None]
        mdds = [wm["mom"].get(y, {})["maxdd"] for y in wm["mom_years"] if wm["mom"].get(y, {}).get("maxdd") is not None]
        lives = [wm["mom"].get(y, {})["livedd"] for y in wm["mom_years"] if wm["mom"].get(y, {}).get("livedd") is not None]
        rmdds = []
        for y in wm["mom_years"]:
            t, d = wm["mom"].get(y, {}).get("total"), wm["mom"].get(y, {}).get("maxdd")
            if t is not None and d:
                rmdds.append(t / abs(d))
        n_sum = sum(tots)
        S.v(r, cc + 14, round(n_sum, 6), PCT_FMT, bg=_sign_fill(n_sum), tx=_sign_tx(n_sum))
        am = _avg(mdds) or 0
        S.v(r, cc + 15, round(am, 6), PCT_FMT, bg=_RED_BG, tx=_RED_TX)
        al = _avg(lives)
        S.v(r, cc + 16, round(al, 6) if al is not None else "", PCT_FMT,
            bg=_RED_BG if (al is not None and al < 0) else _LABEL_BG, tx=_sign_tx(al))
        ar = _avg(rmdds)
        S.v(r, cc + 17, round(ar or 0, 2), RAT_FMT, bg=_sign_fill(ar), tx=_sign_tx(ar))
    return r


def _min_grid_fields(cleaned: List[Dict]) -> Tuple[str, str, str, str]:
    """(mae_field, live_field, mae_title, live_title) for the MIN grids.

    Resolved by PRESENCE, not by the has_midcap flag: a run can carry Combined
    columns from a Midcap overlay, a MIDCPNIFTY overlay, both, or neither, and
    the grids must work in all four cases. Falling back to the plain column when
    no Combined one exists also keeps the title honest about which it summed.
    """
    keys: set = set()
    for t in cleaned:
        if t:
            keys = set(t.keys())
            break
    mae = "Combined Final MAE" if "Combined Final MAE" in keys else "Final MAE"
    live = ("Combined Actual Live DD" if "Combined Actual Live DD" in keys
            else "Actual Live DD")
    return mae, live, f"Min of {mae}", f"Min of {live}"


def _wm_from_cleaned(cleaned: List[Dict], has_midcap: bool, yearly: bool = False) -> Dict[str, Any]:
    ret_field = "Combined Net P&L %" if has_midcap else "% P&L"
    dd_field = "Combined %DD" if has_midcap else "%DD"
    live_field = "Combined Actual Live DD" if has_midcap else "Actual Live DD"
    mae_field, min_live_field, mae_title, live_title = _min_grid_fields(cleaned)
    wm = build_wow_mom(cleaned, ret_field, dd_field, has_midcap,
                       live_field=live_field, yearly=yearly, mae_field=mae_field)
    # The MIN grids read Live DD from the same column the block header does, but
    # unscaled — resolve it independently so a run with no Combined columns still
    # gets a grid instead of a blank one.
    if min_live_field != live_field:
        wm2 = build_wow_mom(cleaned, ret_field, dd_field, has_midcap,
                            live_field=min_live_field, yearly=yearly,
                            mae_field=mae_field)
        wm["wow_ldd"], wm["mom_ldd"] = wm2["wow_ldd"], wm2["mom_ldd"]
    wm["mae_title"], wm["live_title"] = mae_title, live_title
    return wm


# ── MIN-of-MAE / MIN-of-Live-DD grids ──────────────────────────────────────
# Two per axis, written directly beneath their own block so the columns line up
# with the block above. Values are RAW percent points (the source columns
# already are), aggregated with MIN — never summed or averaged.
# Display 2dp, store the FULL float — a formula on the cell must see every digit,
# so the precision lives in the value and only the number_format is truncated.
MIN_GRID_FMT = "0.00"
MIN_GRID_GAP = 4          # => 3 blank rows of margin after EVERY table, incl.
                          #    before the next block (was uneven: 2 then 1)
MIN_WEEKLY_ROWS = 4       # title + Month band + Year header + Grand Total
MIN_MONTHLY_ROWS = 3      # title + column header + Grand Total

# Everything below reuses the sheet's OWN theme constants — the grids sat next to
# the WOW/MOM blocks in a separate invented blue/pink palette and clashed. Header
# rows mirror the blocks exactly: _HEADER_BG title, _SUB_BG column header,
# _LABEL_BG row labels.
_GRID_TITLE_BG, _GRID_TITLE_TX = _HEADER_BG, _HEADER_TX
_GRID_BAND_BG, _GRID_BAND_TX = _SUB_BG, _SUB_TX
_GRID_HDR_BG, _GRID_HDR_TX = _SUB_BG, _SUB_TX
_GRID_YEAR_BG = _LABEL_BG


def _blend(light: str, dark: str, t: float) -> str:
    """Mix two ARGB theme colours; t=0 -> light, t=1 -> dark."""
    lr, lg, lb = int(light[2:4], 16), int(light[4:6], 16), int(light[6:8], 16)
    dr, dg, db = int(dark[2:4], 16), int(dark[4:6], 16), int(dark[6:8], 16)
    mix = tuple(round(l + (d - l) * t) for l, d in ((lr, dr), (lg, dg), (lb, db)))
    return "FF" + "".join(f"{v:02X}" for v in mix)


# Value ramp — design notes, because the obvious choice reads badly here:
#
# MAE and Live DD are almost always negative, so SIGN carries no information;
# magnitude does. A saturated red on every cell therefore just shouts, and
# blending toward the theme's brick red (_RED_TX, an orange-leaning FFC0392B)
# produced a salmon that fought the cool slate/blue headers.
#
# So: one sequential ramp in a COOL rose/crimson, which sits in the same visual
# family as the slate navy header, and stays light for typical values so only a
# genuinely large drawdown darkens. Positives are rare here, so their ramp is
# deliberately muted — they should register without grabbing the eye.
_RAMP_NEG_LIGHT, _RAMP_NEG_DARK = "FFFEF6F7", "FF9B2242"   # near-white -> crimson
_RAMP_POS_LIGHT, _RAMP_POS_DARK = "FFF7FBF8", "FF2C6E4A"    # near-white -> forest
# Kept in the LIGHTEST band of the hue: the whole ramp travels only ~a third of
# the way toward the deep tone, so even the worst cell stays a pale tint. With
# almost every value negative, a scale that reaches saturated red turns the
# sheet into a wall of alarm — magnitude still reads, just gently.
_RAMP_STEPS = (0.0, 0.10, 0.20, 0.32)
_RAMP_NEG = tuple(_blend(_RAMP_NEG_LIGHT, _RAMP_NEG_DARK, t) for t in _RAMP_STEPS)
_RAMP_POS = tuple(_blend(_RAMP_POS_LIGHT, _RAMP_POS_DARK, t) for t in _RAMP_STEPS)
_RAMP_NEG_TX, _RAMP_POS_TX = "FF8E2A3F", "FF2C6E4A"
# Every step stays pale, so dark text reads on all of them — never invert.
_RAMP_INVERT_FROM = len(_RAMP_STEPS)


def _ramp_step(v, worst_neg, best_pos):
    """(band, index) for `v`, scaled against this grid's own extremes."""
    if v is None or v == "":
        return None, 0
    if v < 0:
        span, band = abs(worst_neg or 0), _RAMP_NEG
    elif v > 0:
        span, band = abs(best_pos or 0), _RAMP_POS
    else:
        return _RAMP_NEG, 0
    frac = (abs(v) / span) if span else 1.0
    return band, min(int(frac * len(band)), len(band) - 1)


def _ramp_fill(v, worst_neg, best_pos):
    band, idx = _ramp_step(v, worst_neg, best_pos)
    return band[idx] if band else None


def _ramp_tx(v, worst_neg=None, best_pos=None):
    """Theme text colour, flipped to white once the fill goes too dark to read."""
    if v is None or v == "":
        return _BLACK
    band, idx = _ramp_step(v, worst_neg, best_pos)
    if band and idx >= _RAMP_INVERT_FROM:
        return "FFFFFFFF"
    return _RAMP_NEG_TX if v < 0 else (_RAMP_POS_TX if v > 0 else _BLACK)


def _grid_extremes(data, years):
    """(most negative, most positive) across the whole grid, for the ramp scale."""
    vals = [v for y in years for v in (data.get(y) or {}).values()
            if isinstance(v, (int, float))]
    if not vals:
        return None, None
    return min(vals), max(vals)


def _grid_title(S, row, col, span, text):
    """Merged full-width title band so long names aren't clipped by column A."""
    cell = S.h(row, col, text, tx=_GRID_TITLE_TX, bg=_GRID_TITLE_BG,
               align=Alignment(horizontal="left", vertical="center", indent=1))
    for c in range(col + 1, col + span):
        S.h(row, c, "", tx=_GRID_TITLE_TX, bg=_GRID_TITLE_BG)
    S.ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return cell


def _min_grid_weekly_height(n_years: int) -> int:
    """Rows one weekly MIN grid occupies, including its leading gap."""
    return MIN_GRID_GAP + MIN_WEEKLY_ROWS + n_years


def _min_grid_monthly_height(n_years: int) -> int:
    return MIN_GRID_GAP + MIN_MONTHLY_ROWS + n_years


def _write_min_weekly_grid(S, wm, title: str, key: str,
                           base_row: int, base_col: int = 1) -> int:
    """Year x W1..Wn MIN grid, sharing the WOW block's week columns + Month band.

    Returns the Grand-Total row index.
    """
    cc = base_col - 1
    nw = wm["n_weeks"]
    years = wm["wow_years"]
    data = wm.get(key) or {}
    gt_col = 2 + nw
    sw, ew = get_month_maps(years, nw)
    month_ends = set(ew.values())
    blk = Side(style="medium", color=_BLACK)

    def month_edge(cell, w):
        if w in month_ends:
            cell.border = Border(top=cell.border.top, left=cell.border.left,
                                 bottom=cell.border.bottom, right=blk)

    lo, hi = _grid_extremes(data, years)
    _grid_title(S, base_row, cc + 1, gt_col, title)

    r_month, r_hdr = base_row + 1, base_row + 2
    S.h(r_month, cc + 1, "Month", size=8, tx=_GRID_BAND_TX, bg=_GRID_BAND_BG)
    for w in range(1, nw + 1):
        month_edge(S.h(r_month, cc + 1 + w, "", tx=_GRID_BAND_TX, bg=_GRID_BAND_BG), w)
    S.h(r_month, cc + gt_col, "", tx=_GRID_BAND_TX, bg=_GRID_BAND_BG)
    for mi, mn in enumerate(MONTHS):
        start_w = sw.get(mi + 1, 1)
        if 1 <= start_w <= nw:
            cell = S.ws.cell(r_month, cc + 1 + start_w, mn)
            cell.font = _cached_font(True, 8, _GRID_BAND_TX)
            cell.alignment = _CENTER

    S.h(r_hdr, cc + 1, "Year", size=8, tx=_GRID_HDR_TX, bg=_GRID_HDR_BG)
    for w in range(1, nw + 1):
        month_edge(S.h(r_hdr, cc + 1 + w, f"W{w}", size=7,
                       tx=_GRID_HDR_TX, bg=_GRID_HDR_BG), w)
    S.h(r_hdr, cc + gt_col, "Grand Total", size=8, tx=_GRID_HDR_TX, bg=_GRID_HDR_BG)

    r = r_hdr + 1
    for yr in years:
        S.h(r, cc + 1, yr, size=8, tx=_GRID_HDR_TX, bg=_GRID_YEAR_BG)
        yd = data.get(yr) or {}
        for w in range(1, nw + 1):
            v = yd.get(w)
            cell = S.v(r, cc + 1 + w, v if v is not None else "", MIN_GRID_FMT,
                       size=7, bg=_ramp_fill(v, lo, hi), tx=_ramp_tx(v, lo, hi))
            month_edge(cell, w)
        row_min = min(yd.values()) if yd else None
        S.v(r, cc + gt_col, row_min if row_min is not None else "", MIN_GRID_FMT,
            bg=_ramp_fill(row_min, lo, hi), tx=_ramp_tx(row_min, lo, hi))
        r += 1

    S.h(r, cc + 1, "Grand Total", size=8, tx=_GRID_HDR_TX, bg=_GRID_HDR_BG)
    all_vals = []
    for w in range(1, nw + 1):
        col_vals = [data[y][w] for y in years if w in (data.get(y) or {})]
        cv = min(col_vals) if col_vals else None
        month_edge(S.v(r, cc + 1 + w, cv if cv is not None else "", MIN_GRID_FMT,
                       size=7, bg=_ramp_fill(cv, lo, hi), tx=_ramp_tx(cv, lo, hi)), w)
        all_vals.extend(col_vals)
    gv = min(all_vals) if all_vals else None
    S.v(r, cc + gt_col, gv if gv is not None else "", MIN_GRID_FMT,
        bg=_ramp_fill(gv, lo, hi), tx=_ramp_tx(gv, lo, hi))
    return r


def _write_min_monthly_grid(S, wm, title: str, key: str,
                            base_row: int, base_col: int = 1) -> int:
    """Row Labels x Jan..Dec + Grand Total MIN grid, under the MOM block."""
    cc = base_col - 1
    years = wm["mom_years"]
    data = wm.get(key) or {}
    gt_col = 14                       # Year + 12 months + Grand Total

    lo, hi = _grid_extremes(data, years)
    # Title spans the grid: column A alone clipped "Min of Combined Actual Live
    # DD", and the old "Column Labels" cell next to it just added noise.
    _grid_title(S, base_row, cc + 1, gt_col, title)

    r_hdr = base_row + 1
    S.h(r_hdr, cc + 1, "Year", size=8, tx=_GRID_BAND_TX, bg=_GRID_BAND_BG)
    for mi, mn in enumerate(MONTHS):
        S.h(r_hdr, cc + 2 + mi, mn, size=8, tx=_GRID_BAND_TX, bg=_GRID_BAND_BG)
    S.h(r_hdr, cc + gt_col, "Grand Total", size=8,
        tx=_GRID_BAND_TX, bg=_GRID_BAND_BG)

    r = r_hdr + 1
    for yr in years:
        S.h(r, cc + 1, yr, size=8, tx=_GRID_HDR_TX, bg=_GRID_YEAR_BG)
        yd = data.get(yr) or {}
        for mi in range(12):
            v = yd.get(mi)
            S.v(r, cc + 2 + mi, v if v is not None else "", MIN_GRID_FMT,
                bg=_ramp_fill(v, lo, hi), tx=_ramp_tx(v, lo, hi))
        row_min = min(yd.values()) if yd else None
        S.v(r, cc + gt_col, row_min if row_min is not None else "", MIN_GRID_FMT,
            bg=_ramp_fill(row_min, lo, hi), tx=_ramp_tx(row_min, lo, hi))
        r += 1

    S.h(r, cc + 1, "Grand Total", size=8, tx=_GRID_HDR_TX, bg=_GRID_HDR_BG)
    all_vals = []
    for mi in range(12):
        col_vals = [data[y][mi] for y in years if mi in (data.get(y) or {})]
        cv = min(col_vals) if col_vals else None
        S.v(r, cc + 2 + mi, cv if cv is not None else "", MIN_GRID_FMT,
            bg=_ramp_fill(cv, lo, hi), tx=_ramp_tx(cv, lo, hi))
        all_vals.extend(col_vals)
    gv = min(all_vals) if all_vals else None
    S.v(r, cc + gt_col, gv if gv is not None else "", MIN_GRID_FMT,
        bg=_ramp_fill(gv, lo, hi), tx=_ramp_tx(gv, lo, hi))
    return r


def _write_min_grids(S, wm, axis: str, after_row: int, base_col: int = 1) -> int:
    """Both MIN grids for one axis, stacked under `after_row`. Returns last row."""
    mae_title = wm.get("mae_title") or "Min of Final MAE"
    live_title = wm.get("live_title") or "Min of Actual Live DD"
    write = _write_min_weekly_grid if axis == "wow" else _write_min_monthly_grid
    mae_key, ldd_key = (("wow_mae", "wow_ldd") if axis == "wow"
                        else ("mom_mae", "mom_ldd"))
    r = write(S, wm, mae_title, mae_key, after_row + MIN_GRID_GAP, base_col)
    r = write(S, wm, live_title, ldd_key, r + MIN_GRID_GAP, base_col)
    return r


def _write_min_pivot_sheet(wb: Workbook, placed: List[Tuple[Dict, int, int]],
                           axis: str, sheet_name: str, n_cols: int) -> None:
    """MIN pivots arranged on the SAME axes as the WOW/MOM Summary grid.

    `placed` is [(item, band_row_index, col_index), ...] computed by the caller
    with the summary's own slotting, so a pivot sits at the same grid position
    as its block: adjustment across (No Adj -> Rise -> Fall -> Rise or Fall),
    strike down. Laying these out by a blind "N per line" instead made the order
    look random — a 4-adjustment sweep got split 3+1 across lines and the columns
    no longer meant anything.

    Kept OFF the Summary sheets on purpose: those are cross-combo comparison
    grids and two extra tables under all 24+ blocks made them unreadable. Each
    unit here is caption + Min-of-MAE + Min-of-Live-DD.
    """
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B1"
    S = _Sheet(ws)
    items = [p[0] for p in placed]
    nw = items[0]["wm"]["n_weeks"] if items else 53
    ny = len(items[0]["wm"]["wow_years" if axis == "wow" else "mom_years"]) if items else 0

    span = (2 + nw) if axis == "wow" else 14        # columns one unit occupies
    grid_h = (MIN_WEEKLY_ROWS if axis == "wow" else MIN_MONTHLY_ROWS) + ny
    unit_h = 2 * MIN_GRID_GAP + 2 * grid_h - 1      # caption row through last grid row
    col_stride = span + 2                            # 2-column gutter, as the summary uses
    row_stride = unit_h + (MIN_GRID_GAP - 1)         # 3 blank rows under each band

    for it, ri, ci in placed:
        row = 1 + ri * row_stride
        col = 1 + ci * col_stride
        # Combination caption — the whole point of the separate sheet.
        S.h(row, col, it["title"], size=11, tx=_HEADER_TX, bg=_HEADER_BG,
            align=Alignment(horizontal="left", vertical="center", indent=1))
        for c in range(col + 1, col + span):
            S.h(row, c, "", tx=_HEADER_TX, bg=_HEADER_BG)
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
        ws.row_dimensions[row].height = 18
        _write_min_grids(S, it["wm"], axis, row, col)

    for slot in range(n_cols):
        base = 1 + slot * col_stride
        if axis == "wow":
            _set_wow_widths(ws, nw, base)
        else:
            _set_mom_widths(ws, base)


def _set_wow_widths(ws, nw, base_col: int = 1):
    cc = base_col - 1
    ws.column_dimensions[get_column_letter(cc + 1)].width = 16
    for w in range(1, nw + 1):
        ws.column_dimensions[get_column_letter(cc + 1 + w)].width = 7
    for c in (2 + nw, 3 + nw, 4 + nw, 5 + nw):
        ws.column_dimensions[get_column_letter(cc + c)].width = 8


def _set_mom_widths(ws, base_col: int = 1):
    cc = base_col - 1
    ws.column_dimensions[get_column_letter(cc + 1)].width = 8
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(cc + c)].width = 9


# ── public entry points ────────────────────────────────────────────────────
def write_wow_mom_combined(wb: Workbook, cleaned: List[Dict], has_midcap: bool,
                           title: str, yearly: bool = False) -> bool:
    """Single 'WOW & MOM Summary' sheet (WOW on top, MOM below). Mirrors JS."""
    wm = _wm_from_cleaned(cleaned, has_midcap, yearly=yearly)
    if not (wm["n_trades"] > 0):
        return False
    ws = wb.create_sheet("WOW & MOM Summary")
    ws.freeze_panes = "B1"
    S = _Sheet(ws)
    wow_total = _write_wow_block(S, wm, title, 1)
    # MIN grids sit under their OWN block: weekly pair below WOW, monthly pair
    # below MOM. The MOM block therefore starts after the weekly grids.
    wow_end = _write_min_grids(S, wm, "wow", wow_total)
    # Same MIN_GRID_GAP as between the grids, so the spacing is uniform down the
    # whole sheet instead of 3 rows before a grid and 1 before the next block.
    mom_total = _write_mom_block(S, wm, title, wow_end + MIN_GRID_GAP)
    _write_min_grids(S, wm, "mom", mom_total)
    _set_wow_widths(ws, wm["n_weeks"])
    return True


def _rgb6(color) -> Optional[str]:
    """openpyxl color → visible 6-hex RGB (drop the alpha byte), or None."""
    v = getattr(color, "rgb", None)
    if isinstance(v, str) and len(v) == 8:
        return v[2:].upper()
    return None


def _ws_to_ops(ws) -> Dict[str, Any]:
    """Extract a fully-built openpyxl worksheet into the plain layout-ops dict the
    Rust writer (algotest_native.write_layout_sheet_xlsx) consumes. The WOW/MOM sheet
    leans on openpyxl's mutable cells (per-side medium borders, month-edge, drawdown
    boxes, partial overwrites); extracting the FINAL cell state reproduces all of it
    without re-implementing the block writers. Merge fillers + fully-default cells are
    skipped; borders become per-side style lists so Rust redraws them exactly."""
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.utils import column_index_from_string

    merges: List[Tuple[int, int, int, int]] = []
    covered: set = set()
    for rng in ws.merged_cells.ranges:
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        merges.append((r1, c1, r2, c2))
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if (rr, cc) != (r1, c1):
                    covered.add((rr, cc))

    cells: List[Dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column
            if (r, c) in covered:
                continue
            fill_rgb = (_rgb6(cell.fill.fgColor)
                        if (cell.fill is not None and cell.fill.patternType) else None)
            b = cell.border
            sides = [getattr(getattr(b, s), "style", None)
                     for s in ("top", "left", "bottom", "right")]
            has_border = any(sides)
            val = cell.value
            if (val is None or val == "") and fill_rgb is None and not has_border:
                continue
            f = cell.font
            if not has_border:
                border: Any = False
            elif all(s == "thin" for s in sides):
                border = True
            else:
                border = sides
            nf = cell.number_format
            cells.append({
                "r": r, "c": c,
                "v": (None if val == "" else val),
                "bold": bool(f.bold),
                "size": float(f.size or 11),
                "fc": (_rgb6(f.color) or "000000"),
                "bg": fill_rgb,
                "align": ("L" if cell.alignment.horizontal == "left" else "C"),
                "border": border,
                "nfmt": (None if nf in (None, "General") else nf),
            })

    col_widths = []
    for letter, dim in ws.column_dimensions.items():
        if dim.width is None:
            continue
        lo = dim.min or column_index_from_string(letter)
        hi = dim.max or lo
        for ci in range(lo, hi + 1):
            col_widths.append((ci, float(dim.width)))
    row_heights = [(r, float(dim.height))
                   for r, dim in ws.row_dimensions.items() if dim.height is not None]

    freeze = None
    if ws.freeze_panes:
        c0, r0, *_ = range_boundaries(ws.freeze_panes)
        freeze = (r0 - 1, c0 - 1)   # openpyxl "B1" → set_freeze_panes(0, 1)

    return {"name": ws.title, "cells": cells, "merges": merges,
            "row_heights": row_heights, "col_widths": col_widths, "freeze": freeze}


def wow_mom_ops(cleaned: List[Dict], has_midcap: bool, title: str,
                yearly: bool = False) -> Optional[Dict[str, Any]]:
    """Rust path — build the 'WOW & MOM Summary' sheet via openpyxl (unchanged logic),
    then extract it to a layout-ops dict. Returns None when there are no trades."""
    wm = _wm_from_cleaned(cleaned, has_midcap, yearly=yearly)
    if not (wm["n_trades"] > 0):
        return None
    wb = Workbook()
    wb.remove(wb.active)
    # `yearly` MUST be forwarded: this call builds the sheet that is actually
    # emitted (the `wm` above is only used for the n_trades check). Dropping it
    # here silently rebuilt WOW with yearly=False, collapsing every yearly trade
    # into its December expiry's ISO week.
    if not write_wow_mom_combined(wb, cleaned, has_midcap, title, yearly=yearly):
        return None
    return _ws_to_ops(wb["WOW & MOM Summary"])


def _adj_split(title: str) -> Tuple[str, str]:
    """Fallback: split 'PE ATM | No Adj' into (row_label, adj_label)."""
    if " | " in title:
        left, right = title.split(" | ", 1)
        return left.strip(), right.strip()
    return title.strip(), "No Adj"


def _adj_sort_key(adj_label: str) -> Tuple[int, float]:
    """Order adjustments: No Adj, then Rise/Fall/Rise or Fall by ascending magnitude."""
    s = adj_label.strip().lower()
    if s.startswith("no adj") or s == "":
        return (0, 0.0)
    m = None
    import re as _re
    mm = _re.search(r"([\d.]+)", s)
    mag = float(mm.group(1)) if mm else 0.0
    if s.startswith("rise or fall"):
        grp = 3
    elif s.startswith("fall"):
        grp = 2
    elif s.startswith("rise"):
        grp = 1
    else:
        grp = 4
    return (grp, mag)


# ── Variant axis ────────────────────────────────────────────────────────────
# A sweep can vary parameters the combo LABEL never encodes (SL, target, trail,
# DTE, overall SL/target, lots…): `combo_labeler.label_combo` only emits
# strikes, futures, midcap, adjustment, expiry and shift. Those combos landed on
# an identical (strike, adjustment) grid cell and used to be pushed DOWNWARD as
# "(2)/(3)/(4)". They now get their own sub-column, titled by what differs.

# Params already shown by the strike / expiry / shift / adjustment parts of a
# block title — naming them again in the variant suffix is pure duplication.
_VARIANT_SKIP = re.compile(r"strike_selection|\.expiry$|^expiry|shift|spot_adjustment",
                           re.IGNORECASE)

_VARIANT_NAMES = {
    "stopLoss.value": "SL",
    "targetProfit.value": "TGT",
    "slWithBuffer.value": "SLB",
    "slWithBuffer.buffer_pct": "SLB buf",
    "trailSL.trigger": "Trail trig",
    "trailSL.move": "Trail move",
    "entry_dte": "Entry DTE",
    "exit_dte": "Exit DTE",
    "min_days_to_entry": "Min days",
    "overall_sl_value": "Overall SL",
    "overall_target_value": "Overall TGT",
    "buffer_strike_value": "Buffer strike",
}


def _variant_param_name(path: str) -> str:
    """'legs[1].stopLoss.value' → 'L2 SL'; unknown paths fall back to a
    de-camel-cased last segment so a new sweepable param still reads sanely."""
    m = re.match(r"legs\[(\d+)\]\.(.+)$", path)
    prefix, tail = "", path
    if m:
        prefix, tail = f"L{int(m.group(1)) + 1} ", m.group(2)
    name = _VARIANT_NAMES.get(tail)
    if name is None:
        name = tail.split(".")[-1].replace("_", " ")
        name = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", name).strip()
        name = name[:1].upper() + name[1:]
    return prefix + name


def _fmt_variant_value(v: Any) -> str:
    if isinstance(v, bool):
        return "on" if v else "off"
    if isinstance(v, (int, float)):
        return f"{v:g}"
    return str(v)


def variant_labels(combo_by_safe: Dict[str, Dict[str, Any]]) -> Dict[str, str]:
    """combo_label_safe → short label naming the swept params that DIFFER across
    the job and aren't already in the block title.

    Constant params are dropped (they'd be noise on every block). A sweep where
    nothing extra varies returns {} — every label is empty and the grid layout
    is byte-for-byte what it was before.
    """
    values: Dict[str, set] = {}
    for combo in combo_by_safe.values():
        for path, val in (combo or {}).items():
            values.setdefault(path, set()).add(_fmt_variant_value(val))
    varying = sorted(p for p, vs in values.items()
                     if len(vs) > 1 and not _VARIANT_SKIP.search(p))
    if not varying:
        return {}
    out: Dict[str, str] = {}
    for safe, combo in combo_by_safe.items():
        combo = combo or {}
        parts = [f"{_variant_param_name(p)} {_fmt_variant_value(combo[p])}"
                 for p in varying if p in combo]
        out[safe] = ", ".join(parts)
    return out


#  Matches combo_labeler.py's _adjustment_token: 'adj_{rise|fall|both}_{val}{unit}',
# written inline right after the owning leg's own tokens rather than collected
# at the end with an 'L{n}' tag — so the leg number isn't recoverable from the
# token text alone, only "which adjustments fired" is.
_PER_LEG_ADJ = re.compile(r"adj_(rise|fall|both)_([\d.]+)(pct|pts)", re.IGNORECASE)

_PER_LEG_ADJ_WORD = {"rise": "Rise", "fall": "Fall", "both": "Rise or Fall"}


def adj_label_from_combo_label(combo_label: str) -> str:
    """'..._adj_rise_1pct_...' → 'Rise 1%'; '' when there is none.

    `combo_columns.spot_adjustment` only carries the STRATEGY-level knob, so a
    sweep that adjusts spot PER LEG reports "NoAdjustment" for every combo and
    the whole grid collapses into a single column. The per-leg token IS in the
    combo label (combo_labeler._adjustment_token), so read it back from there.
    Only consulted when the strategy-level knob is off.
    """
    segs = []
    for word, mag, unit in _PER_LEG_ADJ.findall(combo_label or ""):
        pretty = _PER_LEG_ADJ_WORD.get(word.lower(), word)
        segs.append(f"{pretty} {mag}{'pts' if unit.lower() == 'pts' else '%'}")
    return " + ".join(segs)


def write_merged_wow_mom(wb: Workbook, combos: List[Dict]) -> bool:
    """
    Two sheets ('WOW Summary' + 'MOM Summary') laid out as a 2D grid:
      • columns  = spot-adjustment (No Adj, Rise, Fall, Rise or Fall) left→right,
                   then any remaining swept variant, wrapping 4-across
      • rows     = each distinct strike / expiry / shift signature top→bottom,
                   each occupying as many stacked bands as it has variants

    Blocks ALWAYS flow horizontally first — a job with one adjustment and six
    variants gets 4 across + 2 on the band below, not a six-deep vertical stack.

    Deliberately carries NO Min-of-MAE / Min-of-Live-DD grids: those are per-combo
    detail and live in the tradesheet's own 'WOW & MOM Summary' sheet. This sheet
    exists to compare combos side by side.

    combos: [{title, cleaned, has_midcap, adj_key, adj_label, row_key, row_sort,
              variant_label}, ...].
    adj_key/adj_label/row_key are optional — when absent they are derived from the
    'strike | adjustment' title so the layout still works. Per-block content and
    calculations are byte-identical to the per-combo tradesheet; only positioning
    changes.
    """
    # Build each combo's wm + resolve grid axes.
    def _intkeys(o):
        # A stored wm (JSON round-trip through Redis) has its integer year/week
        # dict keys turned into strings; restore them so year lookups match the
        # int values in wow_years/mom_years.
        if isinstance(o, dict):
            return {(int(k) if isinstance(k, str) and k.lstrip("-").isdigit() else k): _intkeys(v)
                    for k, v in o.items()}
        if isinstance(o, list):
            return [_intkeys(x) for x in o]
        return o

    items = []
    for c in combos:
        # Use the pre-computed wm (built inline in the worker during the sweep)
        # when present; otherwise derive it from the cleaned rows (legacy path).
        _pre = c.get("wm")
        wm = _intkeys(_pre) if _pre else _wm_from_cleaned(
            c["cleaned"], c.get("has_midcap", False), yearly=bool(c.get("yearly", False)))
        # A wm stored in Redis BEFORE the MIN grids existed has no wow_mae/etc, so
        # the grids would render empty. Recompute from the cleaned rows when we
        # still have them rather than shipping four blank tables.
        if "wow_mae" not in wm and c.get("cleaned"):
            wm = _wm_from_cleaned(c["cleaned"], c.get("has_midcap", False),
                                  yearly=bool(c.get("yearly", False)))
        if not (wm["n_trades"] > 0):
            continue
        title = c.get("title") or "Strategy"
        row_lbl, adj_lbl = _adj_split(title)
        adj_label = c.get("adj_label") or adj_lbl
        adj_key = c.get("adj_key") or adj_label
        row_key = c.get("row_key") or row_lbl
        row_sort = c.get("row_sort")
        items.append({
            "wm": wm, "title": title,
            "adj_key": adj_key, "adj_label": adj_label,
            "row_key": row_key, "row_sort": row_sort,
            "variant_label": (c.get("variant_label") or "").strip(),
        })
    if not items:
        return False

    # Collapse combos that are genuinely the SAME run. A gated sweep emits these
    # routinely: `legs[1].spot_adjustment.direction` still varies rise/fall/both
    # while `.enabled` is false, so three combos produce one identical result.
    # They land on one cell with nothing to tell them apart, and stacking them
    # is just the same block three times. Identical wm ⇒ identical block, so
    # this only ever removes exact repeats — the same intent as
    # result_store._combo_fingerprint (same label + same P&L = duplicate),
    # applied at layout time instead of by discarding the combo's metadata.
    _seen_blocks: set = set()
    _unique = []
    for it in items:
        # sort_keys: within one workbook some wm come back from Redis and some
        # are rebuilt, so raw dict order is not comparable.
        sig = (it["row_key"], it["adj_key"], it["variant_label"],
               json.dumps(it["wm"], sort_keys=True, default=str))
        if sig in _seen_blocks:
            continue
        _seen_blocks.add(sig)
        _unique.append(it)
    items = _unique

    # Combos that collapse to the SAME (row_key, adj_key) grid cell get their own
    # HORIZONTAL sub-column instead of being stacked downward.
    #
    # This happens whenever the sweep varies something the grid has no axis for:
    # a spread whose two legs are the same option type (the strike display can
    # only show one leg), or — far more commonly — any param the combo label
    # never encodes at all (SL, target, trail, DTE…). Two combos writing into one
    # block would crash on an already-merged header cell ("'MergedCell' object
    # attribute 'value' is read-only"), so each extra one is given the next
    # variant slot; `variant_label` (built by `variant_labels`) names what
    # actually differs. Jobs with NO collisions get variant 0 for every item →
    # the layout math below reduces to the original grid, byte-for-byte.
    _cell_count: Dict[Tuple[str, str], int] = {}
    for it in items:
        cell = (it["row_key"], it["adj_key"])
        n = _cell_count.get(cell, 0)
        _cell_count[cell] = n + 1
        it["variant"] = n
        if it["variant_label"]:
            it["title"] = f'{it["title"]} | {it["variant_label"]}'
        elif n > 0:
            # Nothing distinguishable to name (e.g. a same-option-type spread) —
            # keep the old positional tag so the blocks stay tellable apart.
            it["title"] = f'{it["title"]} ({n + 1})'

    # Column axis: unique adjustments, ordered No Adj → Rise → Fall → Rise or Fall.
    #
    # Sort on the segments that DIFFER, not the whole label: when every combo
    # carries the same constant leg adjustment (e.g. "Rise 1% (L1)") and only a
    # second leg is swept, every label starts with "Rise …" and the ordering
    # collapses to a tie broken alphabetically — putting Fall before Rise. The
    # column that adds nothing beyond the shared part is the "No Adj" one.
    adj_seen: Dict[str, str] = {}   # adj_key → adj_label
    for it in items:
        adj_seen.setdefault(it["adj_key"], it["adj_label"])
    _seg_sets = [set(lbl.split(" + ")) for lbl in adj_seen.values()]
    _common = set.intersection(*_seg_sets) if len(_seg_sets) > 1 else set()
    _distinct = {
        k: (" + ".join([s for s in lbl.split(" + ") if s not in _common]) or "No Adj")
        for k, lbl in adj_seen.items()
    }
    adj_keys = sorted(adj_seen.keys(), key=lambda k: (_adj_sort_key(_distinct[k]), k))
    adj_index = {k: i for i, k in enumerate(adj_keys)}

    # Row axis: unique row signatures, first-seen order (optionally by row_sort).
    row_order: List[str] = []
    row_meta: Dict[str, Any] = {}
    for i, it in enumerate(items):
        rk = it["row_key"]
        if rk not in row_meta:
            row_meta[rk] = {"first": i, "sort": it["row_sort"]}
            row_order.append(rk)
    row_order.sort(key=lambda rk: (
        row_meta[rk]["sort"] if row_meta[rk]["sort"] is not None else 1e18,
        row_meta[rk]["first"],
    ))
    row_index = {rk: i for i, rk in enumerate(row_order)}

    # Unify year axes across ALL combos so every block has identical dimensions
    # (blank where a combo has no trades that year) → grid rows/cols stay aligned.
    all_wow_years = sorted({y for it in items for y in it["wm"]["wow_years"]})
    all_mom_years = sorted({y for it in items for y in it["wm"]["mom_years"]})
    for it in items:
        it["wm"]["wow_years"] = list(all_wow_years)
        it["wm"]["mom_years"] = list(all_mom_years)

    nw = items[0]["wm"]["n_weeks"]
    n_adj = len(adj_keys)
    ny_w = len(all_wow_years)
    ny_m = len(all_mom_years)

    # ── Horizontal slotting ─────────────────────────────────────────────────
    # Every block goes ACROSS first, then wraps to a band below WITHIN its own
    # strike group; the next strike starts under the last band of the previous.
    #
    #   slot        = adj_index + variant * n_adj   (adjustments stay adjacent)
    #   per_band    = n_adj when the sweep varies the adjustment — so No Adj /
    #                 Rise / Fall / Rise or Fall keep a fixed column each and
    #                 every band lines up under the one above it;
    #               = WRAP otherwise — one adjustment and N variants pack WRAP
    #                 to a line before wrapping.
    #   col_in_band = slot % per_band, band = slot // per_band
    #
    # n_bands is global (not per strike) so all strike groups start at the same
    # column offsets and the sheet reads as one grid.
    WRAP = 4
    per_band = n_adj if n_adj > 1 else WRAP
    n_variants = max(it["variant"] for it in items) + 1
    n_bands = -(-(n_adj * n_variants) // per_band)   # ceil
    n_cols = min(per_band, n_adj * n_variants)

    def _slot(it: Dict[str, Any]) -> Tuple[int, int]:
        s = adj_index[it["adj_key"]] + it["variant"] * n_adj
        return s // per_band, s % per_band

    # ── WOW Summary grid ────────────────────────────────────────────────────
    Wb = 5 + nw                # block width (Year + weeks + Total+MaxDD+Live+RMDD)
    Gw = 2                     # column gap between adjustment sections
    Hw = 5 + ny_w              # block height (4 header rows + years + total)
    Gr = 2                     # row gap between strike sections
    # Excel hard limit. A wide sweep lays blocks ACROSS, so past a certain column
    # count the grid silently exceeds it and Excel refuses to open the file —
    # after the entire sweep has already been paid for. Blocks past the limit
    # continue on "<name> (2)", "(3)", ... instead: same slotting, column index
    # re-based per page, so every page reads exactly like the first.
    XL_MAX_COLS = 16384

    def _paged(name, block_w, gap, set_widths):
        per_sheet = max(1, XL_MAX_COLS // (block_w + gap))
        sheets = {}

        def sheet_for(ci):
            page = ci // per_sheet
            if page not in sheets:
                ws = wb.create_sheet(name if page == 0 else f"{name} ({page + 1})")
                ws.freeze_panes = "B1"
                sheets[page] = _Sheet(ws), ws
            return sheets[page][0]

        def finish():
            for page, (_s, ws) in sheets.items():
                lo = page * per_sheet
                for ci in range(lo, min(n_cols, lo + per_sheet)):
                    set_widths(ws, 1 + (ci - lo) * (block_w + gap))
            if len(sheets) > 1:
                logger.warning(
                    "[WOW_MOM] %s split across %d sheets: %d block columns exceed "
                    "Excel's %d-column limit", name, len(sheets), n_cols, XL_MAX_COLS,
                )
        return per_sheet, sheet_for, finish

    w_per_sheet, w_sheet_for, w_finish = _paged(
        "WOW Summary", Wb, Gw, lambda ws, c: _set_wow_widths(ws, nw, c))
    for it in items:
        band, ci = _slot(it)
        ri = row_index[it["row_key"]] * n_bands + band
        base_row = 1 + ri * (Hw + Gr)
        base_col = 1 + (ci % w_per_sheet) * (Wb + Gw)
        # NOTE: no MIN grids here — they belong to the per-combo tradesheet only.
        # The merged summary is a cross-combo comparison grid; stacking two extra
        # tables under all 24+ blocks made it unreadable.
        _write_wow_block(w_sheet_for(ci), it["wm"], it["title"], base_row, base_col)
    w_finish()

    # ── MOM Summary grid ────────────────────────────────────────────────────
    Mb = 18                    # block width (Year + 12 months + Total+MaxDD+Live+RMDD, stats to col 18)
    Gm = 2                     # column gap between adjustment sections (matches Gw)
    Hm = 4 + ny_m
    m_per_sheet, m_sheet_for, m_finish = _paged(
        "MOM Summary", Mb, Gm, lambda ws, c: _set_mom_widths(ws, c))
    for it in items:
        band, ci = _slot(it)
        ri = row_index[it["row_key"]] * n_bands + band
        base_row = 1 + ri * (Hm + Gr)
        base_col = 1 + (ci % m_per_sheet) * (Mb + Gm)
        _write_mom_block(m_sheet_for(ci), it["wm"], it["title"], base_row, base_col)
    m_finish()

    # ── MIN pivots, one sheet per axis ──────────────────────────────────────
    # Placed with the SAME slotting as the blocks above, so each pivot sits at
    # its block's grid position: adjustment across, strike down. Reuse _slot /
    # row_index rather than re-deriving an order — that's what kept the two
    # sheets from agreeing.
    placed = []
    for it in items:
        band, ci = _slot(it)
        placed.append((it, row_index[it["row_key"]] * n_bands + band, ci))
    placed.sort(key=lambda p: (p[1], p[2]))
    _write_min_pivot_sheet(wb, placed, "wow", "WOW Min Pivots", n_cols)
    _write_min_pivot_sheet(wb, placed, "mom", "MOM Min Pivots", n_cols)
    return True


# ── Ops-emitting workbook (Rust render path) ────────────────────────────────
# Drop-in stand-ins for the handful of openpyxl objects the block writers touch,
# so the ~1,500 lines of layout logic above run UNCHANGED and simply produce the
# layout-ops dicts that algotest_native renders.
#
# Why this exists: openpyxl charges for every `cell.font = ...` — the descriptor
# routes into the workbook's style table, which hashes the style object to dedupe
# it. Profiling the merged WOW/MOM for 3,969 combos measured 432s of build vs 42s
# of save, and 31M calls to Serialisable.__hash__ inside that build. Caching the
# style OBJECTS (see _cached_font above) does not help, because the cost is per
# ASSIGNMENT, not per construction. Here assignment is a plain attribute store.
#
# The writers mutate cells after creating them (per-side borders, month edges,
# drawdown boxes, partial overwrites), so cells are kept live in a dict keyed by
# (row, col) and only flattened to ops at the end — mirroring what _ws_to_ops
# extracts from a finished openpyxl sheet.
class _OpsCell:
    __slots__ = ("row", "column", "value", "font", "fill", "alignment",
                 "border", "number_format")

    @property
    def column_letter(self):
        from openpyxl.utils import get_column_letter
        return get_column_letter(self.column)

    def __init__(self, row, column, value=None):
        self.row = row
        self.column = column
        self.value = value
        # set below; declared here so __slots__ stays accurate
        self.font = None
        self.fill = None
        self.alignment = None
        self.border = None
        self.number_format = None


class _OpsDim:
    __slots__ = ("width", "height")

    def __init__(self):
        self.width = None
        self.height = None


class _OpsWorksheet:
    """Quacks like the slice of openpyxl.Worksheet the WOW/MOM writers use."""

    def __init__(self, title):
        self.title = title
        self.freeze_panes = None
        self._cells = {}
        self._merges = []
        self.column_dimensions = _DimMap()
        self.row_dimensions = _DimMap()
        # Row/column extents are tracked incrementally, and the columns present
        # in each row are indexed. Deriving either by scanning _cells is O(cells)
        # per call, and both are called once PER ROW by the row-oriented builders
        # (append → max_row, then ws[max_row]) — that is quadratic, and it made a
        # 2,205-row summary take 6.5s when the actual write is 0.15s.
        self._max_row = 0
        self._row_cols = {}

    def cell(self, row, column, value=None):
        key = (row, column)
        c = self._cells.get(key)
        if c is None:
            c = _OpsCell(row, column, value)
            self._cells[key] = c
            cols = self._row_cols.get(row)
            if cols is None:
                cols = self._row_cols[row] = []
            cols.append(column)
            if row > self._max_row:
                self._max_row = row
        elif value is not None:
            c.value = value
        return c

    def merge_cells(self, start_row=None, start_column=None,
                    end_row=None, end_column=None):
        self._merges.append((start_row, start_column, end_row, end_column))

    # ── row-oriented API (summary_workbook builds this way) ─────────────────
    @property
    def max_row(self):
        return self._max_row

    @property
    def max_column(self):
        return max((c for _, c in self._cells), default=0)

    def append(self, values):
        """openpyxl-compatible append: write `values` as the next row.

        Cells are created even for None so max_row advances on an all-empty
        row, matching openpyxl — otherwise the next append would overwrite it.
        """
        r = self.max_row + 1
        for i, v in enumerate(values, start=1):
            cell = self.cell(r, i)
            cell.value = v
        return r

    def __getitem__(self, key):
        """ws[row_index] → that row's cells, left to right (openpyxl returns a
        tuple). Only integer row access is used by the summary builder."""
        if isinstance(key, int):
            cols = self._row_cols.get(key) or []
            return tuple(self._cells[(key, c)] for c in sorted(cols))
        raise TypeError("only integer row access is supported")

    @property
    def columns(self):
        """Iterate columns as tuples of cells, like openpyxl."""
        by_col = {}
        for (r, c), cell in self._cells.items():
            by_col.setdefault(c, []).append((r, cell))
        for c in sorted(by_col):
            yield tuple(cell for _r, cell in sorted(by_col[c]))

    # ── serialise ───────────────────────────────────────────────────────────
    def to_ops(self):
        covered = set()
        for r1, c1, r2, c2 in self._merges:
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    if (rr, cc) != (r1, c1):
                        covered.add((rr, cc))

        cells = []
        for (r, c), cell in self._cells.items():
            if (r, c) in covered:
                continue
            fill = cell.fill
            fill_rgb = (_rgb6(fill.fgColor)
                        if (fill is not None and fill.patternType) else None)
            b = cell.border
            if b is None:
                sides = [None, None, None, None]
            else:
                sides = [getattr(getattr(b, s), "style", None)
                         for s in ("top", "left", "bottom", "right")]
            has_border = any(sides)
            val = cell.value
            if (val is None or val == "") and fill_rgb is None and not has_border:
                continue
            f = cell.font
            if not has_border:
                border = False
            elif all(s == "thin" for s in sides):
                border = True
            else:
                border = sides
            nf = cell.number_format
            al = cell.alignment
            cells.append({
                "r": r, "c": c,
                "v": (None if val == "" else val),
                "bold": bool(f.bold) if f is not None else False,
                "size": float((f.size if f is not None else None) or 11),
                "fc": ((_rgb6(f.color) if f is not None else None) or "000000"),
                "bg": fill_rgb,
                "align": ("L" if (al is not None and al.horizontal == "left") else "C"),
                "border": border,
                "nfmt": (None if nf in (None, "General") else nf),
            })
        # Deterministic order, so a diff against the openpyxl path lines up.
        cells.sort(key=lambda d: (d["r"], d["c"]))

        col_widths = [(_col_index(letter), float(dim.width))
                      for letter, dim in self.column_dimensions.items()
                      if dim.width is not None]
        row_heights = [(r, float(dim.height))
                       for r, dim in self.row_dimensions.items()
                       if dim.height is not None]
        freeze = None
        if self.freeze_panes:
            from openpyxl.utils.cell import range_boundaries
            c0, r0, *_ = range_boundaries(self.freeze_panes)
            freeze = (r0 - 1, c0 - 1)
        # Sorted, not insertion-ordered: openpyxl keeps merges in an unordered
        # set, so sorting is what makes this path diffable against _ws_to_ops
        # (the parity harness compares them directly). Excel does not care about
        # merge order, only the set.
        return {"name": self.title, "cells": cells,
                "merges": sorted(tuple(m) for m in self._merges),
                "row_heights": row_heights, "col_widths": col_widths,
                "freeze": freeze}


class _DimMap(dict):
    """column_dimensions[letter] / row_dimensions[idx] auto-vivify, like openpyxl."""

    def __getitem__(self, key):
        d = self.get(key)
        if d is None:
            d = _OpsDim()
            self[key] = d
        return d


def _col_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


class _OpsWorkbook:
    """Stand-in for openpyxl.Workbook — only create_sheet is used by the writers."""

    def __init__(self, with_active=False):
        self.worksheets = []
        # openpyxl.Workbook() starts with one sheet; summary_workbook builds on
        # it via wb.active. The WOW/MOM callers create their own sheets and pass
        # with_active=False so no empty tab is emitted.
        self.active = self.create_sheet("Sheet") if with_active else None

    def create_sheet(self, title=None, index=None):
        ws = _OpsWorksheet(title or f"Sheet{len(self.worksheets) + 1}")
        if index is None:
            self.worksheets.append(ws)
        else:
            self.worksheets.insert(index, ws)   # index 0 → first tab
        return ws

    def __getitem__(self, name):
        for ws in self.worksheets:
            if ws.title == name:
                return ws
        raise KeyError(name)

    def to_ops(self):
        return [ws.to_ops() for ws in self.worksheets]
