"""
Backend XLSX builder for optimizer tradesheets.

Replicates buildTradeExcel.js (ExcelJS) logic using openpyxl so the ZIP
endpoint can include the same Trade Sheet + Summary format without requiring
a browser/Node process.
"""
from __future__ import annotations

import calendar
import io
import logging
import math
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from services.trade_anchor import anchor_row as _anchor_row
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Palette (matches buildTradeExcel.js C object) ────────────────────────────
_NAVY_BG    = "1F3864"
_SECTION_BG = "2C5F8A"
_HEADER_BG  = "34495E"
_SUB_HDR_BG = "D6E4F7"
_SUB_HDR_TX = "1F3864"
_GREEN_BG   = "D4EFDF"
_GREEN_TX   = "1E7E34"
_RED_BG     = "FDE8E8"
_RED_TX     = "C0392B"
_LABEL_BG   = "F2F6FA"
_ALT_ROW    = "F9FBFD"
_BORDER_CLR = "B0C4D8"
_WHITE      = "FFFFFF"
_DARK_TXT   = "1A1A2E"
_DARK2_TXT  = "2C3E50"
_WHITE_TXT  = "FFFFFF"

# Style caching — openpyxl deduplicates styles via __eq__/__hash__ on each
# cell assignment, which is O(n) over the style table.  Reusing the SAME
# object reference makes the identity-equality fast-path kick in and cuts
# XLSX build time from ~3s to ~0.6s per file.
_FILL_CACHE: Dict[str, PatternFill] = {}
_FONT_CACHE: Dict[tuple, Font]       = {}
_BORDER_CACHE: Dict[str, Border]     = {}

def _fill(hex_color: str) -> PatternFill:
    f = _FILL_CACHE.get(hex_color)
    if f is None:
        f = PatternFill("solid", fgColor=hex_color)
        _FILL_CACHE[hex_color] = f
    return f

def _font(bold: bool = False, size: int = 10, color: str = "000000") -> Font:
    key = (bold, size, color)
    f = _FONT_CACHE.get(key)
    if f is None:
        f = Font(bold=bold, size=size, color=color, name="Calibri")
        _FONT_CACHE[key] = f
    return f

def _border() -> Border:
    b = _BORDER_CACHE.get(_BORDER_CLR)
    if b is None:
        side = Side(style="thin", color=_BORDER_CLR)
        b = Border(top=side, left=side, bottom=side, right=side)
        _BORDER_CACHE[_BORDER_CLR] = b
    return b

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")


def _to_num(v) -> Optional[float]:
    if isinstance(v, (int, float)) and math.isfinite(float(v)):
        return float(v)
    if v is None or v == "" or (isinstance(v, float) and math.isnan(v)):
        return None
    try:
        s = str(v).replace(",", "").replace("%", "").replace("₹", "").strip()
        f = float(s)
        return f if math.isfinite(f) else None
    except (TypeError, ValueError):
        return None


def _parse_date(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if not v or v == "":
        return None
    s = str(v).strip()
    # Day-first parsing, matching engine_rust._load_filter_segments and the
    # supported-formats help text. Year-first formats are tried first (they're
    # unambiguous); ambiguous slash/dash dates are then read DAY-first, so
    # "10/05/2019" = 10-May-2019, never 5-Oct (no MM/DD here by design).
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d",
                "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _date_ms(val: Any) -> Optional[int]:
    dt = _parse_date(val)
    if dt is None:
        return None
    return int(calendar.timegm(dt.timetuple()) * 1000)


_DATE_COLS = {
    "Entry Date", "Exit Date", "Expiry",
    "Leg Exit Date", "Lazy Entry Date", "Lazy Exit Date",
}
_TRUE_PCT_COLS = {"Spot P&L %", "CE P&L %", "PE P&L %", "FUT P&L %", "%DD"}
_MAE_COLS      = {"MAE", "MFE", "Net MAE 1", "Net MAE 2", "Final MAE",
                  "Midcap MAE", "Midcap MFE",
                  "Combined Net MAE 1", "Combined Net MAE 2", "Combined Final MAE"}
_TRADE_COLS    = {
    "Net MAE 1", "Net MAE 2", "Final MAE",
    "Net P&L", "% P&L", "Cumulative", "Peak", "DD", "%DD",
    "Lowest NAV", "Actual Live DD",
}

_COL_WIDTHS: Dict[str, int] = {
    "Leg": 12, "Entry Date": 13, "Exit Date": 13,
    "Entry Spot": 12, "Exit Spot": 12,
    "buffer_ref_price": 12, "buffer_strike_offset": 10,
    "Re-Entry Type": 14,
    "Raw Entry Price": 12, "Entry Price": 12,
    "Raw Exit Price": 12, "Exit Price": 12,
    "MAE": 9, "MFE": 9, "Net MAE 1": 10, "Net MAE 2": 10, "Final MAE": 10,
    "Net P&L": 10, "% P&L": 8, "Cumulative": 11, "Peak": 10, "DD": 9, "%DD": 8,
    "Lowest NAV": 13, "Actual Live DD": 15,
    "Spot P&L %": 10, "CE P&L %": 10, "PE P&L %": 10, "FUT P&L %": 10,
    "ATM Strike": 11, "ATM Call Price": 13, "ATM Put Price": 13, "ATM Call+Put Price": 16,
    "ATM Straddle Price Source": 40,
    "Exit Reason": 14, "Strike Shift Reason": 40, "Expiry": 12, "STR Segment": 14, "Filter Segment": 22,
    "Midcap Entry Spot": 15, "Midcap Exit Spot": 15, "Midcap Spot P&L": 14,
    "Midcap Spot P&L %": 15, "Midcap No Of Days": 15, "Midcap Rollover Cost %": 18,
    "Midcap Hypo P&L": 15, "Midcap Hypo P&L %": 16, "Midcap MAE": 12, "Midcap MFE": 12,
    "Combined Net P&L": 15, "Combined Net P&L %": 16, "Combined Cumulative": 17,
    "Combined Peak": 13, "Combined DD": 12, "Combined %DD": 12,
    "Combined Net MAE 1": 16, "Combined Net MAE 2": 16, "Combined Final MAE": 15,
    "Combined Lowest NAV": 16, "Combined Actual Live DD": 18,
}


def _is_lazy(row: Dict) -> bool:
    v = row.get("Is Lazy Leg")
    return v is True or str(v).lower() == "true" or bool(row.get("Lazy Leg Name"))


def _main_leg(legs):
    """The trade's ANCHOR leg row — LATEST Entry Date, ties to lowest Leg number.

    Was `next(<first non-reentry leg in list order>, legs[0])`, i.e. whichever
    leg the user happened to put first in the builder. That made Entry Spot —
    the %P&L / NAV denominator — and the patch-segment assignment depend on leg
    ORDER: reordering a carried-yearly + weekly pair flipped Entry Spot between
    23000 and 25000, moving Max DD, CAGR and every NAV-based stat with it.

    `_is_lazy` is kept in the filter because trade_anchor.is_reentry_row does
    not know about the "Is Lazy Leg"/"Lazy Leg Name" markers this module uses.
    See services/trade_anchor.py for why LATEST entry is the right anchor.
    """
    mains = [l for l in legs
             if not l.get("ReEntryIndex") and not l.get("ReEntryTrigger")
             and not l.get("ReEntryMode") and not _is_lazy(l)]
    return _anchor_row(mains) or (legs[0] if legs else {})


def _get_reentry_type(row: Dict) -> str:
    if _is_lazy(row):
        return "Lazy"
    mode = str(row.get("ReEntryMode") or "").strip()
    if mode:
        return mode
    trigger = str(row.get("ReEntryTrigger") or "").strip()
    if trigger:
        return trigger
    return "Re-Entry" if row.get("ReEntryIndex") else ""


def _is_bearish_leg(leg: Dict) -> bool:
    """CE SELL, PE BUY or FUT SELL — profits when market falls."""
    t  = str(leg.get("Type") or "").upper()
    bs = str(leg.get("B/S")  or "").upper()
    return (
        (t in ("CE", "CALL") and bs == "SELL")
        or (t in ("PE", "PUT") and bs == "BUY")
        or (t == "FUT" and bs == "SELL")
    )


def _is_bullish_leg(leg: Dict) -> bool:
    """CE BUY, PE SELL or FUT BUY — profits when market rises."""
    t  = str(leg.get("Type") or "").upper()
    bs = str(leg.get("B/S")  or "").upper()
    return (
        (t in ("CE", "CALL") and bs == "BUY")
        or (t in ("PE", "PUT") and bs == "SELL")
        or (t == "FUT" and bs == "BUY")
    )


def _calc_trade_mae(legs: List[Dict], net_pnl_pct: Optional[float] = None):
    """Replicate calcTradeMae from JS.

    Every leg (option or future) is classified by market direction:
      Bullish (CE BUY / PE SELL / FUT BUY):  adverse when market falls, favorable when rises.
      Bearish (CE SELL / PE BUY / FUT SELL): adverse when market rises, favorable when falls.

    Unified rule (single-leg, multi-leg, options and futures alike):
      nm1 = sum(bullish MAE) + sum(bearish MFE)
      nm2 = sum(bullish MFE) + sum(bearish MAE)
      final = min(nm1, nm2)                              (single directional leg)
      final = min(nm1, nm2, net_pnl_pct)                 (>1 directional leg)

    For MULTI-leg trades the realized Net P&L % is folded into the min so the
    reconstructed combined excursion can never read better than what the trade
    actually booked (nm1/nm2 pair each leg's independent extremes, which may not
    have occurred simultaneously). Single-leg keeps min(nm1, nm2) — its own MAE
    already bounds the realized loss. Mirrors the Midcap Combined Final MAE rule.
    """
    dir_legs = [
        l for l in legs
        if str(l.get("Type") or "").upper() in ("CE", "CALL", "PE", "PUT", "FUT")
    ]
    if not dir_legs:
        return None

    def _sum_field(rows, key):
        total = 0.0
        for r in rows:
            v = _to_num(r.get(key))
            if v is None:
                return None
            total += v
        return total

    def _rnd(v):
        return round(v * 10000) / 10000

    bullish_legs = [l for l in dir_legs if _is_bullish_leg(l)]
    bearish_legs = [l for l in dir_legs if _is_bearish_leg(l)]

    bull_mae = _sum_field(bullish_legs, "MAE"); bull_mfe = _sum_field(bullish_legs, "MFE")
    bear_mae = _sum_field(bearish_legs, "MAE"); bear_mfe = _sum_field(bearish_legs, "MFE")
    if None in (bull_mae, bull_mfe, bear_mae, bear_mfe):
        return None

    nm1 = bull_mae + bear_mfe
    nm2 = bull_mfe + bear_mae
    if len(dir_legs) > 1 and net_pnl_pct is not None:
        final = min(nm1, nm2, net_pnl_pct)
    else:
        final = min(nm1, nm2)
    return (_rnd(nm1), _rnd(nm2), _rnd(final))


# ── Midcap cross-index overlay (mirrors ResultsPanel.exportToCSV exactly) ──────
# Trade-Sheet block appended when a Midcap leg ran. With Midcap, the NIFTY
# trade-level columns (Net MAE 1/2/Final, Net P&L, % P&L, Cumulative, Peak, DD,
# %DD, Lowest NAV, Actual Live DD) are dropped and the Combined versions below
# replace them; leg P&L is labelled "Midcap Hypo P&L".
_MIDCAP_COLS = [
    "Midcap Entry Spot", "Midcap Exit Spot", "Midcap Spot P&L", "Midcap Spot P&L %",
    "Midcap No Of Days", "Midcap Rollover Cost %", "Midcap Hypo P&L", "Midcap Hypo P&L %",
    "Midcap MAE", "Midcap MFE",
    "Combined Net P&L", "Combined Net P&L %", "Combined Cumulative", "Combined Peak",
    "Combined DD", "Combined %DD", "Combined Net MAE 1", "Combined Net MAE 2",
    "Combined Final MAE", "Combined Lowest NAV", "Combined Actual Live DD",
]
_MIDCAP_COLS_SET = set(_MIDCAP_COLS)
# NIFTY trade-level columns dropped from the sheet when Midcap is present.
_NIFTY_DROP_WHEN_MIDCAP = {
    "Net MAE 1", "Net MAE 2", "Final MAE",
    "Net P&L", "% P&L", "Cumulative", "Peak", "DD", "%DD", "Lowest NAV", "Actual Live DD",
}


def _calc_combined_final_mae(legs: List[Dict], mc: Optional[Dict]):
    """Mirror calcCombinedFinalMaePct. NIFTY MAE/MFE are already % of spot, so
    they are summed DIRECTLY (not re-divided by spot), then paired with the
    Midcap MAE/MFE:
        Net MAE 1 = Midcap MFE + NIFTY MAE ;  Net MAE 2 = Midcap MAE + NIFTY MFE
    Returns (nm1, nm2, raw_min) or None if NIFTY MAE/MFE incomplete. The final
    floor with Combined Net P&L % is applied by the caller (chain)."""
    nifty_mae = 0.0
    nifty_mfe = 0.0
    dir_legs = [
        l for l in (legs or [])
        if str(l.get("Type") or "").upper() in ("CE", "CALL", "PE", "PUT", "FUT")
    ]
    for r in dir_legs:
        mae = _to_num(r.get("MAE"))
        mfe = _to_num(r.get("MFE"))
        if mae is None or mfe is None:
            return None
        nifty_mae += mae
        nifty_mfe += mfe
    mid_mae = _to_num((mc or {}).get("Midcap MAE")) or 0.0
    mid_mfe = _to_num((mc or {}).get("Midcap MFE")) or 0.0
    nm1 = mid_mfe + nifty_mae
    nm2 = mid_mae + nifty_mfe

    def _rnd(v):
        return round(v * 10000) / 10000

    return (_rnd(nm1), _rnd(nm2), _rnd(min(nm1, nm2)))


def _build_key_order(rows: List[Dict], has_midcap: bool = False) -> List[str]:
    has_calls     = any(str(r.get("Type") or "").upper() in ("CE", "CALL")  for r in rows)
    has_puts      = any(str(r.get("Type") or "").upper() in ("PE", "PUT")   for r in rows)
    has_futures   = any(str(r.get("Type") or "").upper() == "FUT"           for r in rows)
    has_buffer    = any(
        r.get("buffer_ref_price") not in (None, "", "False", False)         for r in rows
    )
    has_spot_adj  = any(
        r.get("Raw Entry Price") not in (None, "") and
        _to_num(r.get("Raw Entry Price")) != _to_num(r.get("Entry Price"))  for r in rows
    )
    has_reentry   = any(
        r.get("ReEntryIndex") or r.get("ReEntryTrigger") or
        r.get("ReEntryMode") or _is_lazy(r)                                 for r in rows
    )
    has_str       = any(r.get("STR Segment")    not in (None, "") for r in rows)
    has_filter    = any(r.get("Filter Segment") not in (None, "") for r in rows)
    # Strike Shift Reason — shown only when the engine shifted a strike toward
    # ATM because the requested strike had zero turnover (so the user can see
    # WHY a non-ATM strike was taken). Mirrors buildTradeExcel.js hasStrikeShift.
    has_strike_shift = any(r.get("Strike Shift Reason") not in (None, "") for r in rows)
    # Straddle-width context columns present only when the engine populated them
    # (i.e. a straddle_width leg ran). Hidden for every other strike mode.
    has_straddle  = any(r.get("ATM Strike") not in (None, "") for r in rows)
    # Shown only when the ATM CE/PE straddle price itself was illiquid at the
    # leg's own strike gap and had to be sourced from a wider gap — separate
    # from Strike Shift Reason (which is about the final TRADED strike).
    has_straddle_price_source = any(
        r.get("ATM Straddle Price Source") not in (None, "") for r in rows
    )

    order = [
        "Trade", "Leg", "Index", "Entry Date", "Exit Date", "Expiry",
        "Entry Spot", "Exit Spot", "Spot P&L", "Spot P&L %",
        "Type", "Strike",
    ]
    if has_straddle:
        order += ["ATM Strike", "ATM Call Price", "ATM Put Price", "ATM Call+Put Price"]
        if has_straddle_price_source:
            order.append("ATM Straddle Price Source")
    if has_buffer:
        order += ["buffer_ref_price", "buffer_strike_offset"]
    order.append("B/S")
    if has_reentry:
        order.append("Re-Entry Type")
    order.append("Qty")
    if has_spot_adj:
        order.append("Raw Entry Price")
    order.append("Entry Price")
    if has_spot_adj:
        order.append("Raw Exit Price")
    order += ["Exit Price", "MAE", "MFE"]
    # NIFTY trade-level Net MAE — dropped from the Midcap sheet (Combined replaces it).
    if not has_midcap:
        order += ["Net MAE 1", "Net MAE 2", "Final MAE"]
    if has_calls:
        order += ["CE P&L", "CE P&L %"]
    if has_puts:
        order += ["PE P&L", "PE P&L %"]
    if has_futures:
        order += ["FUT P&L", "FUT P&L %"]
    # NIFTY trade-level P&L / NAV / drawdown — dropped from the Midcap sheet
    # (Combined versions in _MIDCAP_COLS replace them).
    if not has_midcap:
        order += ["Net P&L", "% P&L", "Cumulative", "Peak", "DD", "%DD", "Lowest NAV", "Actual Live DD"]
    if has_midcap:
        order += _MIDCAP_COLS
    order.append("Exit Reason")
    if has_strike_shift:
        order.append("Strike Shift Reason")
    if has_str:
        order.append("STR Segment")
    if has_filter:
        order.append("Filter Segment")

    return order, has_calls, has_puts, has_futures


def _aggregate_trades(rows: List[Dict], has_midcap: bool = False,
                      midcap_by_trade: Optional[Dict] = None,
                      patchwise: bool = False,
                      filter_segments: Optional[List] = None) -> Dict[str, Any]:
    """Return per-trade aggregates keyed by str(trade_id), mimicking JS tm dict."""
    grouped: Dict[str, List[Dict]] = {}
    for r in rows:
        k = str(r.get("Trade") or r.get("trade") or 1)
        grouped.setdefault(k, []).append(r)

    # Patchwise reset boundaries: prefer the uploaded filter's segment START dates
    # (reset when a trade's entry crosses into a new segment) so spot-adjustment runs
    # reset too (they never emit a FILTER_END exit reason). Falls back to FILTER_END.
    def _seg_start_ms(s):
        s = s or {}
        return _date_ms(s.get("start") or s.get("Start") or s.get("from")
                        or s.get("start_date") or s.get("startdt"))
    _pw_seg_starts = sorted(ms for ms in (_seg_start_ms(s) for s in (filter_segments or []))
                            if ms is not None)

    def _pw_seg_idx(key: str) -> int:
        legs = grouped.get(key, [])
        mr = _main_leg(legs)
        em = _date_ms(mr.get("Entry Date"))
        i = -1
        if em is not None:
            for j, sm in enumerate(_pw_seg_starts):
                if sm <= em:
                    i = j
                else:
                    break
        return i

    tm: Dict[str, Any] = {}
    for k, legs in grouped.items():
        main = _main_leg(legs)
        spot    = _to_num(main.get("Entry Spot")) or 0.0
        raw_net = _to_num(main.get("Net P&L"))
        if raw_net is None:
            raw_net = sum(
                (_to_num(l.get("CE P&L"))  or 0) +
                (_to_num(l.get("PE P&L"))  or 0) +
                (_to_num(l.get("FUT P&L")) or 0)
                for l in legs
            )
        # Multi-index trade (legs span >1 index): its return is the SUM of each leg's
        # OWN % (leg P&L ÷ that leg's own-index spot), pre-computed as the parent row's
        # "% P&L" — NOT Net ÷ one index's spot. Single-index / Midcap100-overlay trades
        # (one Group Index) keep the Net ÷ Entry Spot behaviour, unchanged.
        _idxs = {str(l.get("Group Index") or "").upper() for l in legs}
        _idxs.discard("")
        _stored_pct = _to_num(main.get("% P&L"))
        if len(_idxs) > 1 and _stored_pct is not None:
            _pct = _stored_pct
        else:
            _pct = (raw_net / spot * 100) if spot != 0 else 0.0
        mae_res = _calc_trade_mae(legs, _pct)

        tm[k] = {
            "net":        raw_net,
            "pct":        _pct,
            "netMae1":    mae_res[0] if mae_res else "",
            "netMae2":    mae_res[1] if mae_res else "",
            "finalMae":   mae_res[2] if mae_res else "",
            "cumulative": "",
            "peak":       "",
            "dd":         "",
            "pctDd":      "",
            "lowestNav":  "",
            "actualLDD":  "",
            # Union of ALL legs' exit reasons (used only for FILTER_END patchwise-
            # reset detection, never displayed). A mixed options+futures trade's
            # MAIN leg is the futures leg (exit reason EXPIRY), so a FILTER_END on
            # the option leg would be missed and the patchwise reset would never
            # fire. Unioning makes the detection leg-order-independent.
            "exitReason": "+".join(
                r for r in ((l.get("Exit Reason") or "").strip() for l in legs) if r
            ),
        }

    # Booked Cumulative / Peak / DD / %DD and Lowest NAV / Actual Live DD pass.
    # Iterate trades in the same CHRONOLOGICAL order used by _build_cleaned_rows
    # so cascade re-entries (higher trade_id, earlier entry date) are placed in
    # the correct time sequence.
    def _parse_entry(r: Dict):
        v = r.get("Entry Date") or r.get("entry_date") or ""
        try:
            return _parse_date(v) or datetime.max
        except Exception:
            return datetime.max

    sorted_rows = sorted(rows, key=lambda r: (
        _parse_entry(r),
        int(str(r.get("Trade") or r.get("trade") or 1)),
        int(str(r.get("Leg")   or r.get("leg")   or 1)),
    ))

    _seen: set = set()
    sorted_keys: List[str] = []
    for _r in sorted_rows:
        _k = str(_r.get("Trade") or _r.get("trade") or 1)
        if _k not in _seen and _k in tm:
            _seen.add(_k)
            sorted_keys.append(_k)

    # Recompute the booked equity curve exactly like the research-sheet
    # formulas: cumulative compounds from prior visible trade rows using
    # `% P&L`; peak is the running max; DD is blank at equity highs; %DD is the
    # Excel ratio DD / Peak.
    cumulative = 100.0
    peak = 100.0
    _prev_pw_key = None
    for k in sorted_keys:
        if patchwise and not has_midcap and _prev_pw_key is not None:
            if _pw_seg_starts:
                new_patch = _pw_seg_idx(k) != _pw_seg_idx(_prev_pw_key)
            else:
                new_patch = "FILTER_END" in (tm[_prev_pw_key].get("exitReason") or "").upper().split("+")
            if new_patch:
                cumulative = 100.0
                peak = 100.0
        _prev_pw_key = k
        t = tm[k]
        pct = t["pct"] if isinstance(t["pct"], float) else 0.0
        cumulative *= (1.0 + pct / 100.0)
        peak = max(peak, cumulative)
        # When at an equity high, drawdown is 0 (not blank).  Previously this
        # was "" which rendered as an empty cell; users expect a clean 0.00.
        dd = cumulative - peak if peak > cumulative else 0.0
        t["cumulative"] = cumulative
        t["peak"] = peak
        t["dd"] = dd
        t["pctDd"] = (dd / peak) if peak != 0 else 0.0

    prev_cum = 100.0
    prev_peak = 100.0
    first_done = False
    _prev_pw_key_ln = None
    for k in sorted_keys:
        if patchwise and not has_midcap and _prev_pw_key_ln is not None:
            if _pw_seg_starts:
                new_patch = _pw_seg_idx(k) != _pw_seg_idx(_prev_pw_key_ln)
            else:
                new_patch = "FILTER_END" in (tm[_prev_pw_key_ln].get("exitReason") or "").upper().split("+")
            if new_patch:
                prev_cum = 100.0
                prev_peak = 100.0
        _prev_pw_key_ln = k
        t = tm[k]
        mae  = t["finalMae"]   if isinstance(t["finalMae"],   float) else None
        peak = t["peak"]       if isinstance(t["peak"],        float) else None
        cum  = t["cumulative"] if isinstance(t["cumulative"],  float) else None
        if mae is not None and peak is not None and prev_peak != 0:
            # Research-verified rule (revised): EVERY trade — including the
            # first — anchors the intra-trade low to the PREVIOUS cumulative
            # pushed down by Final MAE (prev_cum = 100 for the first trade).
            # Formula: AW = AU_prev * (1 + FinalMAE%).
            # Live DD divides by the PREVIOUS trade's peak (AV_prev), not this
            # trade's peak — a trade that closes at a new high is still measured
            # against the peak established going into it.
            lowest_nav = round(prev_cum * (1 + mae / 100) * 100) / 100
            actual_ldd = round((lowest_nav / prev_peak - 1) * 10000) / 100
            t["lowestNav"] = lowest_nav
            t["actualLDD"] = actual_ldd
            first_done = True
        else:
            first_done = True
        if cum is not None:
            prev_cum = cum
        if peak is not None:
            prev_peak = peak

    # Combined NAV / Peak / DD / Net MAE / Lowest NAV chain (Midcap only). Drives
    # the Combined Trade-Sheet columns AND the Combined summary. Final MAE =
    # min(Net MAE 1, Net MAE 2, Combined Net P&L %) per the verified backtest.
    if has_midcap:
        mbt = midcap_by_trade or {}
        nav = 100.0
        peak = 100.0
        prev_nav = 100.0
        prev_peak = 100.0
        first_done = False
        for idx, k in enumerate(sorted_keys):
            if patchwise and idx > 0:
                prev_k = sorted_keys[idx - 1]
                if _pw_seg_starts:
                    new_patch = _pw_seg_idx(k) != _pw_seg_idx(prev_k)
                else:
                    new_patch = "FILTER_END" in (tm[prev_k].get("exitReason") or "").upper().split("+")
                if new_patch:
                    nav = 100.0; peak = 100.0; prev_nav = 100.0; prev_peak = 100.0
            t = tm[k]
            mc = mbt.get(k)
            cpct = _to_num((mc or {}).get("Combined Net P&L %"))
            if cpct is not None and math.isfinite(cpct):
                prev_nav = nav
                prev_peak = peak
                nav *= (1.0 + cpct / 100.0)
                peak = max(peak, nav)
                t["combinedPct"] = cpct
                t["combinedCum"] = round(nav, 4)
                t["combinedPeak"] = round(peak, 4)
                t["combinedDd"] = round(nav - peak, 4)
                t["combinedPctDd"] = round((nav / peak - 1) * 100, 4) if peak != 0 else ""
                cm = _calc_combined_final_mae(grouped.get(k, []), mc)
                if cm is not None:
                    fmae = round(min(cm[0], cm[1], cpct), 4)
                    t["combinedNetMae1"] = cm[0]
                    t["combinedNetMae2"] = cm[1]
                    t["combinedFinalMae"] = fmae
                    # Research-verified rule (revised): every trade (incl. the
                    # first, where prev_nav = 100) anchors the intra-trade low to
                    # prev_nav * (1 + FinalMAE%) — AW = AU_prev * (1 + AM%).
                    lowest_nav = prev_nav * (1.0 + fmae / 100.0)
                    t["combinedLowestNav"] = round(lowest_nav, 4)
                    # Live DD divides by the PREVIOUS trade's peak (AV_prev), not
                    # this trade's peak — AX = AW / AV_prev - 1.
                    t["combinedActualLDD"] = round((lowest_nav / prev_peak - 1) * 100, 4) if prev_peak != 0 else ""
                else:
                    t["combinedNetMae1"] = t["combinedNetMae2"] = ""
                    t["combinedFinalMae"] = t["combinedLowestNav"] = t["combinedActualLDD"] = ""
                first_done = True
            else:
                t["combinedPct"] = None
                for _kk in ("combinedCum", "combinedPeak", "combinedDd", "combinedPctDd",
                            "combinedNetMae1", "combinedNetMae2", "combinedFinalMae",
                            "combinedLowestNav", "combinedActualLDD"):
                    t[_kk] = ""

    # `sorted_keys` is the CANONICAL chronological (entry-date) trade order used
    # to compute every per-trade equity value above (combinedCum / combinedActualLDD
    # / actualLDD). Return it so summary-level reconstructions (max-DD scan,
    # outlier-stripped Live DD) walk trades in the SAME order — otherwise cascade
    # re-entry trades (high engine-id, early date) land in a different slot and
    # the reconstructed numbers diverge between code paths.
    return tm, grouped, sorted_keys


def _build_cleaned_rows(rows: List[Dict], key_order: List[str], tm: Dict,
                        has_midcap: bool = False,
                        midcap_by_trade: Optional[Dict] = None) -> List[Dict]:
    # Sort by Entry Date first so cascade mini-trades (with NEW higher trade
    # IDs but earlier entry dates than later originals) appear interleaved
    # chronologically with the originals.  Secondary keys keep all legs of
    # the same trade grouped together.  The CSV from the engine is already
    # in this order, but we re-sort defensively in case the row dicts came
    # from any source that doesn't preserve order.
    def _parse_entry(r):
        v = r.get("Entry Date") or r.get("entry_date") or ""
        try:
            return _parse_date(v) or datetime.max
        except Exception:
            return datetime.max
    sorted_rows = sorted(rows, key=lambda r: (
        _parse_entry(r),
        int(str(r.get("Trade") or r.get("trade") or 1)),
        int(str(r.get("Leg")   or r.get("leg")   or 1)),
    ))
    # Build engine-trade-id → sequential display number (1, 2, 3, ...) based on
    # FIRST appearance in chronological order.  This is what the "Index" column
    # shows the user, so cascade trades (engine_tid=71+) get the correct
    # sequential number for their chronological position (e.g. Trade=71 enters
    # right after Trade=5, so its display Index is 6, not 71).
    _tid_to_index_no: Dict[str, int] = {}
    _seq_no = 0
    for _r in sorted_rows:
        _ek = str(_r.get("Trade") or _r.get("trade") or 1)
        if _ek not in _tid_to_index_no:
            _seq_no += 1
            _tid_to_index_no[_ek] = _seq_no
    written: set = set()
    cleaned = []
    for trade in sorted_rows:
        k     = str(trade.get("Trade") or trade.get("trade") or 1)
        first = k not in written
        if first:
            written.add(k)
        m   = tm.get(k, {})
        mc  = (midcap_by_trade or {}).get(k) if has_midcap else None
        row = {}
        for key in key_order:
            val = ""
            if key in _TRADE_COLS or (has_midcap and key in _MIDCAP_COLS_SET):
                if not first:
                    val = ""
                elif key == "Net MAE 1":      val = m.get("netMae1", "")
                elif key == "Net MAE 2":      val = m.get("netMae2", "")
                elif key == "Final MAE":      val = m.get("finalMae", "")
                elif key == "Net P&L":        val = m.get("net", "")
                elif key == "% P&L":          val = m.get("pct", "")
                elif key == "Cumulative":     val = m.get("cumulative", "")
                elif key == "Peak":           val = m.get("peak", "")
                elif key == "DD":             val = m.get("dd", "")
                elif key == "%DD":            val = m.get("pctDd", "")
                elif key == "Lowest NAV":     val = m.get("lowestNav", "")
                elif key == "Actual Live DD": val = m.get("actualLDD", "")
                # Renamed Midcap leg P&L (backend keys are "Midcap Leg P&L").
                elif key == "Midcap Hypo P&L":   val = (mc or {}).get("Midcap Leg P&L", "")
                elif key == "Midcap Hypo P&L %": val = (mc or {}).get("Midcap Leg P&L %", "")
                # Combined NAV/DD/Net MAE chain (computed in _aggregate_trades).
                elif key == "Combined Cumulative":     val = m.get("combinedCum", "")
                elif key == "Combined Peak":           val = m.get("combinedPeak", "")
                elif key == "Combined DD":             val = m.get("combinedDd", "")
                elif key == "Combined %DD":            val = m.get("combinedPctDd", "")
                elif key == "Combined Net MAE 1":      val = m.get("combinedNetMae1", "")
                elif key == "Combined Net MAE 2":      val = m.get("combinedNetMae2", "")
                elif key == "Combined Final MAE":      val = m.get("combinedFinalMae", "")
                elif key == "Combined Lowest NAV":     val = m.get("combinedLowestNav", "")
                elif key == "Combined Actual Live DD": val = m.get("combinedActualLDD", "")
                # Remaining Midcap + Combined Net P&L columns from the overlay.
                elif key in _MIDCAP_COLS_SET:          val = (mc or {}).get(key, "")
            elif key == "Leg" and _is_lazy(trade):
                val = trade.get("Lazy Leg Name") or trade.get("Leg", "")
            elif key == "Re-Entry Type":
                val = _get_reentry_type(trade)
            elif key == "Trade":
                # Show user-friendly sequential trade number instead of the
                # engine's internal trade_id (which jumps to 71+ for cascade
                # mini-trades).  Same value as the "Index" column so users
                # see clean 1, 2, 3, ... in both columns.
                val = _tid_to_index_no.get(k, int(str(trade.get("Trade") or trade.get("trade") or 1)))
            elif key == "Index":
                val = _tid_to_index_no.get(k, int(str(trade.get("Trade") or trade.get("trade") or 1)))
            elif key == "Spot P&L %":
                # Spot P&L is a trade-level quantity written on exactly one row
                # per trade (the lowest present leg — see priced_to_tradesheet_records
                # in engine_rust.py). This is purely row-derived: it's blank
                # wherever "Spot P&L" itself is blank on that row, so it
                # automatically follows the same one-row-per-trade placement.
                spot_pnl = _to_num(trade.get("Spot P&L"))
                if spot_pnl is None:
                    val = ""
                else:
                    es2 = _to_num(trade.get("Entry Spot"))
                    val = (spot_pnl / es2) if (es2 and es2 != 0) else ""
            elif key == "CE P&L %":
                pnl = _to_num(trade.get("CE P&L"))
                es  = _to_num(trade.get("Entry Spot"))
                val = (pnl / es) if (pnl is not None and es and es != 0) else ""
            elif key == "PE P&L %":
                pnl = _to_num(trade.get("PE P&L"))
                es  = _to_num(trade.get("Entry Spot"))
                val = (pnl / es) if (pnl is not None and es and es != 0) else ""
            elif key == "FUT P&L %":
                # Per-leg futures P&L as a fraction of Entry Spot (matches CE/PE P&L %).
                pnl = _to_num(trade.get("FUT P&L"))
                es  = _to_num(trade.get("Entry Spot"))
                val = (pnl / es) if (pnl is not None and es and es != 0) else ""
            else:
                val = trade.get(key, "")
            if val is None or (isinstance(val, float) and math.isnan(val)):
                val = ""
            row[key] = val
        cleaned.append(row)
    return cleaned


# ── Sheet 3: Patch wise ───────────────────────────────────────────────────────

def _write_patch_wise_sheet(
    wb: Workbook,
    tm: Dict,
    grouped: Dict,
    sorted_keys: List[str],
    has_midcap: bool,
    midcap_by_trade: Optional[Dict],
    has_calls: bool,
    has_puts: bool,
    filter_segments: Optional[List] = None,
) -> None:
    """openpyxl reference path — build the Patch wise sheet into wb (or nothing
    when there are no patches, matching the original early return)."""
    ws = wb.create_sheet("Patch wise")
    sink = _OpenpyxlSink(ws)
    _patch_wise_layout(sink, tm, grouped, sorted_keys, has_midcap, midcap_by_trade,
                       has_calls, has_puts, filter_segments)
    if not sink.wrote:
        wb.remove(ws)


def _patch_wise_ops(
    tm, grouped, sorted_keys, has_midcap, midcap_by_trade,
    has_calls, has_puts, filter_segments=None,
) -> Optional[Dict]:
    """Rust path — return the Patch wise layout as a plain ops dict, or None when
    there are no patches (no sheet)."""
    sink = _OpsSink("Patch wise")
    _patch_wise_layout(sink, tm, grouped, sorted_keys, has_midcap, midcap_by_trade,
                       has_calls, has_puts, filter_segments)
    return sink.to_dict() if sink.cells else None


def _patch_wise_layout(
    sink,
    tm: Dict,
    grouped: Dict,
    sorted_keys: List[str],
    has_midcap: bool,
    midcap_by_trade: Optional[Dict],
    has_calls: bool,
    has_puts: bool,
    filter_segments: Optional[List] = None,
) -> None:
    """Patch-wise phase distribution sheet — mirrors buildTradeExcel.js Sheet 3."""

    GAP_MS = 30 * 86400000  # 30-day gap separates patches (fallback only)

    def _date_ms(val: Any) -> Optional[int]:
        dt = _parse_date(val)
        if dt is None:
            return None
        return int(calendar.timegm(dt.timetuple()) * 1000)

    tdata = []
    for k in sorted_keys:
        t = tm[k]
        legs = grouped.get(k, [])
        main = _main_leg(legs)
        spot = _to_num(main.get("Entry Spot")) or 0.0
        mc = (midcap_by_trade or {}).get(k) or {}
        # Phase uses whatever DIRECTIONAL leg(s) are present — option legs (CE and/or
        # PE) AND futures legs — so SELL PE / BUY PE / CE+PE / CE+FUT / FUT-only all
        # work. The patch-wise DD must reflect the COMBINED position (options +
        # futures); summing option P&L alone silently dropped a futures leg's P&L from
        # the drawdown. Options-only runs are unchanged (no FUT rows to add).
        opt_legs = [l for l in legs if str(l.get("Type") or "").upper() in ("CE", "CALL", "PE", "PUT", "FUT")]
        nifty_pnl = sum((_to_num(l.get("CE P&L")) or 0.0) + (_to_num(l.get("PE P&L")) or 0.0) + (_to_num(l.get("FUT P&L")) or 0.0) for l in opt_legs)
        nifty_mae = sum((_to_num(l.get("MAE")) or 0.0) for l in opt_legs) if opt_legs else None
        cfm = t.get("combinedFinalMae", "")
        entry = main.get("Entry Date", "")
        exit_ = main.get("Exit Date", "")
        tdata.append({
            "entry": entry, "exit": exit_,
            "entryMs": _date_ms(entry), "exitMs": _date_ms(exit_),
            "midcapPct":   _to_num(mc.get("Midcap Leg P&L %")),
            "midcapMae":   _to_num(mc.get("Midcap MAE")),
            "midcapClose": _to_num(mc.get("Midcap Entry Spot")),
            "callPct":     (nifty_pnl / spot * 100) if (opt_legs and spot != 0) else None,
            "callMae":     nifty_mae,
            "combinedPct": _to_num(mc.get("Combined Net P&L %")),
            "combinedMae": float(cfm) if isinstance(cfm, (int, float)) else None,
        })

    # Patches from the uploaded filter's segment START dates: a new patch begins
    # (equity resets to 100) when a trade's entry reaches the next segment start.
    # Boundary = next start (not the segment's end), so spot-adj cascades that
    # re-enter past a window's end stay in that patch. Falls back to 30-day gap
    # detection only when no filter segments are available.
    def _seg_start_ms(s):
        s = s or {}
        return _date_ms(s.get("start") or s.get("Start") or s.get("from")
                        or s.get("start_date") or s.get("startdt"))
    seg_starts = sorted(ms for ms in (_seg_start_ms(s) for s in (filter_segments or []))
                        if ms is not None)
    patches: List[List[Dict]] = []
    if seg_starts:
        cur_idx = -2
        for td in tdata:
            em = td["entryMs"]
            i = 0
            if em is not None:
                for j, sm in enumerate(seg_starts):
                    if sm <= em:
                        i = j
                    else:
                        break
            if i != cur_idx:
                patches.append([]); cur_idx = i
            patches[-1].append(td)
    else:
        last_exit_ms: Optional[int] = None
        for td in tdata:
            em = td["entryMs"]
            gap = (em - last_exit_ms) if (last_exit_ms is not None and em is not None) else 0
            if not patches or gap > GAP_MS:
                patches.append([])
            patches[-1].append(td)
            if td["exitMs"] is not None:
                last_exit_ms = td["exitMs"]

    if not patches:
        return

    def build_chain(trades, drive_of, mae_of):
        prev_cumm = 100.0; peak = 100.0; prev_peak = 100.0
        rows = []; pnl_sum = 0.0; live_dd_min = float("inf")
        for td in trades:
            dr = drive_of(td); d = dr if (dr is not None and math.isfinite(dr)) else 0.0
            cumm = prev_cumm * (1.0 + d / 100.0)
            peak = max(peak, cumm)
            dd = (cumm - peak) if peak > cumm else ""
            pct_dd = (dd / peak) if isinstance(dd, float) and peak != 0 else 0.0
            mv = mae_of(td); m = mv if (mv is not None and math.isfinite(mv)) else 0.0
            lowest_nav = prev_cumm * (1.0 + m / 100.0)
            live_dd = (lowest_nav / prev_peak - 1.0) * 100.0 if prev_peak != 0 else 0.0
            rows.append({"td": td, "drive": d, "cumm": cumm, "peak": peak,
                         "dd": dd, "pct_dd": pct_dd, "mae": m,
                         "lowest_nav": lowest_nav, "live_dd": live_dd})
            pnl_sum += d
            if live_dd < live_dd_min:
                live_dd_min = live_dd
            prev_cumm = cumm; prev_peak = peak
        if not rows:
            return {"rows": [], "entry": None, "exit": None, "cagr": None,
                    "pnl_sum": 0.0, "live_dd_min": None}
        f, l = trades[0], trades[-1]
        days = ((l["exitMs"] - f["entryMs"]) / 86400000.0
                if (f["entryMs"] is not None and l["exitMs"] is not None) else None)
        last = rows[-1]
        cagr = ((math.pow(last["cumm"] / 100.0, 365.0 / days) - 1.0) * 100.0
                if (days and days > 0 and last["cumm"] > 0) else None)
        return {"rows": rows, "entry": f["entry"], "exit": l["exit"],
                "cagr": cagr, "pnl_sum": pnl_sum,
                "live_dd_min": live_dd_min if live_dd_min != float("inf") else None}

    # Derived locally (not a param) so callers/signatures stay unchanged: does any
    # trade carry a futures leg? Drives the phase title so a mixed run reads e.g.
    # "Nifty CE + FUT" instead of hiding the futures leg.
    has_futures = any(str(l.get("Type") or "").upper() == "FUT"
                      for legs in grouped.values() for l in legs)
    _opt = " + ".join(n for n, h in (("CE", has_calls), ("PE", has_puts), ("FUT", has_futures)) if h) or "Options"
    nifty_title = f"Nifty {_opt}"
    _nifty_phase = {"title": nifty_title, "kind": "std", "dates": False,
         "drive": lambda td: td["callPct"], "mae": lambda td: td["callMae"],
         "detail_hdr": ["Net P&L %","Cumulative","Peak","DD","%DD","MAE","Lowest NAV","Actual Live DD"],
         "side_hdr": ["Entry","Exit","CAGR","Net P&L %","Live DD"]}
    PHASES = [
        {"title": "Midcap Future", "kind": "midcap", "dates": True,
         "drive": lambda td: td["midcapPct"], "mae": lambda td: td["midcapMae"],
         "detail_hdr": ["Entry Date","Exit Date","Midcap Hypo P&L %","cumm","Peak","Close","Hypo MAE","Lowest NAV","Live DD"],
         "side_hdr": ["Entry","Exit","CAGR","Future P&L %","Live DD"]},
        _nifty_phase,
        {"title": f"{nifty_title} + Midcap Future", "kind": "std", "dates": False,
         "drive": lambda td: td["combinedPct"], "mae": lambda td: td["combinedMae"],
         "detail_hdr": ["Net P&L %","Cumulative","Peak","DD","%DD","MAE","Lowest NAV","Actual Live DD"],
         "side_hdr": ["Entry","Exit","CAGR","Net P&L %","Live DD"]},
    ] if has_midcap else [_nifty_phase]

    sink.freeze(4, 0)

    def _hdr(r, c, val, bg=_HEADER_BG, tx=_WHITE_TXT, align="C"):
        sink.cell(r, c, val, bold=True, size=10, fc=tx, bg=bg, align=align, border=True)

    def _val(r, c, val, num_fmt=None):
        nf = num_fmt if (isinstance(val, (int, float)) and num_fmt) else None
        sink.cell(r, c, (val if val is not None else ""), bold=False, size=10,
                  fc="000000", bg=None, align="C", border=True, nfmt=nf)

    col = 1
    for phase in PHASES:
        chains = [build_chain(p, phase["drive"], phase["mae"]) for p in patches]
        dW = len(phase["detail_hdr"])
        detail_start = col
        side_start = col + dW + 1

        # Row 1 — block title
        sink.cell(1, detail_start, phase["title"], bold=True, size=11,
                  fc=_WHITE_TXT, bg=_NAVY_BG, align="L")
        sink.merge(1, detail_start, detail_start + dW - 1)
        # Row 2 — subtitle
        sink.cell(2, detail_start, "Phase wise Distribution", bold=True, size=9,
                  fc=_SUB_HDR_TX, bg=_SUB_HDR_BG, align="L")
        sink.merge(2, detail_start, detail_start + dW - 1)
        # Row 4 — detail headers
        for i, h in enumerate(phase["detail_hdr"]):
            _hdr(4, detail_start + i, h)
        # Side table headers
        for i, h in enumerate(phase["side_hdr"]):
            _hdr(4, side_start + i, h, bg=_SECTION_BG)

        # Detail rows (row 5+)
        rr = 5
        for ch in chains:
            for rw in ch["rows"]:
                c2 = detail_start
                if phase["dates"]:
                    _val(rr, c2, rw["td"]["entry"]); c2 += 1
                    _val(rr, c2, rw["td"]["exit"]); c2 += 1
                if phase["kind"] == "midcap":
                    _val(rr, c2, rw["drive"]); c2 += 1
                    _val(rr, c2, rw["cumm"]); c2 += 1
                    _val(rr, c2, rw["peak"]); c2 += 1
                    _val(rr, c2, rw["td"].get("midcapClose")); c2 += 1
                    _val(rr, c2, rw["mae"], '0.00"%"'); c2 += 1
                    _val(rr, c2, rw["lowest_nav"]); c2 += 1
                    _val(rr, c2, rw["live_dd"], '0.00"%"'); c2 += 1
                else:
                    _val(rr, c2, rw["drive"]); c2 += 1
                    _val(rr, c2, rw["cumm"]); c2 += 1
                    _val(rr, c2, rw["peak"]); c2 += 1
                    _val(rr, c2, rw["dd"] if rw["dd"] != "" else None); c2 += 1
                    _val(rr, c2, rw["pct_dd"]); c2 += 1
                    _val(rr, c2, rw["mae"]); c2 += 1
                    _val(rr, c2, rw["lowest_nav"]); c2 += 1
                    _val(rr, c2, rw["live_dd"]); c2 += 1
                rr += 1

        # Side table rows
        for i, ch in enumerate(chains):
            sr = 5 + i; c3 = side_start
            _val(sr, c3, ch["entry"]); c3 += 1
            _val(sr, c3, ch["exit"]); c3 += 1
            _val(sr, c3, ch["cagr"], '0.00"%"'); c3 += 1
            _val(sr, c3, ch["pnl_sum"]); c3 += 1
            _val(sr, c3, ch["live_dd_min"]); c3 += 1

        # Column widths
        for i in range(dW):
            sink.col_width(detail_start + i, 12)
        for i in range(len(phase["side_hdr"])):
            sink.col_width(side_start + i, 12)

        col = side_start + len(phase["side_hdr"]) + 1


# ── Sheet 1: Trade Sheet ──────────────────────────────────────────────────────

def _rules_layout(sink, rules_sheet: List) -> None:
    """Emit the leg-wise "Rules" sheet through a layout sink.

    `rules_sheet` is a list of typed rows built client-side from the strategy
    payload (frontend buildRulesSheet), so the sheet reflects the full
    configuration: strategy-level rules plus one section per leg (options CE/PE,
    futures, or midcap overlay) with that leg's own strike/SL/target/slippage.
    Row forms: ["title", text] · ["section", text] · ["kv", label, value] ·
    ["spacer"].
    """
    sink.col_width(1, 34)
    sink.col_width(2, 58)
    r = 1
    for row in (rules_sheet or []):
        if not row:
            continue
        kind = str(row[0] or "")
        if kind == "title":
            sink.merge(r, 1, 2)
            sink.cell(r, 1, (row[1] if len(row) > 1 else ""),
                      bold=True, size=14, fc=_WHITE_TXT, bg=_NAVY_BG, align="C")
            sink.row_height(r, 28)
        elif kind == "section":
            sink.merge(r, 1, 2)
            sink.cell(r, 1, "  " + (row[1] if len(row) > 1 else ""),
                      bold=True, size=11, fc=_WHITE_TXT, bg=_SECTION_BG)
            sink.row_height(r, 20)
        elif kind == "kv":
            label = row[1] if len(row) > 1 else ""
            value = row[2] if len(row) > 2 else ""
            sink.cell(r, 1, label, bold=True, size=10, fc=_DARK2_TXT,
                      bg=_LABEL_BG, border=True)
            sink.cell(r, 2, ("" if value is None else value), size=10,
                      fc=_DARK_TXT, bg=_WHITE, border=True)
            sink.row_height(r, 18)
        elif kind == "spacer":
            sink.row_height(r, 6)
        else:
            continue
        r += 1
    sink.freeze(1, 0)


def _write_rules_sheet(wb: Workbook, rules_sheet: List) -> None:
    """Render "Rules" as the FIRST tab of an openpyxl workbook."""
    ws = wb.create_sheet("Rules", 0)  # index 0 → first tab
    _rules_layout(_OpenpyxlSink(ws), rules_sheet)


def _rules_ops(rules_sheet: List) -> Dict:
    """Same sheet, serialized for the Rust layout writer."""
    sink = _OpsSink("Rules")
    _rules_layout(sink, rules_sheet)
    return sink.to_dict()


def _write_trade_sheet(wb: Workbook, cleaned: List[Dict], key_order: List[str]) -> None:
    ws = wb.create_sheet("Trade Sheet")
    ws.freeze_panes = "A2"

    for ci, key in enumerate(key_order, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTHS.get(key, 10)

    # Header
    hdr_row = ws.row_dimensions[1]
    hdr_row.height = 22
    for ci, key in enumerate(key_order, 1):
        cell = ws.cell(row=1, column=ci, value=key)
        cell.font      = _font(bold=True, size=10, color=_WHITE_TXT)
        cell.fill      = _fill(_HEADER_BG)
        cell.alignment = _CENTER
        cell.border    = _border()

    # Data rows
    for ri, row in enumerate(cleaned, 2):
        bg = _WHITE if ri % 2 == 0 else _ALT_ROW
        net_val = row.get("Net P&L")
        net_num = _to_num(net_val)

        ws.row_dimensions[ri].height = 18
        for ci, key in enumerate(key_order, 1):
            raw = row.get(key, "")
            cell = ws.cell(row=ri, column=ci)
            cell.border    = _border()
            cell.fill      = _fill(bg)
            cell.font      = _font(size=10)
            cell.alignment = _LEFT

            # Date columns
            if key in _DATE_COLS:
                d = _parse_date(raw) if not isinstance(raw, datetime) else raw
                if d:
                    cell.value  = d
                    cell.number_format = "DD-MMM-YYYY"
                else:
                    cell.value = raw if raw != "" else None
                continue

            # Numeric coercion
            if raw != "" and raw is not None:
                num = _to_num(raw)
                if num is not None:
                    cell.value = num
                    if key in _TRUE_PCT_COLS:
                        cell.number_format = "0.00%"
                    elif key in _MAE_COLS:
                        cell.number_format = "#,##0.0000"
                    elif num == int(num):
                        cell.number_format = "0"
                    else:
                        cell.number_format = "#,##0.00"
                    continue
            cell.value = raw if raw not in ("", None) else None

        # Color Net P&L and % P&L columns (NIFTY-only sheet)
        if net_num is not None:
            for col_key in ("Net P&L", "% P&L"):
                try:
                    ci2 = key_order.index(col_key) + 1
                    c = ws.cell(row=ri, column=ci2)
                    clr = _GREEN_TX if net_num >= 0 else _RED_TX
                    bg2 = _GREEN_BG if net_num >= 0 else _RED_BG
                    c.font = _font(bold=True, size=10, color=clr)
                    c.fill = _fill(bg2)
                except ValueError:
                    pass
        # Color Combined Net P&L / % columns (Midcap sheet)
        c_net = _to_num(row.get("Combined Net P&L"))
        if c_net is not None:
            for col_key in ("Combined Net P&L", "Combined Net P&L %"):
                try:
                    ci2 = key_order.index(col_key) + 1
                    c = ws.cell(row=ri, column=ci2)
                    clr = _GREEN_TX if c_net >= 0 else _RED_TX
                    bg2 = _GREEN_BG if c_net >= 0 else _RED_BG
                    c.font = _font(bold=True, size=10, color=clr)
                    c.fill = _fill(bg2)
                except ValueError:
                    pass


# ── Sheet 2: Summary ─────────────────────────────────────────────────────────

class _OpenpyxlSink:
    """Writes layout ops directly into an openpyxl worksheet (the reference path)."""
    def __init__(self, ws):
        self.ws = ws
        self.wrote = False

    def cell(self, r, c, v, bold=False, size=10, fc="000000", bg=None,
             align="L", border=False, nfmt=None):
        self.wrote = True
        cell = self.ws.cell(row=r, column=c, value=v)
        cell.font = _font(bold=bold, size=size, color=fc)
        if bg is not None:
            cell.fill = _fill(bg)
        cell.alignment = _CENTER if align == "C" else _LEFT
        if border:
            cell.border = _border()
        if nfmt:
            cell.number_format = nfmt

    def merge(self, r, c1, c2):
        self.ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

    def row_height(self, r, h):
        self.ws.row_dimensions[r].height = h

    def col_width(self, c, w):
        self.ws.column_dimensions[get_column_letter(c)].width = w

    def freeze(self, nrows, ncols):
        # nrows/ncols = count of frozen rows/cols (rust set_freeze_panes semantics).
        self.ws.freeze_panes = f"{get_column_letter(ncols + 1)}{nrows + 1}"


class _OpsSink:
    """Records layout ops as plain data for the Rust layout writer
    (algotest_native.write_layout_sheet_xlsx). Same call surface as _OpenpyxlSink so
    the shared _summary_layout builds an identical sheet through either backend."""
    def __init__(self, name):
        self.name = name
        self.cells = []
        self.merges = []
        self.row_heights = []
        self.col_widths = []
        self._freeze = None

    def cell(self, r, c, v, bold=False, size=10, fc="000000", bg=None,
             align="L", border=False, nfmt=None):
        self.cells.append({"r": r, "c": c, "v": v, "bold": bold, "size": size,
                           "fc": fc, "bg": bg, "align": align, "border": border,
                           "nfmt": nfmt})

    def merge(self, r, c1, c2):
        self.merges.append((r, c1, c2))

    def row_height(self, r, h):
        self.row_heights.append((r, h))

    def col_width(self, c, w):
        self.col_widths.append((c, w))

    def freeze(self, nrows, ncols):
        self._freeze = (nrows, ncols)

    def to_dict(self):
        return {"name": self.name, "cells": self.cells, "merges": self.merges,
                "row_heights": self.row_heights, "col_widths": self.col_widths,
                "freeze": self._freeze}


def _write_summary_sheet(
    wb: Workbook,
    cleaned: List[Dict],
    summary: Dict[str, Any],
    tm: Dict,
    combo_label: str,
    from_date: str,
    to_date: str,
    has_calls: bool,
    has_puts: bool,
    has_futures: bool,
    has_midcap: bool = False,
    midcap_summary: Optional[Dict] = None,
    chron_keys: Optional[List[str]] = None,
    patchwise: bool = False,
    filter_segments: Optional[List] = None,
) -> None:
    """openpyxl reference path — build the Summary sheet into wb."""
    ws = wb.create_sheet("Summary")
    _summary_layout(_OpenpyxlSink(ws), cleaned, summary, tm, combo_label,
                    from_date, to_date, has_calls, has_puts, has_futures,
                    has_midcap, midcap_summary, chron_keys, patchwise, filter_segments)


def _summary_ops(
    cleaned, summary, tm, combo_label, from_date, to_date,
    has_calls, has_puts, has_futures, has_midcap=False,
    midcap_summary=None, chron_keys=None, patchwise=False, filter_segments=None,
) -> Dict:
    """Rust path — return the Summary layout as a plain ops dict (no openpyxl)."""
    sink = _OpsSink("Summary")
    _summary_layout(sink, cleaned, summary, tm, combo_label, from_date, to_date,
                    has_calls, has_puts, has_futures, has_midcap, midcap_summary,
                    chron_keys, patchwise, filter_segments)
    return sink.to_dict()


def _summary_layout(
    sink,
    cleaned: List[Dict],
    summary: Dict[str, Any],
    tm: Dict,
    combo_label: str,
    from_date: str,
    to_date: str,
    has_calls: bool,
    has_puts: bool,
    has_futures: bool,
    has_midcap: bool = False,
    midcap_summary: Optional[Dict] = None,
    chron_keys: Optional[List[str]] = None,
    patchwise: bool = False,
    filter_segments: Optional[List] = None,
) -> None:
    sink.col_width(1, 30)
    sink.col_width(2, 20)
    sink.col_width(3, 12)
    sink.col_width(4, 30)
    sink.col_width(5, 20)

    S = summary or {}
    row = [1]

    # Patchwise reset boundaries (segment START dates), shared by the Max DD scan
    # and the outlier Live DD scan so they reset at the same points as the combined
    # chain. Each cleaned row carries its own Entry Date → segment index.
    def _pw_seg_start_ms(s):
        s = s or {}
        return _date_ms(s.get("start") or s.get("Start") or s.get("from")
                        or s.get("start_date") or s.get("startdt"))
    _pw_seg_starts = sorted(ms for ms in (_pw_seg_start_ms(s) for s in (filter_segments or []))
                            if ms is not None)

    def _pw_row_seg_idx(entry_val) -> int:
        em = _date_ms(entry_val)
        i = -1
        if em is not None:
            for j, sm in enumerate(_pw_seg_starts):
                if sm <= em:
                    i = j
                else:
                    break
        return i

    def _merge(r, c1=1, c2=5):
        sink.merge(r, c1, c2)

    def _title(text, r, bg=_NAVY_BG):
        _merge(r)
        sink.cell(r, 1, text, bold=True, size=13, fc=_WHITE_TXT, bg=bg, align="C")
        sink.row_height(r, 26)

    def _section(text, r):
        _merge(r)
        sink.cell(r, 1, "  " + text, bold=True, size=11, fc=_WHITE_TXT,
                  bg=_SECTION_BG, align="L")
        sink.row_height(r, 20)

    def _kv(label, value, r, col="A", alt=False, val_color=None):
        col_idx = ord(col.upper()) - ord("A") + 1
        num = _to_num(str(value or "").replace("+", "").replace("%", "").replace("₹", ""))
        auto_color = val_color or (_GREEN_TX if (num is not None and num >= 0) else _RED_TX if (num is not None and num < 0) else _DARK_TXT)
        sink.cell(r, col_idx, label, bold=True, size=10, fc=_DARK2_TXT,
                  bg=(_ALT_ROW if alt else _LABEL_BG), align="L", border=True)
        sink.cell(r, col_idx + 1, value, bold=True, size=10, fc=auto_color,
                  bg=(_ALT_ROW if alt else _WHITE), align="L", border=True)
        sink.row_height(r, 18)

    # ── Compute stats from cleaned rows (mirrors JS) ──────────────────────────
    sum_pct = 0.0; sum_pos_pct = 0.0; sum_neg_pct = 0.0
    win_cnt = 0;   loss_cnt = 0;      total_cnt = 0
    sum_net = 0.0; max_net = -math.inf; min_net = math.inf
    final_cum = 100.0; spot_cum = 100.0
    min_entry_ms = None; max_exit_ms = None
    spot_sum_gated = 0.0
    ce_sum = 0.0; pe_sum = 0.0; fut_sum = 0.0
    ce_pct = 0.0; pe_pct = 0.0; spot_pct = 0.0; fut_pct = 0.0

    def _parse_date_ms(v):
        d = _parse_date(v)
        return d.timestamp() * 1000 if d else None

    for t in cleaned:
        ce  = _to_num(t.get("CE P&L"));   ce_sum  += ce  if ce  is not None else 0
        pe  = _to_num(t.get("PE P&L"));   pe_sum  += pe  if pe  is not None else 0
        fu  = _to_num(t.get("FUT P&L"));  fut_sum += fu  if fu  is not None else 0
        cep = _to_num(t.get("CE P&L %")); ce_pct  += cep if cep is not None else 0
        pep = _to_num(t.get("PE P&L %")); pe_pct  += pep if pep is not None else 0
        spp = _to_num(t.get("Spot P&L %")); spot_pct += spp if spp is not None else 0
        # FUT P&L % is not a stored column (computed per-row at write time), so derive
        # it the same way — FUT P&L / Entry Spot — and accumulate for the Summary total.
        _es = _to_num(t.get("Entry Spot"))
        if fu is not None and _es not in (None, 0):
            fut_pct += fu / _es

    # With a Midcap leg, ALL Performance Overview stats run on the COMBINED
    # (NIFTY + Midcap) per-trade P&L; otherwise NIFTY (unchanged). Combined
    # values live on first-leg rows as the Combined columns.
    def _gp(t):
        return _to_num(t.get("Combined Net P&L %")) if has_midcap else _to_num(t.get("% P&L"))

    def _gn(t):
        return _to_num(t.get("Combined Net P&L")) if has_midcap else _to_num(t.get("Net P&L"))

    def _gc(t):
        return _to_num(t.get("Combined Cumulative")) if has_midcap else _to_num(t.get("Cumulative"))

    _init_spot = None; _final_spot = None   # cagr_spot from spot LEVELS (see below)
    for t in cleaned:
        p = _gp(t); n = _gn(t)
        if p is not None and math.isfinite(p):
            sum_pct += p; total_cnt += 1
            if p > 0:  sum_pos_pct += p; win_cnt  += 1
            elif p < 0: sum_neg_pct += p; loss_cnt += 1
        if n is not None and math.isfinite(n):
            sum_net += n
            if n > max_net: max_net = n
            if n < min_net: min_net = n
            sp = _to_num(t.get("Spot P&L"))
            if sp is not None: spot_sum_gated += sp
        cum = _gc(t)
        if cum is not None and math.isfinite(cum): final_cum = cum
        es = _to_num(t.get("Entry Spot")); xs = _to_num(t.get("Exit Spot"))
        # cagr_spot uses spot LEVELS (leg-independent, base.py:1075): first trade's
        # Entry Spot and last trade's Exit Spot in canonical order, NOT compounded
        # per-trade ratios. `cleaned` is already in canonical order.
        if es and es > 0 and _init_spot is None: _init_spot = es
        if xs and xs > 0: _final_spot = xs
        ed = _parse_date_ms(t.get("Entry Date")); xd = _parse_date_ms(t.get("Exit Date"))
        if ed and (min_entry_ms is None or ed < min_entry_ms): min_entry_ms = ed
        if xd and (max_exit_ms  is None or xd > max_exit_ms):  max_exit_ms  = xd

    if not math.isfinite(max_net): max_net = 0.0
    if not math.isfinite(min_net): min_net = 0.0

    avg_win_pct   = (sum_pos_pct / win_cnt)   if win_cnt  > 0 else 0.0
    avg_loss_pct  = (sum_neg_pct / loss_cnt)  if loss_cnt > 0 else 0.0
    win_rate      = (win_cnt  / total_cnt * 100) if total_cnt > 0 else 0.0
    loss_rate     = (loss_cnt / total_cnt * 100) if total_cnt > 0 else 0.0
    avg_net       = (sum_net  / total_cnt)       if total_cnt > 0 else 0.0
    avg_pct       = (sum_pct  / total_cnt)       if total_cnt > 0 else 0.0
    if avg_loss_pct != 0:
        expectancy = (
            ((win_rate / 100) * avg_win_pct - (loss_rate / 100) * abs(avg_loss_pct))
            / abs(avg_loss_pct)
        )
    else:
        expectancy = 0.0
    # Year span + CAGRs use the backtest convention (base.py:999,1075), identical to
    # compute_xlsx_summary_metrics so master == this per-combo sheet == backtest:
    # integer days between first entry and last exit / 365.0, floored 0.01; cagr_spot
    # from spot LEVELS; CAGR(options) clamped +/-99999, -100 on a wiped-out equity.
    _span_days = (round((max_exit_ms - min_entry_ms) / (86400 * 1000))
                  if (min_entry_ms is not None and max_exit_ms is not None) else 0)
    years = max(_span_days / 365.0, 0.01)
    opt_cagr = (
        max(-99999.0, min(99999.0, (math.pow(final_cum / 100, 1 / years) - 1) * 100))
        if (years > 0 and final_cum > 0) else -100.0
    )
    # (cagr_options is taken from the single engine in the AUTHORITATIVE READ block
    # below — unconditionally, so patchwise/midcap are covered too.)
    # SINGLE SOURCE OF TRUTH: the BACKTEST owns cagr_spot (base.py compute_analytics),
    # mirroring summary_metrics.rs. Re-deriving it here walked the per-LEG rows and on a
    # multi-index run paired one index's entry spot with the other's exit spot — two
    # unrelated price scales (measured: -68.49 vs the backtest's +5.82). This layout fn
    # feeds BOTH renderers (openpyxl _write_summary_sheet and the Rust _summary_ops
    # writer), so pinning it here is the one place that fixes the sheet. The derivation
    # survives only as a fallback for callers that pass no summary.
    # `is None`, NOT `or` — must match summary_metrics.rs's unwrap_or_else, which fires
    # only when the key is ABSENT. With `or`, a genuine cagr_spot of 0.0 fell through to
    # the derivation, so Rust reported 0.00 while Python reported the (cross-index)
    # derived value — a fresh divergence on exactly the sync_weekly_roll path.
    _S_cagr_spot = _to_num(S.get("cagr_spot"))
    spot_cagr = _S_cagr_spot if _S_cagr_spot is not None else (
        100 * ((_final_spot / _init_spot) ** (1.0 / years) - 1)
        if (years > 0 and _init_spot and _final_spot and _init_spot > 0 and _final_spot > 0)
        else 0.0
    )
    max_dd_pct   = _to_num(S.get("max_dd_pct")) or 0.0
    max_dd_pts   = _to_num(S.get("max_dd_pts")) or 0.0
    max_win_str  = _to_num(S.get("max_win_streak"))  or 0
    max_loss_str = _to_num(S.get("max_loss_streak")) or 0
    mdd_start    = S.get("mdd_start_date") or ""
    mdd_end      = S.get("mdd_end_date")   or ""
    mdd_dur      = _to_num(S.get("mdd_duration_days")) or ""

    # Max Drawdown = the single worst %DD on the equity curve — Combined NAV with a
    # Midcap leg, NIFTY NAV otherwise — i.e. min over trades of (Cumulative/Peak − 1)×100.
    # Computed straight from the Cumulative/Peak columns (units-safe) so the Summary can
    # NEVER diverge from the per-trade %DD column, in BOTH overall and patchwise (those
    # columns already carry the right peak basis: continuous for overall, per-patch reset
    # for patchwise). Applies to midcap AND non-midcap alike.
    _cum_col  = "Combined Cumulative" if has_midcap else "Cumulative"
    _peak_col = "Combined Peak"       if has_midcap else "Peak"
    _pct_col  = "Combined Net P&L %"  if has_midcap else "% P&L"
    worst_dd = 0.0
    worst_peak_ms = None; worst_trough_ms = None; peak_ms = None
    win_run = loss_run = mx_win = mx_loss = 0
    for t in cleaned:
        cp = _to_num(t.get(_pct_col))
        if cp is not None and math.isfinite(cp):
            if cp > 0:   win_run += 1; loss_run = 0; mx_win  = max(mx_win, win_run)
            elif cp < 0: loss_run += 1; win_run = 0; mx_loss = max(mx_loss, loss_run)
        cc = _to_num(t.get(_cum_col))
        pk = _to_num(t.get(_peak_col))
        xd = _parse_date_ms(t.get("Exit Date"))
        if cc is not None and pk not in (None, 0) and math.isfinite(cc):
            if cc >= pk - 1e-9:
                peak_ms = xd
            else:
                ddp = (cc / pk - 1) * 100
                if ddp < worst_dd:
                    worst_dd = ddp; worst_trough_ms = xd; worst_peak_ms = peak_ms
    max_dd_pct = worst_dd
    # Overall + non-midcap: pin the magnitude to the backtest's OWN max_dd_pct
    # (min of base.py's %DD column) so per-combo == backtest byte-for-byte; the loop
    # above still supplies the MDD peak/trough dates. Patchwise/midcap keep worst_dd.
    if not patchwise and not has_midcap:
        _s_dd = _to_num(summary.get("max_dd_pct"))
        if _s_dd is not None and math.isfinite(_s_dd):
            max_dd_pct = _s_dd
    max_win_str = mx_win; max_loss_str = mx_loss
    if worst_peak_ms and worst_trough_ms:
        mdd_dur = round((worst_trough_ms - worst_peak_ms) / 86400000)
        mdd_start = datetime.utcfromtimestamp(worst_peak_ms / 1000).strftime("%Y-%m-%d")
        mdd_end   = datetime.utcfromtimestamp(worst_trough_ms / 1000).strftime("%Y-%m-%d")
    else:
        mdd_dur = 0; mdd_start = ""; mdd_end = ""

    car_mdd = opt_cagr / abs(max_dd_pct) if max_dd_pct != 0 else 0.0
    # (car_mdd is taken from the single engine in the AUTHORITATIVE READ block below.)

    opt_sum = (
        (ce_sum + pe_sum)   if (has_calls and has_puts) else
        pe_sum              if has_puts else
        ce_sum              if has_calls else
        fut_sum             if has_futures else sum_net
    )
    _scp = _to_num(S.get("spot_change_pct")) or (spot_pct * 100) or 0.0
    roi_pct = sum_pct / abs(_scp) if _scp != 0 else 0.0

    # Live DD outlier analysis — iterate trades chronologically (same order as
    # the cleaned rows / Live DD pass above) so cascade trades are placed in
    # the right time sequence for the outlier-stripped DD computation.
    # Walk trades in the CANONICAL chronological order (same order used to compute
    # each trade's combinedActualLDD / actualLDD in _aggregate_trades). When the
    # caller passes chron_keys (the _aggregate_trades sorted_keys), use it so the
    # outlier-stripped reconstruction is internally consistent with the base Live
    # DD AND identical to the master summary. Fall back to the old cleaned-order
    # logic only if no chron_keys was provided (backward-safe).
    trade_pairs = []
    _seen2: set = set()
    _chron_keys: List[str] = []
    if chron_keys:
        for _k in chron_keys:
            if _k not in _seen2 and _k in tm:
                _seen2.add(_k); _chron_keys.append(_k)
    else:
        for _cr in cleaned:
            _k = str(_cr.get("Trade") or _cr.get("trade") or 1)
            if _k not in _seen2 and _k in tm:
                _seen2.add(_k)
                _chron_keys.append(_k)
    # Fallback to integer ordering for any trades not represented above
    # (shouldn't normally happen, but keeps logic safe).
    for _k in tm.keys():
        if _k not in _seen2:
            _seen2.add(_k); _chron_keys.append(_k)
    # First-seen Entry Date per trade key (for patchwise segment bucketing).
    _key_entry: Dict[str, Any] = {}
    for _cr in cleaned:
        _k = str(_cr.get("Trade") or _cr.get("trade") or 1)
        if _k not in _key_entry and _cr.get("Entry Date"):
            _key_entry[_k] = _cr.get("Entry Date")
    for k in _chron_keys:
        t2 = tm[k]
        # Combined per-trade values when a Midcap leg is present; NIFTY otherwise.
        _pk = "combinedPct"      if has_midcap else "pct"
        _lk = "combinedActualLDD" if has_midcap else "actualLDD"
        _mk = "combinedFinalMae"  if has_midcap else "finalMae"
        pct_v = t2.get(_pk); pct_v = pct_v if isinstance(pct_v, float) and math.isfinite(pct_v) else None
        ldd_v = t2.get(_lk); ldd_v = ldd_v if isinstance(ldd_v, float) and math.isfinite(ldd_v) else None
        mae_v = t2.get(_mk); mae_v = mae_v if isinstance(mae_v, float) and math.isfinite(mae_v) else None
        if pct_v is not None:
            trade_pairs.append({"pct": pct_v, "ldd": ldd_v, "mae": mae_v, "idx": len(trade_pairs),
                                 "exitReason": (t2.get("exitReason") or "").upper(),
                                 "segIdx": _pw_row_seg_idx(_key_entry.get(k))})

    n_trades  = len(trade_pairs)
    by_pct_desc = sorted(trade_pairs, key=lambda x: -x["pct"])

    # READ, don't re-derive (see the unification block in build_combo_xlsx). The local
    # derivation ranks EVERY trade together, while the authoritative engine ranks over
    # the patch-aware chain — so on a patchwise sheet the two disagreed
    # (positive_outlier_1 3.5 here vs the engine's 1.7396). Kept as the fallback for
    # callers that pass no metrics.
    def _outl(key, computed):
        v = _to_num(S.get(key))
        return v if v is not None else computed

    _p1  = _outl("positive_outlier_1", by_pct_desc[0]["pct"] if n_trades > 0 else 0.0)
    _p2  = _outl("positive_outlier_2",
                 (_p1 + by_pct_desc[1]["pct"]) if n_trades > 1 else _p1)
    _p3  = _outl("positive_outlier_3",
                 (_p2 + by_pct_desc[2]["pct"]) if n_trades > 2 else _p2)
    _n1  = _outl("negative_outlier_1",
                 by_pct_desc[n_trades - 1]["pct"] if n_trades > 0 else 0.0)
    _n2  = _outl("negative_outlier_2",
                 (_n1 + by_pct_desc[n_trades - 2]["pct"]) if n_trades > 1 else _n1)
    _n3  = _outl("negative_outlier_3",
                 (_n2 + by_pct_desc[n_trades - 3]["pct"]) if n_trades > 2 else _n2)
    total_pct_sum = sum(p["pct"] for p in trade_pairs)
    pct_no_o1 = total_pct_sum - _p1 - _n1
    pct_no_o2 = total_pct_sum - _p2 - _n2
    pct_no_o3 = total_pct_sum - _p3 - _n3

    def _ldd_exc_stats(exc_top, exc_bot):
        exc_idx = {
            *[p["idx"] for p in by_pct_desc[:exc_top]],
            *[p["idx"] for p in by_pct_desc[max(0, n_trades - exc_bot):]],
        }
        filtered = [p for p in trade_pairs if p["idx"] not in exc_idx]
        if not filtered:
            return (0.0, 0.0)
        cumulative = 100.0
        peak = 100.0
        prev_cum = 100.0
        prev_peak = 100.0
        prev_exit_reason = ""
        prev_seg_idx = None
        ldds = []
        for p in filtered:
            # Reset the chain at each patch boundary (same boundary as the combined chain).
            if patchwise and (
                (prev_seg_idx is not None and p.get("segIdx") != prev_seg_idx) if _pw_seg_starts
                else ("FILTER_END" in (prev_exit_reason or "").split("+"))
            ):
                cumulative = 100.0; peak = 100.0; prev_cum = 100.0; prev_peak = 100.0
            prev_seg_idx = p.get("segIdx")
            pct = p["pct"]
            prev_peak = peak
            cumulative *= (1.0 + pct / 100.0)
            peak = max(peak, cumulative)
            mae = p["mae"]
            if mae is not None and prev_peak != 0:
                lowest_nav = round(prev_cum * (1.0 + mae / 100.0) * 100) / 100
                actual_ldd = round((lowest_nav / prev_peak - 1) * 10000) / 100
                ldds.append(actual_ldd)
            prev_cum = cumulative
            prev_exit_reason = p.get("exitReason") or ""
        if not ldds:
            return (0.0, 0.0)
        return (round(min(ldds), 2), round(sum(ldds) / len(ldds), 2))

    all_ldds   = [p["ldd"] for p in trade_pairs if p["ldd"] is not None]
    live_dd_min = round(min(all_ldds), 2) if all_ldds else 0.0
    live_dd_avg = round(sum(all_ldds) / len(all_ldds), 2) if all_ldds else 0.0
    # Avg (Combined) Final MAE — mean of each trade's Final MAE (Combined when a Midcap
    # leg is present, NIFTY otherwise; trade_pairs["mae"] already holds the right one).
    _final_maes = [p["mae"] for p in trade_pairs if p["mae"] is not None]
    avg_final_mae = round(sum(_final_maes) / len(_final_maes), 2) if _final_maes else 0.0
    # READ, don't re-derive: build_combo_xlsx folded the Rust engine's metrics into `S`,
    # so these come from the SAME computation the master summary renders. _ldd_exc_stats
    # survives only as the fallback for callers that hand us no metrics (it is the
    # implementation that disagreed with both Rust and the Python reference on the
    # patchwise chain — -1.56 vs the -1.81 those two agree on).
    def _ldd_pair(n, computed):
        _mn = _to_num(S.get(f"outlier_dd_{n}"))
        _av = _to_num(S.get(f"outlier_dd_{n}_avg"))
        return (_mn, _av) if (_mn is not None and _av is not None) else computed

    ldd_no_o1  = _ldd_pair(1, _ldd_exc_stats(1, 1))
    ldd_no_o2  = _ldd_pair(2, _ldd_exc_stats(2, 2))
    ldd_no_o3  = _ldd_pair(3, _ldd_exc_stats(3, 3))
    car_mdd_live = opt_cagr / abs(live_dd_min) if live_dd_min != 0 else 0.0

    # ── AUTHORITATIVE READ ────────────────────────────────────────────────────────
    # Everything above this line is now only a FALLBACK for callers that hand us no
    # metrics. build_combo_xlsx folded the single engine's output into `S`, and that
    # output is already mode-correct — Rust's `pin` returns the backtest's value for a
    # plain run and its own per-patch / Combined value for patchwise / midcap — so we
    # take it unconditionally. Reading it here (rather than gating on `not patchwise`)
    # is what removes _summary_layout as an independent implementation: on a patchwise
    # sheet the local derivations disagreed with the engine on cagr_options (23.27 vs
    # 7.81), max_dd_pct (-9.42 vs -3.09) and the avg_* family.
    def _auth(key, computed):
        v = _to_num(S.get(key))
        return v if v is not None else computed

    opt_cagr      = _auth("cagr_options", opt_cagr)
    max_dd_pct    = _auth("max_dd_pct", max_dd_pct)
    car_mdd       = _auth("car_mdd", car_mdd)
    car_mdd_live  = _auth("car_mdd_live", car_mdd_live)
    avg_win_pct   = _auth("avg_win_pct", avg_win_pct)
    avg_loss_pct  = _auth("avg_loss_pct", avg_loss_pct)
    avg_pct       = _auth("avg_profit_per_trade_pct", avg_pct)
    live_dd_min   = _auth("actual_live_dd_max", live_dd_min)
    live_dd_avg   = _auth("actual_live_dd_avg", live_dd_avg)
    avg_final_mae = _auth("avg_final_mae", avg_final_mae)

    def _fmt_pct(v, signed=True):
        prefix = "+" if (signed and v >= 0) else ""
        return f"{prefix}{float(v):.2f}%"

    def _fmt_cur(v):
        return f"₹{float(v):,.2f}"

    # ── Row 1: Title ───────────────────────────────────────────────────────────
    _title("  BACKTEST SUMMARY REPORT", 1)

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    _merge(2)
    parts = []
    if combo_label: parts.append(combo_label)
    if from_date or to_date:
        parts.append(f"{from_date or ''}{' → ' if from_date and to_date else ''}{to_date or ''}")
    parts.append(f"Generated: {datetime.now().strftime('%d %b %Y')}")
    sink.cell(2, 1, "   ·   ".join(parts), bold=False, size=10, fc="555555",
              bg=_SUB_HDR_BG, align="C")
    sink.row_height(2, 16)

    r = 4
    # ── SECTION 1: Performance Overview ──────────────────────────────────────
    _section("PERFORMANCE OVERVIEW", r); r += 1

    _kv("Overall Profit", _fmt_pct(sum_pct), r, "A", False, _GREEN_TX if sum_pct >= 0 else _RED_TX)
    _kv("No. of Trades",  int(total_cnt),   r, "D", False, _DARK_TXT); r += 1

    _kv("Win %",  f"{win_rate:.2f}%",  r, "A", True, _GREEN_TX)
    _kv("Loss %", f"{loss_rate:.2f}%", r, "D", True, _RED_TX); r += 1

    _kv("Avg Profit on Winners", f"{avg_win_pct:.2f}%",  r, "A", False, _GREEN_TX)
    _kv("Avg Loss on Losers",    f"{avg_loss_pct:.2f}%", r, "D", False, _RED_TX); r += 1

    sign = "+" if avg_net >= 0 else ""
    _kv("Avg Profit per Trade",  f"{sign}{avg_net:.2f}", r, "A", True, _GREEN_TX if avg_net >= 0 else _RED_TX)
    _kv("Expectancy Ratio",      f"{expectancy:.4f}",    r, "D", True, _GREEN_TX if expectancy >= 0 else _RED_TX); r += 1
    sign_pct = "+" if avg_pct >= 0 else ""
    _kv("Net P/L Avg %",  f"{sign_pct}{avg_pct:.4f}%", r, "A", False, _GREEN_TX if avg_pct >= 0 else _RED_TX); r += 1

    _kv("Max Profit (Single Trade)", _fmt_cur(max_net), r, "A", False, _GREEN_TX)
    _kv("Max Loss (Single Trade)",   _fmt_cur(min_net), r, "D", False, _RED_TX); r += 1

    _kv("CAGR (Options)", _fmt_pct(opt_cagr),  r, "A", True, _GREEN_TX if opt_cagr  >= 0 else _RED_TX)
    _kv("CAGR (Spot)",    _fmt_pct(spot_cagr), r, "D", True, _GREEN_TX if spot_cagr >= 0 else _RED_TX); r += 1

    r += 1  # blank

    # ROI vs Spot table
    def _hdr_cell(col, txt, rn):
        ci = ord(col.upper()) - ord("A") + 1
        sink.cell(rn, ci, txt, bold=True, size=10, fc=_WHITE_TXT, bg=_HEADER_BG,
                  align="C", border=True)

    _hdr_cell("A", "Type", r); _hdr_cell("B", "Sum", r); _hdr_cell("C", "%", r)
    sink.merge(r, 4, 5)
    _hdr_cell("D", "ROI vs Spot", r)
    sink.row_height(r, 20)
    spot_row = r; r += 1

    sink.merge(spot_row + 1, 4, 5)
    # With Midcap, ROI vs Spot = Combined % / Spot % shown as the raw ratio
    # (e.g. 1.5007); otherwise the existing percent display. sum_pct is already
    # Combined when has_midcap (the accumulation loop uses Combined values).
    _roi_clr = _GREEN_TX if roi_pct >= 0 else _RED_TX
    if has_midcap:
        sink.cell(spot_row + 1, 4, roi_pct, bold=True, size=11, fc=_roi_clr,
                  bg=_WHITE, align="C", border=True, nfmt="General")
    else:
        sink.cell(spot_row + 1, 4, _fmt_pct(roi_pct), bold=True, size=11, fc=_roi_clr,
                  bg=_WHITE, align="C", border=True)

    def _type_row(label, value, pct_val):
        sink.cell(r, 1, label, bold=True, size=10, fc=_DARK2_TXT, bg=_LABEL_BG,
                  align="L", border=True)
        sink.cell(r, 2, f"{float(value):,.2f}", bold=True, size=10,
                  fc=(_GREEN_TX if value >= 0 else _RED_TX), bg=_WHITE, align="L", border=True)
        if pct_val is not None:
            sign2 = "+" if pct_val >= 0 else ""
            sink.cell(r, 3, f"{sign2}{float(pct_val):.2f}%", bold=True, size=10,
                      fc=(_GREEN_TX if pct_val >= 0 else _RED_TX), bg=_WHITE, align="L", border=True)
        sink.row_height(r, 18)

    # Use backend summary for Spot P&L sum and Spot P&L %.  The engine writes
    # Spot P&L on exactly one row per trade (the lowest present leg), so the
    # local sums match backend; but reading from `summary.*` keeps all three
    # Excel builders consistent (single source of truth, set in
    # base.compute_analytics).
    _spot_sum_summary = _to_num(S.get("spot_change"))
    if _spot_sum_summary is None: _spot_sum_summary = spot_sum_gated
    _spot_pct_summary = _to_num(S.get("spot_change_pct"))
    if _spot_pct_summary is None: _spot_pct_summary = spot_pct * 100
    _type_row("Spot P&L", _spot_sum_summary, _spot_pct_summary); r += 1
    if has_calls:   _type_row("CE P&L",        ce_sum,             ce_pct * 100);           r += 1
    if has_puts:    _type_row("PE P&L",         pe_sum,             pe_pct * 100);           r += 1
    if has_futures: _type_row("FUT P&L",        fut_sum,            fut_pct * 100);          r += 1
    if has_calls and has_puts:
        _type_row("CE + PE P&L", ce_sum + pe_sum, (ce_pct + pe_pct) * 100); r += 1
    # Midcap leg P&L + Combined rows (matches the backtest Summary Type block).
    if has_midcap:
        mcs = midcap_summary or {}
        sym = mcs.get("symbol") or "NIFTYMIDCAP100"
        mode_lbl = mcs.get("mode_label") or "Hypothetical Future"
        nifty_prefix = " + ".join(
            x for x, f in (("CE", has_calls), ("PE", has_puts), ("FUT", has_futures)) if f
        ) or "NIFTY"
        _type_row(f"{sym} {mode_lbl} P&L",
                  _to_num(mcs.get("midcap_leg_pnl_sum")) or 0.0,
                  _to_num(mcs.get("midcap_leg_pnl_pct_sum"))); r += 1
        _type_row(f"{nifty_prefix} + {sym} {mode_lbl} P&L",
                  _to_num(mcs.get("combined_pnl_sum")) or 0.0,
                  _to_num(mcs.get("combined_pnl_pct_sum"))); r += 1
    _type_row("Net P&L", sum_net, sum_pct); r += 1

    r += 1

    # ── SECTION 2: Risk Metrics ───────────────────────────────────────────────
    _section("RISK METRICS", r); r += 1
    _kv("Max Drawdown",  f"{max_dd_pct:.2f}%", r, "A", False, _RED_TX)
    _kv("Max DD Days",   str(mdd_dur or "—"),  r, "D", False, _RED_TX); r += 1

    dd_period = f"{mdd_start}  →  {mdd_end}" if (mdd_start and mdd_end) else "—"
    sink.merge(r, 1, 5)
    sink.cell(r, 1, f"Drawdown Period:  {dd_period}", bold=True, size=10, fc=_RED_TX,
              bg=_RED_BG, align="C", border=True)
    sink.row_height(r, 18); r += 1

    _kv("Return / MaxDD", f"{car_mdd:.2f}%", r, "A", True, _GREEN_TX if car_mdd >= 0 else _RED_TX); r += 1

    r += 1

    # ── SECTION 3: Consistency & Streaks ─────────────────────────────────────
    _section("CONSISTENCY & STREAKS", r); r += 1
    _kv("Max Win Streak",    f"{int(max_win_str)} trades",  r, "A", False, _GREEN_TX)
    _kv("Max Losing Streak", f"{int(max_loss_str)} trades", r, "D", False, _RED_TX); r += 1
    r += 1

    # ── SECTION 4: Monthly Returns ────────────────────────────────────────────
    _section("MONTHLY RETURNS (₹ Net P&L)", r); r += 1

    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    mth_hdr = ["Year", *MONTHS, "Total", "Max DD", "R/MDD"]

    # Set wider column widths for month table
    for ci in range(1, len(mth_hdr) + 1):
        if ci == 1: sink.col_width(ci, 8)
        elif ci <= 13: sink.col_width(ci, 9)
        elif ci == 14: sink.col_width(ci, 10)
        elif ci == 15: sink.col_width(ci, 18)
        else: sink.col_width(ci, 10)

    by_ym:     Dict[str, List[float]] = {}
    by_ym_pct: Dict[str, List[float]] = {}
    by_yr_max_dd: Dict[str, float] = {}
    # Per-year RUPEE drawdown, for the ₹ table's R/MDD. by_yr_max_dd holds the %DD
    # FRACTION, so the ₹ table was computing rupees ÷ fraction — 521.28 / 0.001436 =
    # 363,008 — a quotient with no dimension. `cleaned` is already in canonical
    # chronological order, so a running cumsum over it gives the rupee curve directly.
    by_yr_max_dd_rs: Dict[str, float] = {}
    _cum_rs = 0.0
    _peak_rs = 0.0

    def _ym(v):
        d = _parse_date(v)
        if not d: return None
        return str(d.year), d.month - 1

    for t in cleaned:
        # Monthly Net P&L on COMBINED when a Midcap leg is present, else NIFTY.
        if has_midcap:
            net_v = _to_num(t.get("Combined Net P&L"))
            if net_v is None: continue
            pct_v = _to_num(t.get("Combined Net P&L %")) or 0.0
        else:
            net_v = _to_num(t.get("Net P&L"))
            if net_v is None: continue
            # READ the tradesheet's own "% P&L" — do NOT re-derive it. On the sync
            # multi-index path a trade's parent row carries the COMBINED Net P&L (both
            # legs) while its Entry Spot is only the OPTIONS index's, so net_v/spot_v
            # divided a combined P&L by one index's spot: measured 28.474% on the NIFTY
            # rows where the tradesheet says 55.182%. "% P&L" is already the correct
            # per-leg sum. This is the exact field wow_mom.py:583 reads, which is why
            # the MOM sheet always matched and this table did not. Single-index is
            # unaffected — there the two are identical by construction.
            pct_v = _to_num(t.get("% P&L"))
            if pct_v is None:
                spot_v = _to_num(t.get("Entry Spot")) or 0.0
                pct_v  = (net_v / spot_v * 100) if spot_v > 0 else 0.0
        ym = _ym(t.get("Exit Date"))
        if not ym: continue
        yr, mi = ym
        by_ym.setdefault(yr, [0.0]*12)[mi]     += net_v
        by_ym_pct.setdefault(yr, [0.0]*12)[mi] += pct_v
        # Rupee equity curve -> deepest rupee dip whose trade EXITS in this year,
        # mirroring how by_yr_max_dd tracks the worst %DD per year.
        _cum_rs += net_v
        if _cum_rs > _peak_rs:
            _peak_rs = _cum_rs
        _dd_rs = _cum_rs - _peak_rs
        if yr not in by_yr_max_dd_rs or _dd_rs < by_yr_max_dd_rs[yr]:
            by_yr_max_dd_rs[yr] = _dd_rs
        dd_v = _to_num(t.get("Combined %DD") if has_midcap else t.get("%DD"))
        if dd_v is not None:
            if yr not in by_yr_max_dd or dd_v < by_yr_max_dd[yr]:
                by_yr_max_dd[yr] = dd_v

    def _render_mth_rows(data_map, is_pct):
        nonlocal r
        for yi, (yr, mos) in enumerate(sorted(data_map.items())):
            total = sum(mos)
            max_dd_yr = by_yr_max_dd.get(yr)
            # R/MDD — numerator and denominator must share units, and the quotient is
            # clamped to +/-99999 (the ceiling compute_analytics uses for car_mdd) so a
            # year whose deepest dip is float-noise-small can't blow up to six figures.
            # A year with NO drawdown has an undefined ratio -> "—", not a blank cell.
            def _rmdd(num, den):
                if not den or den == 0:
                    return "—"
                if not num:
                    return "—"
                return round(max(-99999.0, min(99999.0, num / abs(den))), 2)

            if is_pct:
                # % table: percent / percent. `total` is already a percent sum
                # (data_map=by_ym_pct); by_yr_max_dd is a fraction (dd/peak), so scale
                # it to percentage-points to match instead of dividing by the fraction.
                max_dd_pct = (max_dd_yr * 100) if max_dd_yr is not None else None
                r_mdd = _rmdd(total, max_dd_pct)
            else:
                # ₹ table: rupees / RUPEES. It used to divide `total` (rupees) by
                # by_yr_max_dd (a %DD fraction) — 521.28 / 0.001436 = 363,008.36, with
                # that same fraction rendered as "-0.14%" in the column beside it.
                r_mdd = _rmdd(total, by_yr_max_dd_rs.get(yr))
            row_data = [yr, *[round(v, 2) for v in mos], round(total, 2),
                        max_dd_yr if max_dd_yr is not None else "", r_mdd]
            sink.row_height(r, 18)
            for ci2, val in enumerate(row_data, 1):
                cell_val = val
                nfmt = None
                is_val  = 2 <= ci2 <= 13
                is_tot  = ci2 == 14
                if is_pct and (is_val or is_tot) and isinstance(val, (int, float)):
                    cell_val = val / 100
                    nfmt = "0.00%"
                    num_v = val
                elif ci2 == 15 and isinstance(val, (int, float)):
                    # Max DD here is `by_yr_max_dd`, sourced from the cleaned
                    # row's "%DD"/"Combined %DD" = _aggregate_trades' pctDd,
                    # which is already a fraction (dd/peak, no *100) — same
                    # convention the per-trade detail sheet uses as-is via
                    # _TRUE_PCT_COLS. Don't divide again here.
                    cell_val = val
                    nfmt = "0.00%"
                    num_v = val
                elif is_pct and ci2 == 14 and isinstance(val, (int, float)):
                    cell_val = val
                    nfmt = "0.00%"
                    num_v = val
                else:
                    num_v = val if isinstance(val, (int, float)) else None
                if (is_val or is_tot) and num_v is not None and num_v != 0:
                    bold = True; fc = _GREEN_TX if num_v >= 0 else _RED_TX
                    bg = _GREEN_BG if num_v >= 0 else _RED_BG
                elif ci2 == 1:
                    bold = True; fc = _SUB_HDR_TX; bg = _SUB_HDR_BG
                else:
                    bold = False; fc = "000000"; bg = _WHITE if yi % 2 == 0 else _ALT_ROW
                sink.cell(r, ci2, cell_val, bold=bold, size=10, fc=fc, bg=bg,
                          align="C", border=True, nfmt=nfmt)
            r += 1

    def _month_header():
        nonlocal r
        sink.row_height(r, 20)
        for ci2, h in enumerate(mth_hdr, 1):
            sink.cell(r, ci2, h, bold=True, size=10, fc=_WHITE_TXT, bg=_HEADER_BG,
                      align="C", border=True)
        r += 1

    # Month header
    _month_header()
    _render_mth_rows(by_ym, False)

    r += 1
    _section("MONTHLY RETURNS (% Net P&L)", r); r += 1
    _month_header()
    _render_mth_rows(by_ym_pct, True)

    r += 1

    # ── SECTION 5: Live DD & Outlier Analysis ────────────────────────────────
    _section("LIVE DD & OUTLIER ANALYSIS", r); r += 1
    _kv("Actual Live DD (min)", f"{live_dd_min:.2f}%", r, "A", False, _RED_TX)
    _kv("Avg Actual Live DD",   f"{live_dd_avg:.2f}%", r, "D", False, _RED_TX); r += 1
    _kv("Avg Combined Final MAE" if has_midcap else "Avg Final MAE",
        f"{avg_final_mae:.2f}%", r, "A", False, _RED_TX); r += 1
    _kv("CAR/MDD (Booked)",     f"{car_mdd:.2f}%",       r, "A", True,  _GREEN_TX if car_mdd     >= 0 else _RED_TX)
    _kv("CAR/MDD Live",         f"{car_mdd_live:.2f}%",  r, "D", True,  _GREEN_TX if car_mdd_live >= 0 else _RED_TX); r += 1

    r += 1
    _kv("+ve Outlier 1", _fmt_pct(_p1), r, "A", False, _GREEN_TX)
    _kv("-ve Outlier 1", _fmt_pct(_n1), r, "D", False, _RED_TX); r += 1
    _kv("Actual Live DD Without Outlier 1",     f"{ldd_no_o1[0]:.2f}%", r, "A", True, _RED_TX)
    _kv("Avg Actual Live DD Without Outlier 1", f"{ldd_no_o1[1]:.2f}%", r, "D", True, _RED_TX); r += 1
    _kv("+ve Outlier 2", _fmt_pct(_p2), r, "A", False, _GREEN_TX)
    _kv("-ve Outlier 2", _fmt_pct(_n2), r, "D", False, _RED_TX); r += 1
    _kv("Actual Live DD Without Outlier 2",     f"{ldd_no_o2[0]:.2f}%", r, "A", True, _RED_TX)
    _kv("Avg Actual Live DD Without Outlier 2", f"{ldd_no_o2[1]:.2f}%", r, "D", True, _RED_TX); r += 1
    _kv("+ve Outlier 3", _fmt_pct(_p3), r, "A", False, _GREEN_TX)
    _kv("-ve Outlier 3", _fmt_pct(_n3), r, "D", False, _RED_TX); r += 1
    _kv("Actual Live DD Without Outlier 3",     f"{ldd_no_o3[0]:.2f}%", r, "A", True, _RED_TX)
    _kv("Avg Actual Live DD Without Outlier 3", f"{ldd_no_o3[1]:.2f}%", r, "D", True, _RED_TX); r += 1

    r += 1
    # Outlier-stripped P&L % label reflects the leg configuration (matches the
    # backtest): "CE + NIFTYMIDCAP100 Hypothetical Future P&L %" with Midcap,
    # else the existing "CE + PE + P&L %".
    _outlier_base = "CE + PE + P&L %"
    if has_midcap:
        _mcs = midcap_summary or {}
        _sym = _mcs.get("symbol") or "NIFTYMIDCAP100"
        _mode = _mcs.get("mode_label") or "Hypothetical Future"
        _np = " + ".join(
            x for x, f in (("CE", has_calls), ("PE", has_puts), ("FUT", has_futures)) if f
        ) or "NIFTY"
        _outlier_base = f"{_np} + {_sym} {_mode} P&L %"
    for si, (label, val) in enumerate([
        (f"{_outlier_base} Without Top 1 Outliers", pct_no_o1),
        (f"{_outlier_base} Without Top 2 Outliers", pct_no_o2),
        (f"{_outlier_base} Without Top 3 Outliers", pct_no_o3),
    ]):
        sink.merge(r, 1, 4)
        sink.cell(r, 1, label, bold=True, size=10, fc=_DARK2_TXT,
                  bg=(_LABEL_BG if si % 2 == 0 else _ALT_ROW), align="L", border=True)
        sink.cell(r, 5, _fmt_pct(val), bold=True, size=10,
                  fc=(_GREEN_TX if val >= 0 else _RED_TX),
                  bg=(_WHITE if si % 2 == 0 else _ALT_ROW), align="C", border=True)
        sink.row_height(r, 18); r += 1


# ── Public API ────────────────────────────────────────────────────────────────

# Parity-reference switch: when True, compute_xlsx_summary_metrics runs the Python
# engine below instead of the Rust path — used ONLY by tools/summary_metrics_parity
# to diff Rust vs Python. The live path keeps it False (Rust-only, no fallback).
_SUMMARY_PYTHON_REF = False


def compute_xlsx_summary_metrics(
    trades_df: pd.DataFrame,
    summary: Dict[str, Any],
    midcap_legs=None,
    midcap_spot_adjustment=None,
    midcap_symbol: str = "NIFTYMIDCAP100",
    patchwise: bool = False,
    filter_segments=None,
) -> Dict[str, Any]:
    """
    Return a dict of every summary stat that _write_summary_sheet computes from
    the trades, using identical formulas.  Called by the optimizer after
    MAE/MFE enrichment so the master summary CSV matches each combo XLSX exactly.
    When midcap_legs is provided, the headline stats run on the COMBINED
    (NIFTY + Midcap) per-trade P&L and Midcap-specific fields are added.
    When patchwise is set (patchwise ZIP download), the equity chain resets to
    100 at each patch boundary — identical to the per-combo XLSX patchwise
    Summary — so the master summary matches the individual combo summary.
    """
    S = summary or {}
    if trades_df is None or (hasattr(trades_df, "empty") and trades_df.empty):
        rows: List[Dict] = []
    else:
        rows = trades_df.where(trades_df.notna(), None).to_dict("records")

    midcap_by_trade, midcap_summary, has_midcap = compute_midcap_for_rows(
        rows, midcap_legs, midcap_spot_adjustment, midcap_symbol,
    )

    # ── Rust-ONLY (no Python fallback — user policy). This entire summary engine
    # is ported to Rust and byte-identical (proven by tools/summary_metrics_parity
    # + an isolated 39-key/0-diff check, overall AND patchwise). It operates on the
    # finished trades, so it is feature-agnostic — safe for ALL strategies incl.
    # midcap / spot-adj / filter. Any Rust failure HARD-FAILS (surfaces the bug)
    # rather than silently falling back. The Python engine below is a PARITY
    # REFERENCE ONLY, reached exclusively when _SUMMARY_PYTHON_REF is set (by
    # tools/summary_metrics_parity).
    if not _SUMMARY_PYTHON_REF:
        import algotest_native as _an_sm
        _rust_sm = _an_sm.compute_summary_metrics(
            rows, S, bool(patchwise), filter_segments,
            (midcap_by_trade or None), midcap_summary,
        )
        if not isinstance(_rust_sm, dict):
            raise RuntimeError(
                "Rust compute_summary_metrics returned a non-dict "
                f"({type(_rust_sm).__name__}) — no Python fallback per policy"
            )
        # The Rust summary engine emits CE/PE totals but not FUT. Add fut_pnl_total /
        # fut_pnl_pct here (same formula as _write_summary_sheet: Σ FUT P&L; pct =
        # Σ FUT P&L / Entry Spot * 100) so the MASTER summary's per-type breakdown
        # includes a real futures leg identically to the per-combo Summary sheet.
        # A pure additive supplement — no existing Rust value is touched.
        if "fut_pnl_total" not in _rust_sm:
            _ft = 0.0
            _fp = 0.0
            for _r in rows:
                _fu = _to_num(_r.get("FUT P&L"))
                if _fu is None:
                    continue
                _ft += _fu
                _es = _to_num(_r.get("Entry Spot"))
                if _es not in (None, 0):
                    _fp += _fu / _es
            _rust_sm["fut_pnl_total"] = round(_ft, 2)
            _rust_sm["fut_pnl_pct"] = round(_fp * 100, 4)
        return _rust_sm

    tm, _grouped, _sorted_keys = _aggregate_trades(
        rows, has_midcap, midcap_by_trade,
        patchwise=patchwise, filter_segments=filter_segments,
    )

    # Patchwise reset boundaries (segment START dates), shared by the Max DD scan
    # and the outlier Live DD scan so they reset at the same points as the per-combo
    # XLSX patchwise Summary (mirrors _write_summary_sheet).
    def _pw_seg_start_ms(s):
        s = s or {}
        return _date_ms(s.get("start") or s.get("Start") or s.get("from")
                        or s.get("start_date") or s.get("startdt"))
    _pw_seg_starts = sorted(ms for ms in (_pw_seg_start_ms(s) for s in (filter_segments or []))
                            if ms is not None)

    def _pw_row_seg_idx(entry_val) -> int:
        em = _date_ms(entry_val)
        i = -1
        if em is not None:
            for j, sm in enumerate(_pw_seg_starts):
                if sm <= em:
                    i = j
                else:
                    break
        return i

    # Single source of truth for trade order: `_sorted_keys` is the CANONICAL
    # chronological (entry-date) order _aggregate_trades used to compute every
    # per-trade equity value. Both the max-DD scan and the outlier-stripped Live
    # DD below walk this exact order, and _write_summary_sheet (combo XLSX) is now
    # handed the same list — so master and XLSX are identical AND internally
    # consistent with the base Live DD. cleaned is still built for the (legacy)
    # column extraction but is no longer used to derive trade order.
    _key_order, _hc, _hp, _hf = _build_key_order(rows, has_midcap)
    cleaned = _build_cleaned_rows(rows, _key_order, tm, has_midcap, midcap_by_trade)

    # First-seen Entry Date / Exit Reason per trade key (for patchwise segment
    # bucketing in the Max DD + outlier Live DD scans below).
    _key_entry_pw: Dict[str, Any] = {}
    _key_exit_reason_pw: Dict[str, str] = {}
    for _cr in cleaned:
        _ck = str(_cr.get("Trade") or _cr.get("trade") or 1)
        if _ck not in _key_entry_pw and _cr.get("Entry Date"):
            _key_entry_pw[_ck] = _cr.get("Entry Date")
        if _cr.get("Exit Reason"):
            # Accumulate the UNION of all legs' exit reasons (not last-leg-wins) so
            # the Max DD / outlier Live DD patchwise scans detect a FILTER_END on
            # ANY leg — matching the per-trade cumulative/peak/lowest-nav reset and
            # keeping every scan resetting at the SAME boundaries for mixed
            # options+futures trades (leg-order-independent).
            _prev_er = _key_exit_reason_pw.get(_ck, "")
            _cur_er = (_cr.get("Exit Reason") or "").upper()
            _key_exit_reason_pw[_ck] = (_prev_er + "+" + _cur_er) if _prev_er else _cur_er

    sum_pct = 0.0; sum_pos_pct = 0.0; sum_neg_pct = 0.0
    win_cnt = 0;   loss_cnt = 0;      total_cnt = 0
    sum_net = 0.0
    final_cum = 100.0; spot_cum = 100.0
    min_entry_ms = None; max_exit_ms = None
    spot_sum_gated = 0.0
    ce_sum = 0.0; pe_sum = 0.0; fut_sum = 0.0
    ce_pct_sum = 0.0; pe_pct_sum = 0.0; fut_pct_sum = 0.0; spot_pct_sum = 0.0

    def _parse_date_ms(v):
        d = _parse_date(v)
        return d.timestamp() * 1000 if d else None

    for t in rows:
        ce = _to_num(t.get("CE P&L")) or _to_num(t.get("Call P&L"))
        pe = _to_num(t.get("PE P&L")) or _to_num(t.get("Put P&L"))
        fu = _to_num(t.get("FUT P&L"))
        ce_sum  += ce if ce is not None else 0
        pe_sum  += pe if pe is not None else 0
        fut_sum += fu if fu is not None else 0
        es = _to_num(t.get("Entry Spot"))
        cep = _to_num(t.get("CE P&L %"))
        if cep is None and ce is not None and es and es != 0:
            cep = ce / es
        ce_pct_sum += cep if cep is not None else 0
        pep = _to_num(t.get("PE P&L %"))
        if pep is None and pe is not None and es and es != 0:
            pep = pe / es
        pe_pct_sum += pep if pep is not None else 0
        # FUT P&L % is not a stored column — derive it FUT P&L / Entry Spot (same as
        # _write_summary_sheet) so the master's futures % matches the per-combo Summary.
        if fu is not None and es and es != 0:
            fut_pct_sum += fu / es
        spp = _to_num(t.get("Spot P&L %"))
        sp = _to_num(t.get("Spot P&L"))
        if spp is None and sp is not None and es and es != 0:
            spp = sp / es
        spot_pct_sum += spp if spp is not None else 0

    _main_of = _main_leg

    # PER-TRADE accumulation. This previously iterated the raw per-LEG `rows`, which
    # double-counted multi-leg trades and produced the cagr_spot (e.g. straddle
    # 57.22 vs 25.37), avg_win/avg_loss/avg_profit_pct and roi_vs_spot divergence
    # from the backtest + the per-combo tradesheet. `tm` holds the per-TRADE pct/net
    # (identical to base.py's net_pnl_pct), so the master now matches them exactly.
    for _k in _sorted_keys:
        _tmr = tm.get(_k) or {}
        p = _tmr.get("pct"); n = _tmr.get("net")
        if isinstance(p, (int, float)) and math.isfinite(p):
            sum_pct += p; total_cnt += 1
            if p > 0:   sum_pos_pct += p; win_cnt  += 1
            elif p < 0: sum_neg_pct += p; loss_cnt += 1
        if isinstance(n, (int, float)) and math.isfinite(n):
            sum_net += n
        _mn = _main_of(_grouped.get(_k) or [])
        _sp = _to_num(_mn.get("Spot P&L"))
        if _sp is not None: spot_sum_gated += _sp
        ed = _parse_date_ms(_mn.get("Entry Date"))
        xd = _parse_date_ms(_mn.get("Exit Date"))
        if ed and (min_entry_ms is None or ed < min_entry_ms): min_entry_ms = ed
        if xd and (max_exit_ms  is None or xd > max_exit_ms):  max_exit_ms  = xd
    # cagr_spot inputs: spot LEVELS of the first/last trade by canonical entry-date
    # order (leg-independent), matching base.py:1069-1075. The old per-leg spot_cum
    # compounding is dropped.
    _init_spot  = _to_num(_main_of(_grouped.get(_sorted_keys[0])  or []).get("Entry Spot")) if _sorted_keys else None
    _final_spot = _to_num(_main_of(_grouped.get(_sorted_keys[-1]) or []).get("Exit Spot"))  if _sorted_keys else None

    # Final booked equity from the CANONICAL patch-aware chain (tm/_sorted_keys),
    # NOT the raw overall "Cumulative" column of rows — so CAGR(Options) and
    # CAR/MDD honor the per-patch reset in patchwise mode. In overall mode this
    # chain equals the raw Cumulative, so the value is unchanged. Uses _sorted_keys
    # (canonical chronological order) so it is the true last trade, order-safe.
    # The midcap block below overrides final_cum with the combined-chain value.
    if _sorted_keys:
        _last_cum = tm.get(_sorted_keys[-1], {}).get("cumulative")
        if isinstance(_last_cum, float) and math.isfinite(_last_cum):
            final_cum = _last_cum

    # With a Midcap leg, recompute the headline aggregates on the COMBINED
    # per-trade P&L (chronological), mirroring the combo Summary sheet. CAGR(Spot)
    # stays NIFTY (spot_cum, above). combined_max_dd drives CAR/MDD.
    # Use the CANONICAL chronological order returned by _aggregate_trades — the
    # exact same order the per-trade equity values were computed in, and the same
    # list now handed to _write_summary_sheet. Single source of truth.
    _chron_from_rows: List[str] = list(_sorted_keys)

    combined_max_dd = None
    if has_midcap:
        _seen_c: set = set(); _chron_c: List[str] = []
        for _k in _chron_from_rows:
            if _k not in _seen_c:
                _seen_c.add(_k); _chron_c.append(_k)
        sum_pct = sum_pos_pct = sum_neg_pct = 0.0
        win_cnt = loss_cnt = total_cnt = 0
        sum_net = 0.0; final_cum = 100.0
        worst_dd = 0.0
        win_net_sum = loss_net_sum = 0.0
        max_net_c = -math.inf; min_net_c = math.inf
        for _k in _chron_c:
            t2 = tm[_k]; mc2 = midcap_by_trade.get(_k) or {}
            cp = t2.get("combinedPct"); cp = cp if isinstance(cp, float) and math.isfinite(cp) else None
            cnet = _to_num(mc2.get("Combined Net P&L"))
            cc = t2.get("combinedCum"); cc = cc if isinstance(cc, float) and math.isfinite(cc) else None
            cdd = t2.get("combinedPctDd"); cdd = cdd if isinstance(cdd, float) and math.isfinite(cdd) else None
            if cp is not None:
                sum_pct += cp; total_cnt += 1
                if cp > 0:   sum_pos_pct += cp; win_cnt  += 1
                elif cp < 0: sum_neg_pct += cp; loss_cnt += 1
            if cnet is not None:
                sum_net += cnet
                if cnet > max_net_c: max_net_c = cnet
                if cnet < min_net_c: min_net_c = cnet
                if cnet > 0:   win_net_sum  += cnet
                elif cnet < 0: loss_net_sum += cnet
            if cc is not None:
                final_cum = cc
            # Max DD = worst Combined %DD (identical to the tradesheet column), both modes.
            if cdd is not None and cdd < worst_dd:
                worst_dd = cdd
        combined_max_dd = worst_dd
        if not math.isfinite(max_net_c): max_net_c = 0.0
        if not math.isfinite(min_net_c): min_net_c = 0.0

    avg_win_pct  = (sum_pos_pct / win_cnt)  if win_cnt  > 0 else 0.0
    avg_loss_pct = (sum_neg_pct / loss_cnt) if loss_cnt > 0 else 0.0
    avg_pct      = (sum_pct / total_cnt)    if total_cnt > 0 else 0.0

    # Year span + CAGRs use the backtest's exact convention (base.py:999,1075):
    # integer calendar days between first entry and last exit, / 365.0, floored at
    # 0.01; CAGR(options) clamped to +/-99999, and -100 on a wiped-out equity.
    _span_days = (round((max_exit_ms - min_entry_ms) / (86400 * 1000))
                  if (min_entry_ms is not None and max_exit_ms is not None) else 0)
    years = max(_span_days / 365.0, 0.01)
    opt_cagr = (
        max(-99999.0, min(99999.0, (math.pow(final_cum / 100, 1 / years) - 1) * 100))
        if (years > 0 and final_cum > 0) else -100.0
    )
    # cagr_spot from spot LEVELS (leg-independent) — base.py:1075.
    # Mirrors summary_metrics.rs: the BACKTEST owns cagr_spot. This Python engine is a
    # PARITY REFERENCE for the Rust one, so it must pin the same value — otherwise
    # tools/summary_metrics_parity would report a false PASS on multi-index payloads.
    # `is None`, NOT `or` — must match summary_metrics.rs's unwrap_or_else, which fires
    # only when the key is ABSENT. With `or`, a genuine cagr_spot of 0.0 fell through to
    # the derivation, so Rust reported 0.00 while Python reported the (cross-index)
    # derived value — a fresh divergence on exactly the sync_weekly_roll path.
    _S_cagr_spot = _to_num(S.get("cagr_spot"))
    spot_cagr = _S_cagr_spot if _S_cagr_spot is not None else (
        100 * ((_final_spot / _init_spot) ** (1.0 / years) - 1)
        if (years > 0 and _init_spot and _final_spot and _init_spot > 0 and _final_spot > 0)
        else 0.0
    )

    # Max Drawdown = min %DD on the equity curve (Combined for midcap, NIFTY otherwise),
    # read straight from Cumulative/Peak so the master == the per-combo tradesheet's %DD,
    # for midcap AND non-midcap. Mirrors _write_summary_sheet exactly.
    _mdd_cum_col  = "Combined Cumulative" if has_midcap else "Cumulative"
    _mdd_peak_col = "Combined Peak"       if has_midcap else "Peak"
    max_dd_pct = 0.0
    # OVERALL + non-midcap: read the backtest's OWN %DD column (base.py:939 stamps
    # it, base.py:1008 does float(%DD.min())) so the master is byte-identical to the
    # backtest instead of re-deriving cum/peak and drifting at the 7th decimal.
    # Patchwise (per-patch reset) and midcap (Combined) keep the recompute below.
    _src_dd_used = False
    if not patchwise and not has_midcap:
        _src_dd = [_to_num(r.get("%DD")) for r in rows]
        _src_dd = [v for v in _src_dd if v is not None and math.isfinite(v)]
        if _src_dd:
            max_dd_pct = float(min(_src_dd))
            _src_dd_used = True
    if not _src_dd_used:
        for t in cleaned:
            cc = _to_num(t.get(_mdd_cum_col)); pk = _to_num(t.get(_mdd_peak_col))
            if cc is not None and pk not in (None, 0) and math.isfinite(cc):
                ddp = (cc / pk - 1) * 100
                if ddp < max_dd_pct:
                    max_dd_pct = ddp
    if has_midcap:
        combined_max_dd = max_dd_pct
    car_mdd = opt_cagr / abs(max_dd_pct) if max_dd_pct != 0 else 0.0

    opt_sum = (
        (ce_sum + pe_sum) if (ce_sum != 0 or pe_sum != 0)
        else (fut_sum if fut_sum != 0 else sum_net)
    )
    _scp = _to_num(S.get("spot_change_pct")) or (spot_pct_sum * 100) or 0.0
    roi_pct = sum_pct / abs(_scp) if _scp != 0 else 0.0

    trade_pairs: List[Dict] = []
    _seen2: set = set()
    _chron_keys: List[str] = []
    for _k in _chron_from_rows:
        if _k not in _seen2:
            _seen2.add(_k); _chron_keys.append(_k)
    for _k in tm.keys():
        if _k not in _seen2:
            _seen2.add(_k); _chron_keys.append(_k)
    _pk = "combinedPct"       if has_midcap else "pct"
    _lk = "combinedActualLDD" if has_midcap else "actualLDD"
    _mk = "combinedFinalMae"  if has_midcap else "finalMae"
    for k in _chron_keys:
        t2 = tm[k]
        pct_v = t2.get(_pk)
        pct_v = pct_v if isinstance(pct_v, float) and math.isfinite(pct_v) else None
        ldd_v = t2.get(_lk)
        ldd_v = ldd_v if isinstance(ldd_v, float) and math.isfinite(ldd_v) else None
        mae_v = t2.get(_mk)
        mae_v = mae_v if isinstance(mae_v, float) and math.isfinite(mae_v) else None
        if pct_v is not None:
            trade_pairs.append({"pct": pct_v, "ldd": ldd_v, "mae": mae_v, "idx": len(trade_pairs),
                                 "exitReason": _key_exit_reason_pw.get(k, ""),
                                 "segIdx": _pw_row_seg_idx(_key_entry_pw.get(k))})

    n_trades = len(trade_pairs)
    by_pct_desc = sorted(trade_pairs, key=lambda x: -x["pct"])

    _p1 = by_pct_desc[0]["pct"]  if n_trades > 0 else 0.0
    _p2 = _p1 + by_pct_desc[1]["pct"] if n_trades > 1 else _p1
    _p3 = _p2 + by_pct_desc[2]["pct"] if n_trades > 2 else _p2
    _n1 = by_pct_desc[n_trades - 1]["pct"] if n_trades > 0 else 0.0
    _n2 = _n1 + by_pct_desc[n_trades - 2]["pct"] if n_trades > 1 else _n1
    _n3 = _n2 + by_pct_desc[n_trades - 3]["pct"] if n_trades > 2 else _n2
    total_pct_s = sum(p["pct"] for p in trade_pairs)
    pct_no_o1 = total_pct_s - _p1 - _n1
    pct_no_o2 = total_pct_s - _p2 - _n2
    pct_no_o3 = total_pct_s - _p3 - _n3

    def _ldd_exc(exc_top: int, exc_bot: int):
        exc_idx = {
            *[p["idx"] for p in by_pct_desc[:exc_top]],
            *[p["idx"] for p in by_pct_desc[max(0, n_trades - exc_bot):]],
        }
        filtered = [p for p in trade_pairs if p["idx"] not in exc_idx]
        if not filtered:
            return (0.0, 0.0)
        cumulative = 100.0
        peak = 100.0
        prev_cum = 100.0
        prev_peak = 100.0
        first_done = False
        prev_exit_reason = ""
        prev_seg_idx = None
        ldds = []
        for p in filtered:
            # Reset the chain at each patch boundary (same boundary as the per-combo
            # XLSX patchwise Summary) so the cumulative reset isn't read as a DD.
            if patchwise and (
                (prev_seg_idx is not None and p.get("segIdx") != prev_seg_idx) if _pw_seg_starts
                else ("FILTER_END" in (prev_exit_reason or "").split("+"))
            ):
                cumulative = 100.0; peak = 100.0; prev_cum = 100.0; prev_peak = 100.0
            prev_seg_idx = p.get("segIdx")
            pct = p["pct"]
            prev_peak = peak
            cumulative *= (1.0 + pct / 100.0)
            peak = max(peak, cumulative)
            mae = p["mae"]
            if mae is not None and prev_peak != 0:
                # Revised rule: every trade (incl. first, prev_cum = 100) anchors
                # the low to prev_cum * (1 + FinalMAE%); Live DD divides by the
                # PREVIOUS trade's peak (AV_prev), not this trade's peak.
                lowest_nav = round(prev_cum * (1.0 + mae / 100.0) * 100) / 100
                actual_ldd = round((lowest_nav / prev_peak - 1) * 10000) / 100
                ldds.append(actual_ldd)
                first_done = True
            else:
                first_done = True
            prev_cum = cumulative
            prev_exit_reason = p.get("exitReason") or ""
        if not ldds:
            return (0.0, 0.0)
        return (round(min(ldds), 4), round(sum(ldds) / len(ldds), 4))

    all_ldds = [p["ldd"] for p in trade_pairs if p["ldd"] is not None]
    live_dd_min  = round(min(all_ldds), 4)                if all_ldds else 0.0
    live_dd_avg  = round(sum(all_ldds) / len(all_ldds), 4) if all_ldds else 0.0
    # Avg (Combined) Final MAE — mean of each trade's Final MAE (Combined when a Midcap
    # leg is present, NIFTY otherwise; identical to the per-combo XLSX summary).
    _final_maes = [p["mae"] for p in trade_pairs if p["mae"] is not None]
    avg_final_mae = round(sum(_final_maes) / len(_final_maes), 4) if _final_maes else 0.0
    ldd_no_o1    = _ldd_exc(1, 1)
    ldd_no_o2    = _ldd_exc(2, 2)
    ldd_no_o3    = _ldd_exc(3, 3)
    car_mdd_live = opt_cagr / abs(live_dd_min) if live_dd_min != 0 else 0.0

    _spot_chg     = _to_num(S.get("spot_change"))     or round(spot_sum_gated, 2)
    _spot_chg_pct = _to_num(S.get("spot_change_pct")) or round(spot_pct_sum * 100, 4)

    _metrics = {
        "cagr_options":                          round(opt_cagr, 2),
        "cagr_spot":                             round(spot_cagr, 2),
        # Max Drawdown = min %DD on the equity curve (Combined for midcap, NIFTY otherwise),
        # for BOTH midcap and non-midcap — so the master summary == the per-combo tradesheet.
        "max_dd_pct":                            max_dd_pct,
        "car_mdd":                               round(car_mdd, 4),
        "roi_vs_spot":                           round(roi_pct, 4),
        "avg_profit_per_trade_pct":              round(avg_pct, 4),
        "avg_win_pct":                           round(avg_win_pct, 4),
        "avg_loss_pct":                          round(avg_loss_pct, 4),
        "ce_pnl_total":                          round(ce_sum, 2),
        "ce_pnl_pct":                            round(ce_pct_sum * 100, 4),
        "pe_pnl_total":                          round(pe_sum, 2),
        "pe_pnl_pct":                            round(pe_pct_sum * 100, 4),
        "fut_pnl_total":                         round(fut_sum, 2),
        "fut_pnl_pct":                           round(fut_pct_sum * 100, 4),
        "long_spot_pnl":                         _spot_chg,
        "long_spot_pnl_pct":                     _spot_chg_pct,
        "actual_live_dd_max":                    live_dd_min,
        "actual_live_dd_avg":                    live_dd_avg,
        "avg_final_mae":                         avg_final_mae,
        "car_mdd_live":                          round(car_mdd_live, 4),
        "positive_outlier_1":                    round(_p1, 4),
        "negative_outlier_1":                    round(_n1, 4),
        "positive_outlier_2":                    round(_p2, 4),
        "negative_outlier_2":                    round(_n2, 4),
        "positive_outlier_3":                    round(_p3, 4),
        "negative_outlier_3":                    round(_n3, 4),
        "outlier_dd_1":                          ldd_no_o1[0],
        "outlier_dd_1_avg":                      ldd_no_o1[1],
        "outlier_dd_2":                          ldd_no_o2[0],
        "outlier_dd_2_avg":                      ldd_no_o2[1],
        "outlier_dd_3":                          ldd_no_o3[0],
        "outlier_dd_3_avg":                      ldd_no_o3[1],
        "ce_pe_pnl_pct_without_top_1_outliers": round(pct_no_o1, 4),
        "ce_pe_pnl_pct_without_top_2_outliers": round(pct_no_o2, 4),
        "ce_pe_pnl_pct_without_top_3_outliers": round(pct_no_o3, 4),
        # Midcap overlay (present only when a Midcap leg ran; headline stats above
        # are already COMBINED in that case).
        "has_midcap":             bool(has_midcap),
        "midcap_leg_pnl_sum":     round(_to_num((midcap_summary or {}).get("midcap_leg_pnl_sum")) or 0.0, 2) if has_midcap else None,
        "midcap_leg_pnl_pct_sum": round(_to_num((midcap_summary or {}).get("midcap_leg_pnl_pct_sum")) or 0.0, 4) if has_midcap else None,
        "combined_pnl_sum":       round(_to_num((midcap_summary or {}).get("combined_pnl_sum")) or 0.0, 2) if has_midcap else None,
        "combined_pnl_pct_sum":   round(_to_num((midcap_summary or {}).get("combined_pnl_pct_sum")) or 0.0, 4) if has_midcap else None,
        "max_dd_pct_combined":    round(combined_max_dd, 4) if (has_midcap and combined_max_dd is not None) else None,
    }

    # With a Midcap leg, overwrite the headline P&L stats with the COMBINED
    # values so the master summary table matches the per-combo Summary sheet
    # (and the tradesheet). Non-Midcap combos keep the stored NIFTY values.
    if has_midcap:
        _win_rate = (win_cnt / total_cnt * 100) if total_cnt > 0 else 0.0
        _loss_rate = (loss_cnt / total_cnt * 100) if total_cnt > 0 else 0.0
        _expectancy = (
            ((_win_rate / 100) * avg_win_pct - (_loss_rate / 100) * abs(avg_loss_pct)) / abs(avg_loss_pct)
            if avg_loss_pct != 0 else 0.0
        )
        _metrics.update({
            "total_pnl":            round(sum_net, 2),
            "total_pnl_pct":        round(sum_pct, 4),
            "count":                int(total_cnt),
            "win_pct":              round(_win_rate, 2),
            "loss_pct":             round(_loss_rate, 2),
            "avg_profit_per_trade": round(sum_net / total_cnt, 2) if total_cnt > 0 else 0.0,
            "avg_win":              round(win_net_sum / win_cnt, 2) if win_cnt > 0 else 0.0,
            "avg_loss":             round(loss_net_sum / loss_cnt, 2) if loss_cnt > 0 else 0.0,
            "max_win":              round(max_net_c, 2),
            "max_loss":             round(min_net_c, 2),
            "expectancy":           round(_expectancy, 6),
            "max_dd_pct":           max_dd_pct,
        })

    return _metrics


def _project_rows_for_midcap(rows: List[Dict]) -> List[Dict]:
    """Project per-leg trade rows into one row per trade for compute_midcap_legs:
    {trade_id, entry_date, exit_date, nifty_pnl, nifty_pnl_pct}."""
    grouped: Dict[str, List[Dict]] = {}
    for r in rows:
        k = str(r.get("Trade") or r.get("trade") or 1)
        grouped.setdefault(k, []).append(r)
    out: List[Dict] = []
    for k, legs in grouped.items():
        mains = [l for l in legs
                 if not l.get("ReEntryIndex") and not l.get("ReEntryTrigger")
                 and not l.get("ReEntryMode") and not _is_lazy(l)] or list(legs)
        # PARENT row = LOWEST Leg number, which is where simulate.rs:1794-1806
        # writes the trade total. Selecting it by list position instead made the
        # total depend on how the rows happened to be ordered.
        parent = min(mains, key=lambda l: (_to_num(l.get("Leg")) or 0))
        # WINDOW row = the ANCHOR (latest leg entry, ties -> lowest Leg). A
        # CARRIED yearly leg holds an older entry date than the weekly leg that
        # re-enters each cycle; reading the window off "the first row" therefore
        # priced the Midcap overlay from the carried leg's anchor whenever the
        # user configured the yearly leg first, which drove every trade to
        # available=False and stripped the whole Midcap column block from the
        # Trade Sheet. See services/trade_anchor.py.
        main = _anchor_row(mains) or parent
        net = _to_num(parent.get("Net P&L"))
        if net is None:
            net = sum(
                (_to_num(l.get("CE P&L"))  or 0) +
                (_to_num(l.get("PE P&L"))  or 0) +
                (_to_num(l.get("FUT P&L")) or 0)
                for l in legs
            )
        # Recompute the NIFTY leg % at full (4-dp) precision to mirror the
        # backtest engine exactly (engine_rust.py:1952 → round(net/spot*100, 4)).
        # Reading the pre-rounded "% P&L" here (runner.py:1295 sets it via
        # .round(2)) feeds a 2-dp NIFTY % into the Combined Net P&L % formula
        # (Combined % = nifty% + Σ midcap leg %, midcap_overlay.py:357), drifting
        # the optimizer's Combined % vs a direct backtest by ~0.003%/trade
        # (~0.03% over a full run) even though the rupee P&L is identical.
        es = _to_num(main.get("Entry Spot")) or 0.0
        pct = round(net / es * 100.0, 4) if es else 0.0
        def _iso(v):
            d = _parse_date(v)
            return d.strftime("%Y-%m-%d") if d else None
        out.append({
            "trade_id":      k,
            "entry_date":    _iso(main.get("Entry Date") or main.get("entry_date")),
            "exit_date":     _iso(main.get("Exit Date") or main.get("exit_date")),
            "nifty_pnl":     net,
            "nifty_pnl_pct": pct,
        })
    return out


def compute_midcap_for_rows(rows: List[Dict], midcap_legs, midcap_spot_adjustment,
                            symbol: str = "NIFTYMIDCAP100"):
    """Compute the Midcap overlay per combo via the SAME native engine the
    backtest uses (rust_fast_path.compute_midcap_legs) — RUST ONLY, no Python
    fallback (matches the backtest's rust-only policy). If the native path is
    unavailable the overlay is simply not applied (NIFTY-only sheet), never
    silently computed in Python.
    Returns (midcap_by_trade {str(trade_id): fields}, midcap_summary, has_midcap)."""
    if not midcap_legs or not rows:
        return {}, None, False
    proj = _project_rows_for_midcap(rows)
    symbol = (symbol or "NIFTYMIDCAP100")
    # NOTE: the overlay does NOT apply midcap_spot_adjustment — the ENGINE
    # (run_rust_engine_pipeline) already truncates the trade at the Midcap breach
    # and re-enters, so the trade rows here already reflect it. The overlay just
    # prices the Midcap leg over each trade's window. Re-applying spot-adj here
    # would double-handle it AND use the base (not per-combo swept) value,
    # diverging from the engine and between single-combo vs ZIP. Pass None.
    result = None
    try:
        from services import index_ohlc_store, rust_fast_path
        index_ohlc_store.ensure_index_ohlc_loaded(symbol)
        if rust_fast_path.index_ohlc_is_loaded() and rust_fast_path.compute_midcap_legs_available():
            result = rust_fast_path.compute_midcap_legs(proj, midcap_legs, None, symbol)
        else:
            logger.warning("[OPTIM_MIDCAP] native midcap engine unavailable (rust-only) — skipping overlay")
    except Exception as _exc:
        logger.warning("[OPTIM_MIDCAP] rust compute_midcap_legs failed: %s", _exc)
        result = None
    # A CONFIGURED Midcap leg that prices nothing must never vanish quietly: the
    # whole _MIDCAP_COLS block is gated on the has_midcap flag returned here, so
    # a False slips 21 columns out of the Trade Sheet with no other symptom.
    if not result or not result.get("available"):
        logger.warning(
            "[OPTIM_MIDCAP] %d Midcap leg(s) configured but the overlay returned "
            "nothing for any of %d trades — the Midcap/Combined columns will be "
            "ABSENT from the Trade Sheet. First projected window: %s",
            len(midcap_legs), len(proj), (proj[0] if proj else None),
        )
        if os.environ.get("OPTIM_MIDCAP_ALLOW_EMPTY") != "1":
            # The workbook would silently become NIFTY-only: 21 Midcap/Combined
            # columns vanish AND the headline numbers change (measured: Net P&L
            # 353.46 -> 40.00), under a combo label that still says Midcap.
            raise RuntimeError(
                "Midcap overlay produced nothing for any of %d trades though %d "
                "Midcap leg(s) are configured. Refusing to emit a NIFTY-only "
                "workbook under a Midcap label. Set OPTIM_MIDCAP_ALLOW_EMPTY=1 to "
                "override deliberately." % (len(proj), len(midcap_legs))
            )
        return {}, None, False
    by_trade = {
        str(rr.get("trade_id")): rr
        for rr in (result.get("results") or [])
        if rr.get("available")
    }
    if not by_trade:
        _unpriced = [rr for rr in (result.get("results") or []) if not rr.get("available")]
        logger.warning(
            "[OPTIM_MIDCAP] %d Midcap leg(s) configured but NO trade could be "
            "priced (%d unpriced of %d projected) — the Midcap/Combined columns "
            "will be ABSENT from the Trade Sheet. First projected window: %s",
            len(midcap_legs), len(_unpriced), len(proj), (proj[0] if proj else None),
        )
        if os.environ.get("OPTIM_MIDCAP_ALLOW_EMPTY") != "1":
            raise RuntimeError(
                "Midcap overlay priced 0 of %d projected trades though %d Midcap "
                "leg(s) are configured. Refusing to emit a NIFTY-only workbook "
                "under a Midcap label. Set OPTIM_MIDCAP_ALLOW_EMPTY=1 to override."
                % (len(proj), len(midcap_legs))
            )
        return {}, None, False
    _missing = len(proj) - len(by_trade)
    if _missing > 0:
        logger.warning(
            "[OPTIM_MIDCAP] %d of %d trades could not be priced on the Midcap "
            "overlay and will show blank Combined values.", _missing, len(proj),
        )
    summ = dict(result.get("summary") or {})
    is_hypo = any(
        str((l or {}).get("midcap_mode") or (l or {}).get("mode") or "").lower() == "hypothetical"
        for l in midcap_legs
    )
    summ["mode_label"] = "Hypothetical Future" if is_hypo else "Spot"
    return by_trade, summ, True


def _build_combo_xlsx_rust(
    cleaned, key_order, summary, tm, grouped, sorted_keys,
    combo_label, from_date, to_date, has_calls, has_puts, has_futures,
    has_midcap, midcap_summary, midcap_by_trade, patchwise, filter_segments, want_patch,
    yearly: bool = False, rules_sheet=None,
) -> bytes:
    """Rust workbook path for build_combo_xlsx (OPTIMIZE_RUST_XLSX=1). Same sheets,
    same order, cell-identical. Trade Sheet is built in Rust from `cleaned`; Summary /
    Patch / WOW come from their ops builders. Hard-fails if the native module or the
    combined writer is unavailable (no openpyxl fallback)."""
    import tempfile
    import algotest_native as _an
    if not hasattr(_an, "write_workbook_xlsx"):
        raise RuntimeError("algotest_native.write_workbook_xlsx unavailable — rebuild the wheel")

    summary_ops = _summary_ops(
        cleaned, summary, tm, combo_label, from_date, to_date,
        has_calls, has_puts, has_futures, has_midcap, midcap_summary,
        chron_keys=sorted_keys, patchwise=patchwise, filter_segments=filter_segments,
    )
    patch_ops = None
    if want_patch:
        patch_ops = _patch_wise_ops(
            tm, grouped, sorted_keys, has_midcap, midcap_by_trade,
            has_calls, has_puts, filter_segments=filter_segments,
        )
    # Mirror the openpyxl path: never block the tradesheet on the extra WOW/MOM sheet.
    wow_ops = None
    try:
        from services.optimizer.wow_mom import wow_mom_ops
        wow_ops = wow_mom_ops(cleaned, has_midcap, combo_label or "Strategy", yearly=yearly)
    except Exception as _wm_exc:
        logging.getLogger(__name__).warning("[XLSX] WOW/MOM ops skipped: %s", _wm_exc)

    fd, tp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        _an.write_workbook_xlsx(cleaned, key_order, summary_ops, patch_ops, wow_ops, tp,
                                _rules_ops(rules_sheet) if rules_sheet else None)
        with open(tp, "rb") as f:
            return f.read()
    finally:
        try:
            os.remove(tp)
        except OSError:
            pass


def build_combo_xlsx(
    trades_df: pd.DataFrame,
    summary: Dict[str, Any],
    combo_label: str = "",
    from_date: str = "",
    to_date: str = "",
    midcap_legs=None,
    midcap_spot_adjustment=None,
    midcap_symbol: str = "NIFTYMIDCAP100",
    filter_name: str = "",
    patchwise: bool = False,
    filter_segments=None,
    force_patch_wise: bool = False,
    rules_sheet=None,
    yearly: bool = False,
) -> bytes:
    """
    Build a complete XLSX workbook (Trade Sheet + Summary + optional Patch wise)
    from a trades DataFrame and a summary dict. Returns raw bytes for ZIP embedding.
    When midcap_legs is provided the Midcap overlay is applied via the Rust engine.
    When filter_name is set (or force_patch_wise is True), a Patch wise sheet is
    added — force_patch_wise=True is used for the backtest download so the
    phase-wise tab always appears (matching the frontend ExcelJS workbook), even
    when there is no filter.
    """
    if trades_df is None or (hasattr(trades_df, "empty") and trades_df.empty):
        rows: List[Dict] = []
    else:
        rows = trades_df.where(trades_df.notna(), None).to_dict("records")

    midcap_by_trade, midcap_summary, has_midcap = compute_midcap_for_rows(
        rows, midcap_legs, midcap_spot_adjustment, midcap_symbol,
    )



    key_order, has_calls, has_puts, has_futures = _build_key_order(rows, has_midcap)
    tm, _grouped, _sorted_keys = _aggregate_trades(rows, has_midcap, midcap_by_trade,
                                                   patchwise=patchwise, filter_segments=filter_segments)
    cleaned = _build_cleaned_rows(rows, key_order, tm, has_midcap, midcap_by_trade)

    # ── SINGLE SOURCE OF TRUTH ────────────────────────────────────────────────────
    # Compute the summary metrics ONCE — in Rust (compute_xlsx_summary_metrics
    # delegates to algotest_native.compute_summary_metrics) — and fold them into the
    # `summary` dict that BOTH sheet builders already receive. _summary_layout then
    # READS these numbers instead of re-deriving them, so the per-combo Summary sheet
    # cannot drift from the master summary or from the backtest.
    #
    # Metrics WIN the merge, in every mode:
    #   * plain (non-patchwise, non-midcap) — the Rust engine already pins each value
    #     to the backtest's own (`pin` in summary_metrics.rs), so taking it IS taking
    #     the backtest's number;
    #   * patchwise / midcap — it holds the per-patch-reset / NIFTY+Midcap COMBINED
    #     value, which the plain backtest summary does not describe at all.
    # This also retires a THIRD implementation: _summary_layout's own patchwise
    # outlier-stripped Live DD disagreed with BOTH Rust and the Python reference engine
    # (outlier_dd_1 -1.56 vs the -1.81 those two agree on — gated by
    # tools/summary_metrics_parity at 39 keys / 0 diverging). Two independent
    # implementations against one; the merge removes the outlier.
    try:
        summary = {**(summary or {}), **compute_xlsx_summary_metrics(
            trades_df, summary,
            midcap_legs=midcap_legs, midcap_spot_adjustment=midcap_spot_adjustment,
            midcap_symbol=midcap_symbol, patchwise=patchwise,
            filter_segments=filter_segments,
        )}
    except Exception as exc:
        # NOT skippable. This block is what makes the workbook's Summary sheet
        # agree with the requested basis; skipping it leaves the OVERALL numbers
        # in place inside a workbook that is otherwise built as patchwise, and
        # the sheet gives no sign of it (project rule: no silent basis switch).
        logger.error("[XLSX] summary metric unification FAILED for %s (patchwise=%s): %s",
                     combo_label, patchwise, exc)
        raise RuntimeError(
            "Summary metric unification failed for combo %r (patchwise=%s): %s. "
            "Refusing to emit a workbook whose Summary sheet silently carries the "
            "other basis." % (combo_label, patchwise, exc)
        ) from exc

    # Rust workbook path — ALWAYS. One Rust call builds every sheet: Trade Sheet from
    # `cleaned`; Rules/Summary/Patch/WOW from their ops builders. No Python fallback:
    # a Rust error propagates instead of silently degrading to openpyxl.
    # The openpyxl code below is reference-only, reachable solely when the parity gate
    # (tools/workbook_verify) sets XLSX_PARITY_PY=1. It is never set in production.
    if os.environ.get("XLSX_PARITY_PY") != "1":
        return _build_combo_xlsx_rust(
            cleaned, key_order, summary, tm, _grouped, _sorted_keys,
            combo_label, from_date, to_date, has_calls, has_puts, has_futures,
            has_midcap, midcap_summary, midcap_by_trade, patchwise, filter_segments,
            bool(filter_name or force_patch_wise),
            rules_sheet=rules_sheet,
            # Without this the Rust workbook path silently builds WOW with
            # yearly=False, so every yearly trade collapses into its December
            # expiry's ISO week (~7 cells) instead of spreading by Exit Date.
            yearly=yearly,
        )

    # Built HERE, not before the Rust return above: this workbook is only used by
    # the openpyxl reference path, and constructing one costs 0.277 ms — paid on
    # every combo and thrown away, i.e. ~17.6 s of pure waste on a 63,504 sweep.
    wb = Workbook()
    wb.remove(wb.active)   # drop the default sheet

    _write_trade_sheet(wb, cleaned, key_order)
    _write_summary_sheet(
        wb, cleaned, summary, tm,
        combo_label, from_date, to_date,
        has_calls, has_puts, has_futures,
        has_midcap, midcap_summary,
        chron_keys=_sorted_keys,
        patchwise=patchwise,
        filter_segments=filter_segments,
    )
    if filter_name or force_patch_wise:
        _write_patch_wise_sheet(
            wb, tm, _grouped, _sorted_keys,
            has_midcap, midcap_by_trade,
            has_calls, has_puts,
            filter_segments=filter_segments,
        )

    # WOW & MOM Summary (shared logic with the backtest export + merged summary).
    try:
        from services.optimizer.wow_mom import write_wow_mom_combined
        write_wow_mom_combined(wb, cleaned, has_midcap, combo_label or "Strategy", yearly=yearly)
    except Exception as _wm_exc:  # never block the tradesheet on the extra sheet
        logging.getLogger(__name__).warning("[XLSX] WOW/MOM sheet skipped: %s", _wm_exc)

    # Standalone leg-wise "Rules" sheet as the FIRST tab (backtest download only).
    if rules_sheet:
        try:
            _write_rules_sheet(wb, rules_sheet)
        except Exception as _rs_exc:  # never block the download on the rules sheet
            logging.getLogger(__name__).warning("[XLSX] Rules sheet skipped: %s", _rs_exc)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cleaned_for_combo(
    trades_df: pd.DataFrame,
    midcap_legs=None,
    midcap_spot_adjustment=None,
    midcap_symbol: str = "NIFTYMIDCAP100",
    patchwise: bool = False,
    filter_segments=None,
):
    """
    Return (cleaned_rows, has_midcap) for a combo's trades — the same `cleaned`
    list build_combo_xlsx feeds to the sheets. Used by the merged WOW/MOM
    summary so each combo's block matches its individual tradesheet exactly.
    """
    if trades_df is None or (hasattr(trades_df, "empty") and trades_df.empty):
        rows: List[Dict] = []
    else:
        rows = trades_df.where(trades_df.notna(), None).to_dict("records")
    midcap_by_trade, _midcap_summary, has_midcap = compute_midcap_for_rows(
        rows, midcap_legs, midcap_spot_adjustment, midcap_symbol,
    )
    key_order, _hc, _hp, _hf = _build_key_order(rows, has_midcap)
    tm, _grouped, _sorted_keys = _aggregate_trades(
        rows, has_midcap, midcap_by_trade, patchwise=patchwise, filter_segments=filter_segments)
    cleaned = _build_cleaned_rows(rows, key_order, tm, has_midcap, midcap_by_trade)
    return cleaned, has_midcap
