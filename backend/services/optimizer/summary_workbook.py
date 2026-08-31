"""Master "Optimization Summary" workbook — the SINGLE backend builder.

Replaces the ExcelJS builder that lived in frontend/src/utils/optimSummaryExport.js,
so every .xlsx this product emits is built by openpyxl on the backend:

    tradesheet (backtest download / optim combo / optim ZIP) -> excel_builder.build_combo_xlsx
    WOW & MOM                                                -> wow_mom.write_merged_wow_mom
    Optimization Summary                                     -> THIS MODULE

The frontend still DERIVES what varies per run (the Rules block and the per-combo
`combo_columns`), because that derivation reads the sweep config it already holds —
re-implementing it in Python would create exactly the second implementation this
consolidation exists to remove. It POSTs those derived values here; this module owns
the workbook: columns, layout, styling, number formats.

Cell layout is a 1:1 port of the JS it replaces (rules rows -> blank -> bold header on
#1E3A8A -> data rows, numeric cells "0.00", all column widths 16, freeze at
xSplit=1/ySplit=header row).
"""
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Ported verbatim from MASTER_SUMMARY_COLUMNS (strategyParamSchema.js). `conditional`
# hides a column when that leg type is absent from the run.
MASTER_SUMMARY_COLUMNS: List[Dict[str, str]] = [
    {"key": "sr_no", "label": "Sr. No."},
    {"key": "expiry", "label": "Expiry"},
    {"key": "shifting", "label": "Shifting"},
    {"key": "put_strike_label", "label": "Put ATM or ITM"},
    {"key": "call_strike_label", "label": "Call ATM or ITM"},
    {"key": "spot_adjustment", "label": "Spot Adjustment"},
    {"key": "count", "label": "Trades Count"},
    {"key": "total_pnl", "label": "Net P/L Sum"},
    {"key": "total_pnl_pct", "label": "Net P/L Sum %"},
    {"key": "avg_profit_per_trade", "label": "Net P/L Avg."},
    {"key": "avg_profit_per_trade_pct", "label": "Net P/L Avg. %"},
    {"key": "win_pct", "label": "Winners %"},
    {"key": "avg_win", "label": "Avg. win"},
    {"key": "avg_win_pct", "label": "Avg. win %"},
    {"key": "loss_pct", "label": "Looser %"},
    {"key": "avg_loss", "label": "Avg. Loss"},
    {"key": "avg_loss_pct", "label": "Avg. Loss %"},
    {"key": "expectancy", "label": "Expectancy"},
    {"key": "cagr_options", "label": "CAGR(Options)"},
    {"key": "max_dd_pct", "label": "DD %"},
    {"key": "spot_change", "label": "Spot Change"},
    {"key": "spot_change_pct", "label": "Spot Change %"},
    {"key": "roi_vs_spot", "label": "ROI vs Spot"},
    {"key": "cagr_spot", "label": "CAGR(Spot)"},
    {"key": "car_mdd", "label": "CAR/MDD Booked"},
    {"key": "max_dd_pct", "label": "DD"},
    {"key": "actual_live_dd_max", "label": "Actual Live DD"},
    {"key": "actual_live_dd_avg", "label": "Avg Actual Live DD"},
    {"key": "avg_final_mae", "label": "Avg Combined Final MAE", "conditional": "hasMidcap"},
    {"key": "avg_final_mae", "label": "Avg Final MAE", "conditional": "notMidcap"},
    {"key": "car_mdd_live", "label": "CAR/MDD Live"},
    {"key": "positive_outlier_1", "label": "+ve Outlier 1"},
    {"key": "negative_outlier_1", "label": "-ve Outlier 1"},
    {"key": "outlier_dd_1", "label": "Actual Live DD Without Outlier 1"},
    {"key": "outlier_dd_1_avg", "label": "Avg Actual Live DD Without Outlier 1"},
    {"key": "positive_outlier_2", "label": "+ve Outlier 2"},
    {"key": "negative_outlier_2", "label": "-ve Outlier 2"},
    {"key": "outlier_dd_2", "label": "Actual Live DD Without Outlier 2"},
    {"key": "outlier_dd_2_avg", "label": "Avg Actual Live DD Without Outlier 2"},
    {"key": "positive_outlier_3", "label": "+ve Outlier 3"},
    {"key": "negative_outlier_3", "label": "-ve Outlier 3"},
    {"key": "outlier_dd_3", "label": "Actual Live DD Without Outlier 3"},
    {"key": "outlier_dd_3_avg", "label": "Avg Actual Live DD Without Outlier 3"},
    {"key": "ce_pe_pnl_pct_without_top_1_outliers", "label": "CE + PE + P&L % Without Top 1 Outliers"},
    {"key": "ce_pe_pnl_pct_without_top_2_outliers", "label": "CE + PE + P&L % Without Top 2 Outliers"},
    {"key": "ce_pe_pnl_pct_without_top_3_outliers", "label": "CE + PE + P&L % Without Top 3 Outliers"},
    {"key": "ce_pnl_total", "label": "CE P&L", "conditional": "hasCE"},
    {"key": "ce_pnl_pct", "label": "CE P&L %", "conditional": "hasCE"},
    {"key": "pe_pnl_total", "label": "PE P&L", "conditional": "hasPE"},
    {"key": "pe_pnl_pct", "label": "PE P&L %", "conditional": "hasPE"},
    {"key": "fut_pnl_total", "label": "FUT P&L", "conditional": "hasFUT"},
    {"key": "fut_pnl_pct", "label": "FUT P&L %", "conditional": "hasFUT"},
    {"key": "long_spot_pnl", "label": "Long Spot P&L", "conditional": "hasSpot"},
    {"key": "long_spot_pnl_pct", "label": "Long Spot P&L %", "conditional": "hasSpot"},
    {"key": "midcap_leg_pnl_sum", "label": "Midcap Leg P&L", "conditional": "hasMidcap"},
    {"key": "midcap_leg_pnl_pct_sum", "label": "Midcap Leg P&L %", "conditional": "hasMidcap"},
    {"key": "combined_pnl_sum", "label": "Combined Net P&L", "conditional": "hasMidcap"},
    {"key": "combined_pnl_pct_sum", "label": "Combined Net P&L %", "conditional": "hasMidcap"},
]

_HEADER_FILL = PatternFill("solid", fgColor="FF1E3A8A")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def compute_leg_presence(rows: List[Dict[str, Any]]) -> Dict[str, bool]:
    """Which leg types appear, for hiding conditional columns (mirrors computeLegPresence)."""
    has_ce = has_pe = has_fut = has_spot = has_midcap = False
    for r in rows:
        s = r.get("summary") or {}
        def _nz(key):
            v = s.get(key)
            try:
                return v is not None and abs(float(v)) > 0.01
            except (TypeError, ValueError):
                return False
        has_ce = has_ce or _nz("ce_pnl_total")
        has_pe = has_pe or _nz("pe_pnl_total")
        has_fut = has_fut or _nz("fut_pnl_total")
        has_spot = has_spot or _nz("long_spot_pnl")
        has_midcap = has_midcap or bool(s.get("has_midcap"))
        if has_ce and has_pe and has_fut and has_spot and has_midcap:
            break
    return {"hasCE": has_ce, "hasPE": has_pe, "hasFUT": has_fut, "hasSpot": has_spot,
            "hasMidcap": has_midcap, "notMidcap": not has_midcap}


def visible_columns_for(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    presence = compute_leg_presence(rows)
    return [c for c in MASTER_SUMMARY_COLUMNS
            if not c.get("conditional") or presence.get(c["conditional"]) is True]


def legwise_columns_for(rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Extra master-summary columns: one BLOCK PER LEG, plus the strategy-wide
    adjustment and the Midcap overlay.

    Why: the three packed columns (Put/Call ATM or ITM, Spot Adjustment) squeeze
    every leg into one cell — two same-type legs collapse to "L1_ATM/L3_OTM2",
    adjustments concatenate to "L1RiseBy1%_L2MoveBy1%", a futures leg vanishes,
    and nothing says which INDEX a leg belongs to. These columns are ADDITIVE:
    the originals keep their position so saved filters and downstream sheets are
    unaffected.

    Emitted per leg that ACTUALLY EXISTS (a 2-leg sweep gets no L3 columns), and
    ordered BY LEG — L1 strike, L1 adj, L2 strike, L2 adj — so one leg reads as a
    block instead of being scattered across the sheet. Headers come from the
    first row's leg_cols; leg identity (index/type/expiry/side) is constant for a
    sweep because only strike_selection.* and spot_adjustment.* are ever swept.
    """
    first = next((r for r in rows if (r.get("combo_columns") or {}).get("leg_cols")), None)
    if not first:
        return []                       # pre-change jobs: no leg_cols stored -> unchanged sheet
    cols: List[Dict[str, str]] = [{"key": "overall_adjustment", "label": "Overall Adjustment"}]
    if any((r.get("combo_columns") or {}).get("midcap_leg") for r in rows):
        cols.append({"key": "midcap_leg", "label": "Midcap Leg"})
        cols.append({"key": "midcap_adj", "label": "Midcap Adj"})
    for i, lc in enumerate((first.get("combo_columns") or {}).get("leg_cols") or []):
        # The header can bake in the leg's expiry cadence (e.g. "L1 NF CE Wkly
        # Sell"), but expiry IS a swept axis in some sweeps — asserting row 0's
        # cadence as a static column header over every row silently mislabels
        # every combo whose leg traded a DIFFERENT cadence. Fall back to the
        # expiry-free header when any row's own leg_cols[i]["hdr"] disagrees
        # with row 0's — a leg identity (index/type/side) IS constant for a
        # sweep (only strike_selection.*/spot_adjustment.*/expiry are ever
        # swept), so the stable header is always honest.
        _all_hdrs = {
            (((r.get("combo_columns") or {}).get("leg_cols") or [{}] * (i + 1))[i] or {}).get("hdr")
            for r in rows
            if len(((r.get("combo_columns") or {}).get("leg_cols")) or []) > i
        }
        if len(_all_hdrs) > 1:
            hdr = str(lc.get("hdr_stable") or lc.get("hdr") or f"L{i + 1}")
        else:
            hdr = str(lc.get("hdr") or f"L{i + 1}")
        cols.append({"key": f"__leg{i}_strike", "label": hdr})
        cols.append({"key": f"__leg{i}_adj", "label": f"L{i + 1} Adj"})
        # Per-leg RISK controls (Target / Stop Loss / SL-Buffer / Trail /
        # Re-entry on Target / Re-entry on SL). Emitted ONLY when at least one
        # combo actually carries that control for this leg — a strike/adj-only
        # sweep is byte-identical to before (no empty risk columns). Values come
        # from combo_labeler's leg_cols; _legwise_value resolves __leg{i}_{field}
        # generically, so no extra plumbing is needed.
        for _fld, _lbl in (("tp", "Target"), ("sl", "Stop Loss"), ("slb", "SL Buffer"),
                           ("ts", "Trail SL"), ("rot", "Re-Entry Tgt"), ("ros", "Re-Entry SL")):
            _present = any(
                str((((r.get("combo_columns") or {}).get("leg_cols") or [{}] * (i + 1))[i] or {}).get(_fld) or "").strip()
                not in ("", "-")
                for r in rows
                if len(((r.get("combo_columns") or {}).get("leg_cols")) or []) > i
            )
            if _present:
                cols.append({"key": f"__leg{i}_{_fld}", "label": f"L{i + 1} {_lbl}"})
    return cols


def _legwise_value(key: str, cols: Dict[str, Any]) -> Any:
    """Resolve a __leg{i}_{field} key against that row's stored leg_cols."""
    try:
        idx, field = key[len("__leg"):].split("_", 1)
        lc = (cols.get("leg_cols") or [])[int(idx)]
        return lc.get(field, "")
    except (ValueError, IndexError, TypeError, AttributeError):
        return ""


def build_summary_workbook(rows: List[Dict[str, Any]],
                           rule_rows: Optional[List[List[Any]]] = None,
                           rules_sheet: Optional[List[List[Any]]] = None) -> bytes:
    """Build the Optimization Summary workbook.

    rows        — one per combo: {"summary": {...}, "combo_columns": {...}}
    rule_rows   — [[label, value], ...] for the flat Rules block (frontend-derived).
    rules_sheet — typed rows (title/section/kv/spacer) for the leg-wise "Rules"
                  sheet, built by the SAME frontend buildRulesSheet the backtest
                  uses and rendered by the SAME _write_rules_sheet — so the optim
                  master summary's Rules sheet is identical to the backtest's.
    """
    visible = visible_columns_for(rows)
    # Insert the leg-wise block immediately AFTER the packed "Spot Adjustment"
    # column, so the sweep-config columns stay together and the metrics that
    # follow keep their relative order.
    _extra = legwise_columns_for(rows)
    if _extra:
        _at = next((i for i, c in enumerate(visible)
                    if c["key"] == "spot_adjustment"), len(visible) - 1)
        visible = visible[:_at + 1] + _extra + visible[_at + 1:]
    # Rust render path — the LAST openpyxl builder on a user-facing path.
    # The block below is unchanged; it just runs against the ops-emitting
    # stand-in and is rendered by algotest_native. openpyxl charged a
    # style-table hash per cell assignment, which at 63,504 combos (one row
    # each, ~40 columns) is millions of hashes inside an API request.
    from services.optimizer.wow_mom import _OpsWorkbook

    wb = _OpsWorkbook(with_active=True)
    ws = wb.active
    ws.title = "Optimization Summary"

    # The Rules block no longer sits at the top of this sheet — the full leg-wise
    # config lives in a dedicated "Rules" sheet (rules_sheet below), so the
    # Optimization Summary sheet starts straight at the combos table. rule_rows is
    # accepted for API compatibility but intentionally not rendered here.
    header_row_idx = 1

    data_cols = [c for c in visible if c["key"] != "sr_no"]
    ws.append(["Sr. No."] + [c["label"] for c in data_cols])
    for cell in ws[header_row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        # openpyxl spells the JS "middle" as "center" — same rendered alignment.
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # String coordinate, NOT ws.cell(...): ws.cell() instantiates the cell and bumps
    # max_row, which made append() skip a row and leave a blank line under the header.
    ws.freeze_panes = f"B{header_row_idx + 1}"

    for i, row in enumerate(rows):
        summary = row.get("summary") or {}
        cols = row.get("combo_columns") or {}
        # combo_columns wins over summary — identical precedence to the JS.
        vals = [_legwise_value(c["key"], cols) if c["key"].startswith("__leg")
                else (cols[c["key"]] if c["key"] in cols else summary.get(c["key"]))
                for c in data_cols]
        ws.append([i + 1] + vals)
        row_cells = ws[ws.max_row]
        for j, cell in enumerate(row_cells):
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                # CAR/MDD (and everything else) render as a plain ratio, not a percent.
                cell.number_format = "0.00"

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    # Leg-wise "Rules" sheet as the FIRST tab — identical to the backtest download
    # (same buildRulesSheet rows, same _write_rules_sheet renderer).
    # `_rules_ops` already renders this sheet for the Rust writer (it is what the
    # per-combo tradesheet uses), so the ops go straight in front of the summary
    # sheet — index 0, same "Rules is the first tab" contract as before.
    sheets = []
    if rules_sheet:
        from services.optimizer.excel_builder import _rules_ops
        sheets.append(_rules_ops(rules_sheet))
    sheets.append(ws.to_ops())

    # No openpyxl fallback: a native failure raises rather than quietly handing
    # back a workbook that is missing its Rules tab (what the old try/except
    # around _write_rules_sheet did — the download simply arrived incomplete).
    import os
    import tempfile

    import algotest_native as _an

    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        _an.write_layout_workbook_xlsx(sheets, tmp)
        with open(tmp, "rb") as fh:
            return fh.read()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass
