"""Ingest the EOD actual-delta history workbook into a lean feather the engine
reads for delta strike selection.

Source: eod_delta_full_history_updated.xlsx — one sheet per index (NIFTY,
MIDCPNIFTY), columns:
  Date, Symbol, Expiry, DaysToExpiry, Strike,
  CE SpotClose, CE Premium, CE IV%, CE Delta, CE Timestamp,
  PE SpotClose, PE Premium, PE IV%, PE Delta, PE Timestamp

We keep only what strike selection needs — symbol, date, expiry, strike, and the
two actual deltas — so the on-disk feather stays small and loads fast. Deltas are
stored as signed (CE ~ +, PE ~ -); the lookup takes abs() against the target.

Usage (inside the backend container so pyarrow/polars + /data/cache are present):
    python ingest_delta_history.py /app/_delta_ingest.xlsx
Writes /data/cache/delta_history.feather. Re-run whenever the workbook updates —
it fully overwrites (idempotent).
"""
import sys
from datetime import date, datetime

import openpyxl
import polars as pl

OUT = "/data/cache/delta_history.feather"
# 0-based column positions in the source sheets (see header above).
C_DATE, C_SYMBOL, C_EXPIRY, C_STRIKE, C_CE_DELTA, C_PE_DELTA = 0, 1, 2, 4, 8, 13


def _iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    return s[:10] if s else None


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def ingest(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    syms, dates, exps, strikes, ce_d, pe_d = [], [], [], [], [], []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        first = True
        n = 0
        for row in ws.iter_rows(values_only=True):
            if first:                       # header
                first = False
                continue
            d = _iso(row[C_DATE]); ex = _iso(row[C_EXPIRY]); st = _f(row[C_STRIKE])
            if not (d and ex and st):
                continue
            cd = _f(row[C_CE_DELTA]); pd_ = _f(row[C_PE_DELTA])
            if cd is None and pd_ is None:  # nothing usable on this row
                continue
            sym = str(row[C_SYMBOL] or sheet).strip().upper()
            syms.append(sym); dates.append(d); exps.append(ex); strikes.append(st)
            ce_d.append(cd); pe_d.append(pd_)
            n += 1
        print(f"  {sheet}: {n} rows kept")
    df = pl.DataFrame({
        "symbol": syms, "date": dates, "expiry": exps,
        "strike": strikes, "ce_delta": ce_d, "pe_delta": pe_d,
    })
    df.write_ipc(OUT)
    print(f"wrote {OUT}: {df.height} rows, symbols={sorted(df['symbol'].unique().to_list())}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "/app/_delta_ingest.xlsx"
    ingest(src)
