#!/usr/bin/env python3
"""End-to-end check: build a REAL per-combo tradesheet XLSX via the production
write_combo_xlsx path, then read back the Summary sheet's CAGR (Spot) cell and
compare it to the backtest's cagr_spot. Proves the user-facing tradesheet Summary
now uses the leg-independent spot-level cagr_spot (no per-leg/per-trade compound).

Run inside a worker container:
    docker exec algotest-worker-backtests python /app/tools/verify_combo_xlsx.py
Read-only w.r.t. production code; writes a throwaway XLSX under a temp job id.
"""
from __future__ import annotations
import os, sys, glob
for _p in ("/app", os.path.dirname(os.path.dirname(os.path.abspath(__file__)))):
    if _p not in sys.path: sys.path.insert(0, _p)
import pandas as pd
from openpyxl import load_workbook

# Reuse the harness payloads so this matches the parity baseline exactly.
from tools.parity_harness import PAYLOADS  # type: ignore
from services.algotest_job import execute_algotest_job
from services.optimizer import result_store
from services.optimizer.result_store import write_combo_xlsx, get_trades_dir

JOB = "_parity_verify_tmp"

def _read_cagr_spot_cell(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for i, c in enumerate(row):
                if isinstance(c.value, str) and c.value.strip() == "CAGR (Spot)":
                    # value is in a nearby cell on the same row
                    for c2 in row[i+1:]:
                        if c2.value not in (None, ""):
                            return c2.value, ws.title
    return None, None

def main() -> int:
    for name, payload in PAYLOADS:
        print("=" * 70); print("PAYLOAD:", name)
        try:
            res = execute_algotest_job(dict(payload))
            trades = res.get("trades") or []
            summary = res.get("summary") or {}
            bt_cagr_spot = summary.get("cagr_spot")
            df = pd.DataFrame(trades)
            label = "_test_combo"
            write_combo_xlsx(JOB, label, df, summary, index_str="NIFTY",
                             from_date=payload["from_date"], to_date=payload["to_date"])
            path = os.path.join(get_trades_dir(JOB), f"{label}.xlsx")
            files = glob.glob(os.path.join(get_trades_dir(JOB), "*.xlsx"))
            xlsx = path if os.path.isfile(path) else (files[0] if files else None)
            if not xlsx:
                print("  [ERROR] combo XLSX not written"); continue
            sheet_val, sheet = _read_cagr_spot_cell(xlsx)
            print(f"  backtest cagr_spot   = {bt_cagr_spot}")
            print(f"  tradesheet XLSX cell = {sheet_val}  (sheet '{sheet}')")
            # tradesheet cell is formatted like '25.37%'; strip for compare
            sv = None
            if isinstance(sheet_val, str):
                try: sv = float(sheet_val.replace("%", "").strip())
                except Exception: sv = None
            elif isinstance(sheet_val, (int, float)):
                sv = float(sheet_val)
            ok = (sv is not None and bt_cagr_spot is not None
                  and abs(sv - float(bt_cagr_spot)) < 0.01)
            print(f"  MATCH: {'YES' if ok else 'NO'}")
            # cleanup this combo's files
            for f in glob.glob(os.path.join(get_trades_dir(JOB), "*")):
                try: os.remove(f)
                except OSError: pass
        except Exception as exc:
            import traceback
            print(f"  [ERROR] {type(exc).__name__}: {exc}")
            print("  " + "\n  ".join(traceback.format_exc().strip().splitlines()[-6:]))
    return 0

if __name__ == "__main__":
    try: main()
    except Exception as exc:
        print("[VERIFY ERROR]", exc)
    sys.exit(0)
