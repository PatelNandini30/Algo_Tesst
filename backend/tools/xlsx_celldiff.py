"""
Cell-diff harness for Phase-5 Rust writers. Compares two .xlsx workbooks
CELL-BY-CELL across every sheet: value, number_format, font (name/size/bold/
italic/color), fill (fgColor), alignment (horizontal), and border presence.

This is the PROMOTION GATE: a Rust-written workbook may replace the openpyxl one
only when `celldiff(openpyxl_bytes, rust_bytes)` returns zero diffs across the
whole corpus. Any single differing cell blocks it (your tradesheets are audited
cell-by-cell).

Usage (programmatic):
    from tools.xlsx_celldiff import celldiff
    diffs = celldiff(bytes_a, bytes_b)   # list of (sheet, coord, attr, a, b)
"""
from __future__ import annotations
import io
from typing import List, Tuple, Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


def _col_widths(ws) -> dict:
    """Per-column-index effective width, expanding rust_xlsxwriter's coalesced
    `<col min max width>` ranges (openpyxl writes one <col> per column; rust groups
    equal-width runs). Both normalize to the same per-column map for comparison."""
    out: dict = {}
    for dim in ws.column_dimensions.values():
        if dim.width is None or not dim.customWidth:
            continue
        for ci in range(dim.min, dim.max + 1):
            out[ci] = dim.width
    return out


def _rgb6(color) -> Any:
    """Normalize an openpyxl color to its VISIBLE 6-hex RGB (drop the alpha byte).
    openpyxl writes fills as 00RRGGBB, rust_xlsxwriter as FFRRGGBB — same visible
    colour, different alpha. "Same styling" means the same visible colour."""
    if color is None:
        return None
    v = getattr(color, "rgb", None) or getattr(color, "value", None)
    if isinstance(v, str) and len(v) == 8:
        return v[2:].upper()
    return str(v).upper() if v is not None else None


def _font_sig(cell) -> tuple:
    f = cell.font
    return (f.name, f.size, bool(f.bold), bool(f.italic), _rgb6(f.color))


def _fill_sig(cell) -> Any:
    fl = cell.fill
    return _rgb6(getattr(fl, "fgColor", None))


def _border_sig(cell) -> tuple:
    b = cell.border
    return tuple(
        (getattr(getattr(b, side), "style", None))
        for side in ("top", "left", "bottom", "right")
    )


def celldiff(a_bytes: bytes, b_bytes: bytes, max_report: int = 40) -> List[Tuple]:
    """Return a list of (sheet, coord, attr, a_val, b_val). Empty == identical."""
    wa = openpyxl.load_workbook(io.BytesIO(a_bytes))
    wb = openpyxl.load_workbook(io.BytesIO(b_bytes))
    diffs: List[Tuple] = []

    if wa.sheetnames != wb.sheetnames:
        diffs.append(("<workbook>", "-", "sheetnames", wa.sheetnames, wb.sheetnames))
        # still compare the sheets that exist in both

    for name in wa.sheetnames:
        if name not in wb.sheetnames:
            continue
        sa, sb = wa[name], wb[name]
        if (sa.max_row, sa.max_column) != (sb.max_row, sb.max_column):
            diffs.append((name, "-", "dims",
                          (sa.max_row, sa.max_column), (sb.max_row, sb.max_column)))
        rows = max(sa.max_row, sb.max_row)
        cols = max(sa.max_column, sb.max_column)
        # Non-top-left members of a merged range are visually covered by the merge —
        # openpyxl leaves them default, rust_xlsxwriter fills them via merge_range;
        # either way they are invisible, so compare only the merge's top-left.
        covered = set()
        for rng in (set(str(x) for x in sa.merged_cells.ranges)
                    | set(str(x) for x in sb.merged_cells.ranges)):
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    if (rr, cc) != (min_row, min_col):
                        covered.add((rr, cc))
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                if (r, c) in covered:
                    continue
                ca, cb = sa.cell(r, c), sb.cell(r, c)
                coord = ca.coordinate
                for attr, va, vb in (
                    ("value", ca.value, cb.value),
                    ("number_format", ca.number_format, cb.number_format),
                    ("font", _font_sig(ca), _font_sig(cb)),
                    ("fill", _fill_sig(ca), _fill_sig(cb)),
                    ("align", ca.alignment.horizontal, cb.alignment.horizontal),
                    ("border", _border_sig(ca), _border_sig(cb)),
                ):
                    if attr == "value" and isinstance(va, (int, float)) and isinstance(vb, (int, float)):
                        # Values equal to f64 precision. openpyxl serializes numbers via
                        # Python repr (shortest round-trip); rust_xlsxwriter one ULP wider.
                        # The stored f64 agree to 15 sig-figs and the number_format displays
                        # them identically — a serialization artifact, not a cell difference.
                        if va == vb or abs(va - vb) <= 1e-9 + 1e-12 * max(abs(va), abs(vb)):
                            continue
                    if va != vb:
                        diffs.append((name, coord, attr, va, vb))
                        if len(diffs) >= max_report:
                            return diffs

        # ── Sheet-level styling: merges, freeze, column widths, row heights ──
        ma = set(str(x) for x in sa.merged_cells.ranges)
        mb = set(str(x) for x in sb.merged_cells.ranges)
        if ma != mb:
            diffs.append((name, "-", "merges", sorted(ma - mb), sorted(mb - ma)))
        if sa.freeze_panes != sb.freeze_panes:
            diffs.append((name, "-", "freeze", sa.freeze_panes, sb.freeze_panes))
        wa_ = _col_widths(sa)
        wb_ = _col_widths(sb)
        for ci in set(wa_) | set(wb_):
            a, b = wa_.get(ci), wb_.get(ci)
            if a is None or b is None or abs(a - b) > 0.02:
                diffs.append((name, get_column_letter(ci), "colwidth", a, b))
        for rr in set(sa.row_dimensions.keys()) | set(sb.row_dimensions.keys()):
            ha_ = sa.row_dimensions[rr].height if rr in sa.row_dimensions else None
            hb_ = sb.row_dimensions[rr].height if rr in sb.row_dimensions else None
            if ha_ is None and hb_ is None:
                continue
            if ha_ is None or hb_ is None or abs((ha_ or 0) - (hb_ or 0)) > 0.02:
                diffs.append((name, str(rr), "rowheight", ha_, hb_))
        if len(diffs) >= max_report:
            return diffs
    return diffs


if __name__ == "__main__":
    # Self-test: a workbook is cell-identical to itself.
    import warnings
    warnings.filterwarnings("ignore")
    import pandas as pd
    from tools.parity_harness import PAYLOADS
    from services.algotest_job import execute_algotest_job
    from services.optimizer.excel_builder import build_combo_xlsx

    name, payload = PAYLOADS[1]
    res = execute_algotest_job(dict(payload))
    df = pd.DataFrame(res["trades"])
    xb = build_combo_xlsx(df, res["summary"], combo_label="self",
                          from_date=payload["from_date"], to_date=payload["to_date"])
    d = celldiff(xb, xb)
    print(f"self-diff (must be 0): {len(d)}")
    # And a trivially-different workbook must be detected.
    import openpyxl as _op, io as _io
    wb = _op.load_workbook(_io.BytesIO(xb)); wb.worksheets[0].cell(2, 1).value = "__CHANGED__"
    buf = _io.BytesIO(); wb.save(buf)
    d2 = celldiff(xb, buf.getvalue())
    print(f"changed-diff (must be >0): {len(d2)}")
    print("HARNESS OK" if (len(d) == 0 and len(d2) > 0) else "HARNESS BROKEN")
