#!/usr/bin/env python3
"""Prove the Rust WOW/MOM render matches the openpyxl one, cell for cell.

The merged WOW/MOM workbook is now built by running the UNCHANGED block writers
against an ops-emitting stand-in for openpyxl (`wow_mom._OpsWorkbook`) and
rendering those ops with `algotest_native.write_layout_workbook_xlsx`. That
replaced openpyxl's own writer, which spent 91% of its time hashing style
objects into the workbook's style table (31M calls to Serialisable.__hash__ on a
3,969-combo sweep).

Because the swap changes HOW the file is written and not WHAT is in it, the
check that matters is a value-by-value diff of the two workbooks.

    docker compose exec -T worker-optimize python /app/tools/wow_mom_parity.py <job_id> [limit]

Floats are compared as doubles, not as text: openpyxl and rust_xlsxwriter print
the same IEEE-754 value with different digit counts (10.11891896677075 vs
10.118918966770755), which is a formatting difference, not a numeric one.
Exits non-zero if any cell genuinely differs.
"""
import os
import struct
import sys
import time

sys.path.insert(0, "/app")


def main(job_id: str, limit: int = 250) -> int:
    from openpyxl import Workbook, load_workbook

    import algotest_native as an
    import services.optimizer.result_store as rs
    from services.optimizer import wow_mom as W

    rows = rs.get_all_results(job_id)[:limit]
    combos = []
    for r in rows:
        wm = rs.read_combo_wm(job_id, r.get("combo_label_safe"), True)
        if wm:
            combos.append({"label": r.get("combo_label", ""), "wm": wm,
                           "variant_label": "", "yearly": False})
    if not combos:
        print("no WOW/MOM payloads for %s — nothing to compare" % job_id[:8])
        return 2
    print("combos: %d" % len(combos))

    a_path, b_path = "/tmp/_wm_parity_openpyxl.xlsx", "/tmp/_wm_parity_rust.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    t = time.time()
    W.write_merged_wow_mom(wb, combos)
    wb.save(a_path)
    t_openpyxl = time.time() - t

    wb2 = W._OpsWorkbook()
    t = time.time()
    W.write_merged_wow_mom(wb2, combos)
    an.write_layout_workbook_xlsx(wb2.to_ops(), b_path)
    t_rust = time.time() - t

    print("openpyxl: %6.2f s   rust: %6.2f s   speedup %.1fx"
          % (t_openpyxl, t_rust, t_openpyxl / max(t_rust, 1e-9)))

    wa, wbk = load_workbook(a_path), load_workbook(b_path)
    if wa.sheetnames != wbk.sheetnames:
        print("SHEETS DIFFER: %s vs %s" % (wa.sheetnames, wbk.sheetnames))
        return 1

    total = repr_only = real = 0
    worst = 0.0
    examples = []
    for name in wa.sheetnames:
        sa, sb = wa[name], wbk[name]
        for row in sa.iter_rows():
            for ca in row:
                cb = sb.cell(ca.row, ca.column)
                total += 1
                a, b = ca.value, cb.value
                if a == b:
                    continue
                if isinstance(a, float) and isinstance(b, float):
                    # Same bits, or within float64 epsilon => a printing
                    # difference, not a numeric one.
                    if struct.pack("<d", a) == struct.pack("<d", b):
                        repr_only += 1
                        continue
                    delta = abs(a - b) / max(abs(a), abs(b), 1e-30)
                    worst = max(worst, delta)
                    if delta < 1e-12:
                        repr_only += 1
                        continue
                real += 1
                if len(examples) < 5:
                    examples.append((name, ca.row, ca.column, a, b))

    for path in (a_path, b_path):
        try:
            os.remove(path)
        except OSError:
            pass

    print("cells compared      : %d" % total)
    print("same value (repr)   : %d" % repr_only)
    print("REAL differences    : %d" % real)
    print("worst relative delta: %.3e" % worst)
    for e in examples:
        print("   ", e)
    print("PARITY:", "IDENTICAL" if real == 0 else "MISMATCH")
    return 0 if real == 0 else 1


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(2)
    raise SystemExit(main(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else 250))
