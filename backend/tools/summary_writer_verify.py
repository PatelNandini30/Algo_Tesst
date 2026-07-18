"""Verify the Rust Summary-sheet layout writer is cell-identical to openpyxl.

    docker exec -w /app algotest-backend python -m tools.summary_writer_verify

openpyxl reference = _write_summary_sheet (Summary-only workbook).
Rust = _summary_ops(...) -> algotest_native.write_layout_sheet_xlsx.
Also runs a synthetic numeric-merge smoke (the midcap ROI cell path).
"""
import warnings; warnings.filterwarnings("ignore")
import io, os, tempfile
from openpyxl import Workbook

from tools.parity_harness import PAYLOADS
from services.algotest_job import execute_algotest_job
from services.optimizer import excel_builder as eb
from tools.xlsx_celldiff import celldiff
import algotest_native as an

total = 0
for name, payload in PAYLOADS:
    res = execute_algotest_job(dict(payload))
    rows = res.get("trades") or []
    if not rows:
        print(f"{name}: no trades, skip"); continue

    mbt, msumm, has_mc = eb.compute_midcap_for_rows(rows, None, None, "NIFTYMIDCAP100")
    tm, _g, sk = eb._aggregate_trades(rows, has_mc, mbt, patchwise=False, filter_segments=None)
    ko, hc, hp, hf = eb._build_key_order(rows, has_mc)
    cleaned = eb._build_cleaned_rows(rows, ko, tm, has_mc, mbt)

    args = ("combo-" + name, payload["from_date"], payload["to_date"], hc, hp, hf,
            has_mc, msumm)
    kw = dict(chron_keys=sk, patchwise=False, filter_segments=None)

    # openpyxl reference (Summary only)
    wb = Workbook(); wb.remove(wb.active)
    eb._write_summary_sheet(wb, cleaned, res["summary"], tm, *args, **kw)
    buf = io.BytesIO(); wb.save(buf); py_bytes = buf.getvalue()

    # Rust
    ops = eb._summary_ops(cleaned, res["summary"], tm, *args, **kw)
    tp = tempfile.mktemp(suffix=".xlsx")
    an.write_layout_sheet_xlsx(ops, tp)
    with open(tp, "rb") as f:
        rust_bytes = f.read()
    os.remove(tp)

    diffs = celldiff(py_bytes, rust_bytes, max_report=30)
    total += len(diffs)
    print(f"{name}: {len(ops['cells'])} cells, {len(ops['merges'])} merges -> {len(diffs)} diffs")
    for d in diffs[:15]:
        print("   ", d)

# Synthetic numeric-merge smoke (midcap ROI path): a float value in a merged range.
tp = tempfile.mktemp(suffix=".xlsx")
synth = {
    "name": "Summary",
    "cells": [
        {"r": 1, "c": 4, "v": 1.5007, "bold": True, "size": 11, "fc": "1E7E34",
         "bg": "FFFFFF", "align": "C", "border": True, "nfmt": "General"},
        {"r": 2, "c": 1, "v": "hello", "bold": True, "size": 13, "fc": "FFFFFF",
         "bg": "1F3864", "align": "C", "border": False, "nfmt": None},
    ],
    "merges": [(1, 4, 5), (2, 1, 5)],
    "row_heights": [(1, 20.0), (2, 26.0)],
    "col_widths": [(4, 30.0)],
    "freeze": None,
}
an.write_layout_sheet_xlsx(synth, tp)
import openpyxl
w = openpyxl.load_workbook(tp); ws = w["Summary"]
os.remove(tp)
ok = (abs((ws["D1"].value or 0) - 1.5007) < 1e-9 and ws["A2"].value == "hello")
print(f"numeric-merge smoke: D1={ws['D1'].value!r} A2={ws['A2'].value!r} -> {'OK' if ok else 'FAIL'}")

print(f"\nSUMMARY TOTAL cell-diffs: {total}  {'-> CELL-IDENTICAL' if total == 0 else '(iterate)'}")
