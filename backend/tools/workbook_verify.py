"""Full-workbook parity gate: build_combo_xlsx openpyxl (XLSX_PARITY_PY=1) vs Rust.
Compares the ENTIRE workbook — all sheets, sheet order + names — cell-by-cell.

    docker exec -w /app algotest-backend python -m tools.workbook_verify
"""
import warnings; warnings.filterwarnings("ignore")
import os
import pandas as pd

from tools.parity_harness import PAYLOADS
from services.algotest_job import execute_algotest_job
from services.optimizer.excel_builder import build_combo_xlsx
from tools.xlsx_celldiff import celldiff


def _one(name, payload, force_patch, synth_midcap):
    res = execute_algotest_job(dict(payload))
    rows = res.get("trades") or []
    if not rows:
        print(f"{name}: no trades, skip"); return 0
    df = pd.DataFrame(rows)
    midcap_legs = None
    if synth_midcap:
        # exercise the midcap branch (3-phase patch + combined summary/wow)
        midcap_legs = [{"symbol": "NIFTYMIDCAP100", "mode": "hypothetical_future",
                        "lots": 1, "position": "SELL"}]
    # A rules_sheet is passed on every case so the leg-wise "Rules" tab (first sheet,
    # rendered in Rust via build_layout_sheet) stays inside the parity gate too.
    rules_sheet = [["title", "Strategy Rules"], ["section", "Leg 1 (CE)"],
                   ["kv", "Strike Gap", "100"], ["kv", "Slippage", "0.5%"], ["spacer"],
                   ["section", "Leg 2 (FUT)"], ["kv", "Own Spot Adjustment", "2% (Yearly)"]]
    kw = dict(combo_label="combo-" + name, from_date=payload["from_date"],
              to_date=payload["to_date"], force_patch_wise=force_patch,
              midcap_legs=midcap_legs, rules_sheet=rules_sheet)

    # Rust is the only production path; XLSX_PARITY_PY=1 is what reaches the
    # reference openpyxl builder, and exists purely for this gate.
    os.environ["XLSX_PARITY_PY"] = "1"
    try:
        py_bytes = build_combo_xlsx(df, res["summary"], **kw)
    finally:
        os.environ.pop("XLSX_PARITY_PY", None)
    rust_bytes = build_combo_xlsx(df, res["summary"], **kw)

    diffs = celldiff(py_bytes, rust_bytes, max_report=40)
    import openpyxl, io
    wsa = openpyxl.load_workbook(io.BytesIO(py_bytes)).sheetnames
    wsb = openpyxl.load_workbook(io.BytesIO(rust_bytes)).sheetnames
    tag = "midcap" if synth_midcap else "plain"
    tag += "+patch" if force_patch else ""
    ok = (wsa == wsb)
    print(f"{name} [{tag}]: sheets {wsa} {'==' if ok else '!= ' + str(wsb)} -> {len(diffs)} diffs")
    for d in diffs[:15]:
        print("    ", d)
    return len(diffs) + (0 if ok else 1)


total = 0
for name, payload in PAYLOADS:
    total += _one(name, payload, force_patch=True, synth_midcap=False)
    total += _one(name, payload, force_patch=True, synth_midcap=True)

print(f"\nWORKBOOK TOTAL diffs: {total}  {'-> CELL-IDENTICAL' if total == 0 else '(iterate)'}")
