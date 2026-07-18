"""Verify the refactored Patch-wise sheet:
  1. REGRESSION: refactored openpyxl _write_patch_wise_sheet == git-HEAD version.
  2. PARITY:     Rust write_layout_sheet_xlsx(_patch_wise_ops(...)) == openpyxl.

    docker exec -w /app algotest-backend python -m tools.patchwise_writer_verify
"""
import warnings; warnings.filterwarnings("ignore")
import io, os, tempfile, importlib.util
from openpyxl import Workbook

from tools.parity_harness import PAYLOADS
from services.algotest_job import execute_algotest_job
from services.optimizer import excel_builder as eb
from tools.xlsx_celldiff import celldiff
import algotest_native as an

# Load git-HEAD excel_builder as a separate module for the regression baseline.
ORIG = None
orig_path = "/tmp/excel_builder_orig.py"
if os.path.exists(orig_path):
    spec = importlib.util.spec_from_file_location("excel_builder_orig", orig_path)
    ORIG = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(ORIG)
    except Exception as e:
        print("orig load failed:", e); ORIG = None

reg_total = 0
par_total = 0
for name, payload in PAYLOADS:
    res = execute_algotest_job(dict(payload))
    rows = res.get("trades") or []
    if not rows:
        print(f"{name}: no trades, skip"); continue

    mbt, msumm, has_mc = eb.compute_midcap_for_rows(rows, None, None, "NIFTYMIDCAP100")
    tm, g, sk = eb._aggregate_trades(rows, has_mc, mbt, patchwise=True, filter_segments=None)
    ko, hc, hp, hf = eb._build_key_order(rows, has_mc)

    pw_args = (tm, g, sk, has_mc, mbt, hc, hp)

    # openpyxl reference (Patch wise only)
    wb = Workbook(); wb.remove(wb.active)
    eb._write_patch_wise_sheet(wb, *pw_args, filter_segments=None)
    if "Patch wise" not in wb.sheetnames:
        print(f"{name}: no patch sheet, skip"); continue
    buf = io.BytesIO(); wb.save(buf); py_bytes = buf.getvalue()

    # regression vs git-HEAD
    if ORIG is not None:
        wbo = Workbook(); wbo.remove(wbo.active)
        ORIG._write_patch_wise_sheet(wbo, *pw_args, filter_segments=None)
        bo = io.BytesIO(); wbo.save(bo)
        dreg = celldiff(bo.getvalue(), py_bytes)
        reg_total += len(dreg)
        print(f"{name}: REGRESSION {len(dreg)} diffs")
        for d in dreg[:8]: print("    ", d)

    # Rust parity
    ops = eb._patch_wise_ops(*pw_args, filter_segments=None)
    tp = tempfile.mktemp(suffix=".xlsx")
    an.write_layout_sheet_xlsx(ops, tp)
    with open(tp, "rb") as f: rust_bytes = f.read()
    os.remove(tp)
    dpar = celldiff(py_bytes, rust_bytes, max_report=30)
    par_total += len(dpar)
    print(f"{name}: {len(ops['cells'])} cells, {len(ops['merges'])} merges -> PARITY {len(dpar)} diffs")
    for d in dpar[:15]: print("    ", d)

print(f"\nREGRESSION TOTAL: {reg_total}  {'(orig not loaded)' if ORIG is None else ('SAFE' if reg_total==0 else 'BROKE')}")
print(f"PARITY TOTAL:     {par_total}  {'-> CELL-IDENTICAL' if par_total==0 else '(iterate)'}")
