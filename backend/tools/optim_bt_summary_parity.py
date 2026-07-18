"""Parity: OPTIM master summary (compute_xlsx_summary_metrics) VS the BACKTEST
tradesheet's Summary sheet (build_combo_xlsx -> _write_summary_sheet), for the SAME
strategy. Proves the two summary sites produce identical numbers (the HARD RULE:
optim summary == backtest summary). Runs on the LIVE deployed code.

    docker exec -w /app algotest-backend python -m tools.optim_bt_summary_parity
"""
import warnings; warnings.filterwarnings("ignore")
import io
import pandas as pd
import openpyxl

from tools.parity_harness import PAYLOADS
from services.algotest_job import execute_algotest_job
from services.optimizer.excel_builder import compute_xlsx_summary_metrics, build_combo_xlsx


def _num(s):
    """Parse a displayed Summary-sheet value ('+9.05%', '₹-190.50', '8 trades') -> float."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).replace("+", "").replace("%", "").replace("₹", "").replace(",", "")
    t = t.replace(" trades", "").strip()
    try:
        return float(t)
    except ValueError:
        return None


# sheet-label  ->  metric-dict key  (only metrics that appear in BOTH)
MAP = {
    "Avg Profit on Winners": "avg_win_pct",
    "Avg Loss on Losers": "avg_loss_pct",
    "Net P/L Avg %": "avg_profit_per_trade_pct",
    "CAGR (Options)": "cagr_options",
    "CAGR (Spot)": "cagr_spot",
    "Max Drawdown": "max_dd_pct",
    "Return / MaxDD": "car_mdd",
    "CAR/MDD (Booked)": "car_mdd",
    "CAR/MDD Live": "car_mdd_live",
    "Actual Live DD (min)": "actual_live_dd_max",
    "Avg Actual Live DD": "actual_live_dd_avg",
    "Avg Final MAE": "avg_final_mae",
    "CE P&L": "ce_pnl_total",
    "PE P&L": "pe_pnl_total",
    "Spot P&L": "long_spot_pnl",
    "+ve Outlier 1": "positive_outlier_1",
    "+ve Outlier 2": "positive_outlier_2",
    "+ve Outlier 3": "positive_outlier_3",
    "-ve Outlier 1": "negative_outlier_1",
    "-ve Outlier 2": "negative_outlier_2",
    "-ve Outlier 3": "negative_outlier_3",
}


def _sheet_kv(ws):
    """label -> value across the A/B, D/E, and A/E (merged outlier) column pairs."""
    kv = {}
    for r in range(1, ws.max_row + 1):
        a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
        d, e = ws.cell(r, 4).value, ws.cell(r, 5).value
        if a and b not in (None, ""):
            kv.setdefault(str(a).strip(), b)
        if d and e not in (None, ""):
            kv.setdefault(str(d).strip(), e)
        if a and e not in (None, "") and (b in (None, "")):
            kv.setdefault(str(a).strip(), e)
    return kv


grand = 0
for name, payload in PAYLOADS:
    res = execute_algotest_job(dict(payload))
    if not res.get("trades"):
        print(f"{name}: no trades, skip"); continue
    df = pd.DataFrame(res["trades"])

    optim = compute_xlsx_summary_metrics(df, res["summary"])          # OPTIM master
    xb = build_combo_xlsx(df, res["summary"], combo_label=name,
                          from_date=payload["from_date"], to_date=payload["to_date"])
    kv = _sheet_kv(openpyxl.load_workbook(io.BytesIO(xb))["Summary"])  # BACKTEST sheet

    diffs = []
    checked = 0
    for label, key in MAP.items():
        if label not in kv or key not in optim or optim[key] is None:
            continue
        sheet_v = _num(kv[label])
        opt_v = optim[key]
        if sheet_v is None:
            continue
        checked += 1
        # sheet displays rounded (2dp % / 4dp ratios); compare at the coarser display precision
        if abs(round(opt_v, 2) - round(sheet_v, 2)) > 0.005:
            diffs.append((label, key, f"optim={opt_v}", f"sheet={sheet_v}"))
    grand += len(diffs)
    tag = "OK" if not diffs else "MISMATCH"
    print(f"{name}: checked {checked} shared metrics -> {len(diffs)} diffs  [{tag}]")
    for d in diffs:
        print("    ", d)

print(f"\nOPTIM-vs-BACKTEST SUMMARY TOTAL diffs: {grand}  "
      f"{'-> IDENTICAL' if grand == 0 else '(DIVERGED)'}")
