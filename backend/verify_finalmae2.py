import sys, pandas as pd
sys.path.insert(0, '/app')
from services.algotest_job import execute_algotest_job
from services.optimizer import excel_builder as eb
import openpyxl, io

def run(legs, tag, multi):
    req={"index":"NIFTY","from_date":"2020-01-01","to_date":"2024-12-31",
     "strategy_type":"positional","underlying":"cash","expiry_type":"WEEKLY",
     "entry_dte":1,"exit_dte":0,"slippage_pct":0,"charges_enabled":False,
     "square_off_mode":"partial","rollover_toggle":True,"no_cache":True,"legs":legs}
    res = execute_algotest_job(req)
    df = pd.DataFrame(res.get("trades") or [])
    xb = eb.build_combo_xlsx(df, res.get("summary") or {}, combo_label=tag,
                             from_date=req["from_date"], to_date=req["to_date"])
    wb = openpyxl.load_workbook(io.BytesIO(xb))
    ws = wb["Trade Sheet"]
    hdr = [c.value for c in ws[1]]; idx = {h:i for i,h in enumerate(hdr) if h}
    total=0; parity_ok=0; floor_bit=0; single_would_differ=0
    examples=[]
    for r in range(2, ws.max_row+1):
        nm1 = ws.cell(row=r, column=idx.get('Net MAE 1',0)+1).value
        nm2 = ws.cell(row=r, column=idx.get('Final MAE',0)+1).value and ws.cell(row=r, column=idx.get('Net MAE 2',0)+1).value
        fm  = ws.cell(row=r, column=idx.get('Final MAE',0)+1).value
        pct = ws.cell(row=r, column=idx.get('% P&L',0)+1).value
        if nm1 in (None,'') or nm2 in (None,'') or fm in (None,'') or pct in (None,''): continue
        total+=1
        exp = round(min(nm1,nm2,pct),4) if multi else round(min(nm1,nm2),4)
        if abs(fm-exp) < 1e-6: parity_ok+=1
        single_val = round(min(nm1,nm2),4)
        if multi and round(fm,4)==round(pct,4) and pct < single_val - 1e-9:
            floor_bit+=1
            if len(examples)<5: examples.append((r,nm1,nm2,pct,fm,single_val))
    print(f"=== {tag} (multi={multi}) ===")
    print(f"  total checked: {total}, parity(Final==expected): {parity_ok}/{total}")
    if multi:
        print(f"  rows where pct floor BIT (Final=pct < min(nm1,nm2)): {floor_bit}")
        for e in examples:
            print(f"    row {e[0]}: nm1={e[1]} nm2={e[2]} pct={e[3]:.4f} -> Final={e[4]} (was min(nm1,nm2)={e[5]})")

run([
 {"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
 {"segment":"OPTIONS","option_type":"PE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
], "2LEG_CE_PE_SELL", True)

run([
 {"segment":"OPTIONS","option_type":"CE","position":"SELL","lots":1,"expiry":"WEEKLY","strike_interval":50,"strike_selection":{"type":"STRIKE_TYPE","strike_type":"ATM"}},
], "1LEG_CE_SELL", False)
