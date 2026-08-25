#!/usr/bin/env python3
"""
update_ohlc_xlsx.py — fetch missing dates from NSE archives and append to the OHLC xlsx.

Reads the xlsx from SMB, finds the last date per sheet, fetches NSE archive closes
for any missing trading days, appends rows, and writes back to SMB.

Usage:
    python3 scripts/update_ohlc_xlsx.py
"""
from __future__ import annotations

import csv
import datetime
import io
import json
import logging
import subprocess
import sys
import tempfile
import urllib.request
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("update_ohlc_xlsx")

SMB_XLSX = "smb://192.168.4.50/share/Yash Darji/DATA OF OHLC - Copy.xlsx"

# Sheet → (NSE Index Name, xlsx Ticker label)
SHEET_MAP = {
    "Nifty_D":      ("Nifty 50",                "NIFTY 50"),
    "BNF_D":        ("Nifty Bank",              "NIFTYBANK"),
    "Mid_Select_D": ("Nifty Midcap Select",     "NIFTYMIDCAPSELECT"),
    "Mid_D":        ("NIFTY Midcap 100",        "NIFTYMIDCAP100"),
}


def _nse_fetch(date: datetime.date) -> dict[str, float]:
    """Fetch all index closes for a date from NSE archives. Returns {IndexName: close}."""
    url = f"https://archives.nseindia.com/content/indices/ind_close_all_{date:%d%m%Y}.csv"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=15) as r:
        text = r.read().decode()
    return {
        row["Index Name"].strip(): float(row["Closing Index Value"].replace(",", ""))
        for row in csv.DictReader(io.StringIO(text))
        if row.get("Closing Index Value", "").strip()
    }


def _last_date_in_sheet(ws) -> datetime.date | None:
    last = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] is not None:
            v = row[1]
            if isinstance(v, datetime.datetime):
                last = v.date()
            elif isinstance(v, datetime.date):
                last = v
    return last


def run() -> None:
    import openpyxl

    # Download xlsx from SMB
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    logger.info("Downloading xlsx from SMB...")
    r = subprocess.run(["gio", "cat", SMB_XLSX], capture_output=True)
    if r.returncode != 0 or not r.stdout:
        logger.error("Could not read xlsx from SMB: %s", r.stderr.decode()[:200])
        sys.exit(1)
    tmp.write_bytes(r.stdout)

    wb = openpyxl.load_workbook(str(tmp), data_only=True)
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)

    # Find overall date range we need to fetch
    sheet_last: dict[str, datetime.date | None] = {}
    for sheet_name in SHEET_MAP:
        if sheet_name in wb.sheetnames:
            sheet_last[sheet_name] = _last_date_in_sheet(wb[sheet_name])
        else:
            sheet_last[sheet_name] = None

    earliest_missing = min(
        (d + datetime.timedelta(days=1) for d in sheet_last.values() if d),
        default=yesterday,
    )

    if earliest_missing > yesterday:
        logger.info("All sheets already up to date (last date: %s)", yesterday)
        wb.close(); tmp.unlink(missing_ok=True); return

    # Fetch NSE archive data for each missing date
    logger.info("Fetching NSE archives from %s to %s...", earliest_missing, yesterday)
    nse_by_date: dict[datetime.date, dict[str, float]] = {}
    d = earliest_missing
    while d <= yesterday:
        try:
            nse_by_date[d] = _nse_fetch(d)
            logger.info("  %s: ok", d)
        except Exception as e:
            logger.debug("  %s: skip (%s)", d, e)
        d += datetime.timedelta(days=1)

    if not nse_by_date:
        logger.info("No new data from NSE archives")
        wb.close(); tmp.unlink(missing_ok=True); return

    # Append rows to each sheet
    rows_added = 0
    for sheet_name, (nse_name, ticker) in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        last = sheet_last[sheet_name]
        for date, closes in sorted(nse_by_date.items()):
            if last and date <= last:
                continue
            if nse_name not in closes:
                continue
            close = closes[nse_name]
            # Columns: Ticker, Date/Time, Open, High, Low, Close, %Chg, Log Return
            # NSE archives only have close — use close for all OHLC
            ws.append([ticker, datetime.datetime(date.year, date.month, date.day),
                        close, close, close, close, None, None])
            rows_added += 1

    if rows_added == 0:
        logger.info("No new rows to append")
        wb.close(); tmp.unlink(missing_ok=True); return

    # Save and write back to SMB
    out = Path(tempfile.mktemp(suffix=".xlsx"))
    wb.save(str(out))
    wb.close()
    logger.info("Appended %d rows total. Writing back to SMB...", rows_added)
    r2 = subprocess.run(["gio", "copy", "-T", str(out), SMB_XLSX], capture_output=True)
    if r2.returncode != 0:
        logger.error("Failed to write xlsx back to SMB: %s", r2.stderr.decode()[:300])
        sys.exit(1)
    logger.info("Done — xlsx updated on SMB.")
    tmp.unlink(missing_ok=True)
    out.unlink(missing_ok=True)


if __name__ == "__main__":
    run()
