#!/usr/bin/env python3
"""Trade Sheet parity check: shadow tradesheet.xlsx vs a live reference xlsx.

Reports, for the "Trade Sheet": which reference columns the shadow produces,
header order alignment, and — for columns present in both — per-cell value match
on the overlapping rows (numeric tolerance). Structural gaps and per-column
value deltas become an explicit punch-list, so the port is measured, not guessed.

Usage: tradesheet_parity.py <shadow.xlsx> <reference.xlsx> [tol]
"""
import sys
import openpyxl


def load_sheet(path, name="Trade Sheet"):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        return None, []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h) if h is not None else "" for h in rows[0]]
    return header, rows[1:]


def as_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    shadow_path, ref_path = sys.argv[1], sys.argv[2]
    tol = float(sys.argv[3]) if len(sys.argv) > 3 else 0.01

    s_hdr, s_rows = load_sheet(shadow_path)
    r_hdr, r_rows = load_sheet(ref_path)
    if s_hdr is None:
        print("SHADOW has no 'Trade Sheet' sheet"); sys.exit(2)

    s_idx = {h: i for i, h in enumerate(s_hdr)}
    present = [h for h in r_hdr if h in s_idx]
    missing = [h for h in r_hdr if h not in s_idx]
    extra = [h for h in s_hdr if h not in r_hdr]

    print(f"Reference cols: {len(r_hdr)} | Shadow cols: {len(s_hdr)}")
    print(f"Rows: reference {len(r_rows)} | shadow {len(s_rows)}")
    print(f"\nPRESENT in both ({len(present)}/{len(r_hdr)}): {present}")
    print(f"\nMISSING from shadow ({len(missing)}): {missing}")
    print(f"\nEXTRA in shadow ({len(extra)}): {extra}")

    n = min(len(r_rows), len(s_rows))
    if n == 0 or not present:
        print("\nNo overlapping rows/columns to value-check.")
        return
    print(f"\nPer-column value match over {n} overlapping rows (tol={tol}):")
    for h in present:
        ri, si = r_hdr.index(h), s_idx[h]
        ok = tot = 0
        for k in range(n):
            rv, sv = r_rows[k][ri], s_rows[k][si]
            rn, sn = as_num(rv), as_num(sv)
            tot += 1
            if rn is not None and sn is not None:
                if abs(rn - sn) <= tol + 1e-9 * max(abs(rn), abs(sn)):
                    ok += 1
            elif str(rv).strip() == str(sv).strip():
                ok += 1
        flag = "OK " if ok == tot else "DIFF"
        print(f"  [{flag}] {h:18s} {ok}/{tot}")


if __name__ == "__main__":
    main()
